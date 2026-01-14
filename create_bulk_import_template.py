import os
import django
import sys

# Set up Django environment
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'portal.settings')
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from backend.models import Empstatus, Position, ExtensionName, Aoa, Section, Division

def create_excel_template():
    # Create a new workbook
    wb = Workbook()
    
    # Create the main data entry sheet
    ws_main = wb.active
    ws_main.title = "Data Entry"
    
    # Create reference sheets
    ws_empstatus = wb.create_sheet(title="Employment Status")
    ws_position = wb.create_sheet(title="Position")
    ws_extension = wb.create_sheet(title="Extension")
    ws_aoa = wb.create_sheet(title="Area of Assignment")
    ws_section = wb.create_sheet(title="Section")
    ws_division = wb.create_sheet(title="Division")
    
    # Set up headers for main sheet - include the required columns plus the new ones
    headers = ["ID NUMBER", "LAST NAME", "FIRST NAME", "MIDDLE NAME", "EXTENSION", 
               "SEX", "USERNAME", "EMPLOYMENT STATUS", "POSITION", "SALARY RATE", 
               "SALARY GRADE", "STEP INCREMENT", "AREA OF ASSIGNMENT", "SECTION", "DIVISION"]
    
    # Apply formatting to headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws_main.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Set up data validation for SEX column (F)
    sex_validation = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
    ws_main.add_data_validation(sex_validation)
    sex_validation.add('F2:F1000')  # Apply to SEX column
    
    # Set up Employment Status reference sheet
    ws_empstatus.append(["ID", "Name", "Acronym"])
    for cell in ws_empstatus[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all employment statuses
    empstatuses = Empstatus.objects.filter(status=1).order_by('name')
    
    for row_num, status in enumerate(empstatuses, 2):
        ws_empstatus.append([status.id, status.name, status.acronym])
    
    # Set up Position reference sheet
    ws_position.append(["ID", "Name", "Acronym"])
    for cell in ws_position[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all positions
    positions = Position.objects.filter(status=1).order_by('name')
    
    for row_num, position in enumerate(positions, 2):
        ws_position.append([position.id, position.name, position.acronym])
    
    # Set up Extension reference sheet
    ws_extension.append(["ID", "Name"])
    for cell in ws_extension[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all extensions
    extensions = ExtensionName.objects.filter(status=1).order_by('name')
    
    for row_num, extension in enumerate(extensions, 2):
        ws_extension.append([extension.id, extension.name])
        
    # Set up Area of Assignment reference sheet
    ws_aoa.append(["ID", "Name"])
    for cell in ws_aoa[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all areas of assignment
    aoas = Aoa.objects.filter(status=1).order_by('name')
    
    for row_num, aoa in enumerate(aoas, 2):
        ws_aoa.append([aoa.id, aoa.name])
        
    # Set up Section reference sheet
    ws_section.append(["ID", "Name"])
    for cell in ws_section[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all sections
    sections = Section.objects.all().order_by('sec_name')
    
    for row_num, section in enumerate(sections, 2):
        ws_section.append([section.id, section.sec_name])
        
    # Set up Division reference sheet
    ws_division.append(["ID", "Name"])
    for cell in ws_division[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all divisions
    divisions = Division.objects.all().order_by('div_name')
    
    for row_num, division in enumerate(divisions, 2):
        ws_division.append([division.id, division.div_name])
    
    # Set up data validation for EMPLOYMENT STATUS column (H)
    empstatus_validation = DataValidation(type="list", formula1=f'=\'Employment Status\'!$C$2:$C${len(empstatuses)+1}', allow_blank=True)
    ws_main.add_data_validation(empstatus_validation)
    empstatus_validation.add('H2:H1000')  # Apply to EMPLOYMENT STATUS column
    
    # Set up data validation for POSITION column (I)
    position_validation = DataValidation(type="list", formula1=f'=\'Position\'!$B$2:$B${len(positions)+1}', allow_blank=True)
    ws_main.add_data_validation(position_validation)
    position_validation.add('I2:I1000')  # Apply to POSITION column
    
    # Set up data validation for EXTENSION column (E)
    extension_validation = DataValidation(type="list", formula1=f'=\'Extension\'!$B$2:$B${len(extensions)+1}', allow_blank=True)
    ws_main.add_data_validation(extension_validation)
    extension_validation.add('E2:E1000')  # Apply to EXTENSION column
    
    # Set up data validation for AREA OF ASSIGNMENT column (M)
    aoa_validation = DataValidation(type="list", formula1=f'=\'Area of Assignment\'!$B$2:$B${len(aoas)+1}', allow_blank=True)
    ws_main.add_data_validation(aoa_validation)
    aoa_validation.add('M2:M1000')  # Apply to AREA OF ASSIGNMENT column
    
    # Set up data validation for SECTION column (N)
    section_validation = DataValidation(type="list", formula1=f'=\'Section\'!$B$2:$B${len(sections)+1}', allow_blank=True)
    ws_main.add_data_validation(section_validation)
    section_validation.add('N2:N1000')  # Apply to SECTION column
    
    # Set up data validation for DIVISION column (O)
    division_validation = DataValidation(type="list", formula1=f'=\'Division\'!$B$2:$B${len(divisions)+1}', allow_blank=True)
    ws_main.add_data_validation(division_validation)
    division_validation.add('O2:O1000')  # Apply to DIVISION column
    
    # Add instructions to the main sheet
    ws_main.insert_rows(1)
    instructions = ws_main.cell(row=1, column=1, value="Instructions: Use the dropdown lists for EXTENSION, SEX, EMPLOYMENT STATUS, POSITION, AREA OF ASSIGNMENT, SECTION, and DIVISION columns. For reference, see the other sheets.")
    instructions.font = Font(bold=True, color="FF0000")
    ws_main.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    
    # Add example row
    example_row = ["10-2304001", "DELA CRUZ", "JUAN", "SANTOS", "", "M", "jsdcruz", "REG", "Administrative Officer V", "29000", "18", "1", "Field Office", "HRPPMS", "Admin"]
    for col_num, value in enumerate(example_row, 1):
        cell = ws_main.cell(row=2, column=col_num, value=value)
        cell.font = Font(italic=True, color="808080")
    
    # Adjust column widths
    for ws in [ws_empstatus, ws_position, ws_extension, ws_aoa, ws_section, ws_division]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save the workbook
    template_path = 'static/templates/backend/Bulk Import Employees.xlsx'
    wb.save(template_path)
    print(f"Excel template created successfully at {template_path}")

if __name__ == "__main__":
    create_excel_template()
