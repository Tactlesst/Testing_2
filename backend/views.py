import os
import random
import subprocess
import threading
import json
import re
import math

import ldap3
import qrcode
import requests
from django.contrib.staticfiles import finders
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.core.exceptions import EmptyResultSet
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt

from datetime import datetime, date

from dotenv import load_dotenv
from ldap3 import MODIFY_REPLACE, Server, Connection, ALL
from mozilla_django_oidc.utils import is_authenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from django.core.paginator import Paginator
from django.db.models.functions import Concat, Upper, Substr
from django.db.models import Value, Sum, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.contrib import messages

from api.wiserv import send_notification, send_sms_notification
from backend.convocation.models import ConvocationQRCode
from backend.documents.models import DocsIssuancesType, DocsIssuancesFiles
from backend.libraries.gamification.models import GamifyPoints
from backend.libraries.leave.models import LeaveApplication, CTDORequests
from backend.libraries.pas.forms import UploadPictureForm

from frontend.lds.models import LdsRso
from backend.models import Empprofile, Personalinfo, Empstatus, WfhTime, AuthUser, HrwsHealthchecklist, \
    AuthUserUserPermissions, WfhType, DjangoLoggedInUser, Patches, PortalSuccessLogs, PortalErrorLogs, \
    Position, ExtensionName, Aoa, Section, Division
from backend.pas.employee.functions import get_age
from backend.templatetags.tags import get_date_duration_from_now, get_residentialadd, get_section, get_division
from frontend.models import Ritodetails, Feedback, IncaseOfEmergency, PortalAnnouncements, PortalConfiguration, \
    DeskClassification, DeskServices, DeskServicesTransaction, DeskServicesAttachment, \
    Quotes, Civilservice, DeskServicesTransactionAttachment, DeskServicesAdInfo, TsObservationForm, \
    TsObservationFormTotal, PortalShortcutLinks
from landing.models import PositionVacancy

from frontend.templatetags.tags import gamify

from portal import settings, global_variables
from portal.active_directory import searchSamAccountName
from django.utils.timezone import now

from PIL import Image


load_dotenv()


def generate_serial_string(oldstring, prefix=None):
    current_year = datetime.now().year
    current_month = datetime.now().month
    if oldstring:
        oldstring_list = oldstring.split("-")
        if oldstring_list[1] == str(current_year).zfill(4):
            number = int(oldstring_list[3]) + 1
            return "{}-{}-{}-{}".format(str(oldstring_list[0]), str(current_year).zfill(4), str(current_month).zfill(2),
                                        '%04d' % int(number)).strip()
        else:
            return "{}-{}-{}-{}".format(str(oldstring_list[0]), str(current_year).zfill(4), str(current_month).zfill(2),
                                        str("1").zfill(4)).strip()
    else:
        return "{}-{}-{}-{}".format(str(prefix), str(current_year).zfill(4), str(current_month).zfill(2),
                                    str("1").zfill(4)).strip()


class AjaxableResponseMixin:
    def form_invalid(self, form):
        response = super().form_invalid(form)
        if self.request.is_ajax():
            return JsonResponse({'error': True, 'errors': form.errors})
        else:
            return response


def login(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('backend-dashboard')

    if request.method == 'POST':
        data_privacy = request.POST.get('data_privacy_act')

        if not data_privacy:
            return JsonResponse({'msg': 'You must agree to the Data Privacy Act to proceed with the login.', 'type': 'privacy'})

        user = authenticate(username=request.POST.get('username'), password=request.POST.get('password'), request=request)

        if user is None or not user.is_staff or not user.is_active:
            return JsonResponse({'msg': 'Invalid username and password.', 'type': 'auth'})
        auth_login(request, user)
        emp = Empprofile.objects.filter(pi__user_id=user.id).first()
        request.session['user_id'] = user.id
        request.session['username'] = emp.pi.user.username
        request.session['full_name'] = "{}".format(emp.pi.user.get_fullname)

        request.session['pi_id'] = emp.pi_id
        request.session['fname'] = emp.pi.user.first_name

        request.session['emp_id'] = emp.id
        request.session['picture'] = str(emp.picture)

        request.session['div_id'] = emp.section.div.id
        messages.success(request, 'Hi {}! Please secure your ISSSO Account by requesting ticket support '
                                      'through <a href="https://ictsupport.dswd.gov.ph" '
                                      'target="_blank"><strong>ICT Support</strong></a> '
                                      'or email us at ictsupport.focrg@dswd.gov.ph. Thank you!'.format(
                emp.pi.user.first_name), extra_tags='portal_login')

        gamify(1, request.session['emp_id'])

        return JsonResponse({'data': 'Success'})

    return render(request, 'login_new.html')



@login_required
def get_logout_url(request):
    keycloak_redirect_url = settings.OIDC_OP_LOGOUT_ENDPOINT or None
    return keycloak_redirect_url + "?redirect_uri=" + request.build_absolute_uri(settings.LOGOUT_REDIRECT_URL)


@login_required
def logout(request):
    django_logout_url = settings.LOGOUT_REDIRECT_URL or '/'
    if is_authenticated(request.user):
        logout_url = get_logout_url(request)
        request.session.flush()

        auth_logout(request)
        return HttpResponseRedirect('/')



@login_required
def session_expired(request):
    last_activity = request.session.get('last_activity')

    if last_activity:
        elapsed_time = (now() - now().fromisoformat(last_activity)).total_seconds()
        if elapsed_time > settings.INACTIVITY_TIMEOUT:
            request.session['pending_logout'] = True

    return JsonResponse({'pending_logout': request.session.get('pending_logout', False)})

@login_required
def savehealth_checklist(request):
    if request.method == "POST":
        answers = HrwsHealthchecklist.objects.create(
            temperature=request.POST.get('temp'),
            q1=request.POST.get('q1') if request.POST.get('q1') == 'YES' or request.POST.get(
                'q1') == 'NO' else request.POST.get('answer1'),
            q2=request.POST.get('q2') if request.POST.get('q2') == 'YES' or request.POST.get(
                'q2') == 'NO' else request.POST.get('answer2'),
            q3=request.POST.get('q3') if request.POST.get('q3') == 'YES' or request.POST.get(
                'q3') == 'NO' else request.POST.get('answer3'),
            q4=request.POST.get('q4') if request.POST.get('q4') == 'YES' or request.POST.get(
                'q4') == 'NO' else request.POST.get('answer4'),
            q5=request.POST.get('q5') if request.POST.get('q5') == 'YES' or request.POST.get(
                'q5') == 'NO' else request.POST.get('answer5'),
            q6=request.POST.get('q6') if request.POST.get('q6') == 'YES' or request.POST.get(
                'q6') == 'NO' else request.POST.get('answer6'),
            q7=request.POST.get('q7') if request.POST.get('q7') == 'YES' or request.POST.get(
                'q7') == 'NO' else request.POST.get('answer7'),
            q8=request.POST.get('q8') if request.POST.get('q8') == 'YES' or request.POST.get(
                'q8') == 'NO' else request.POST.get('answer8'),
            q9=request.POST.get('q9') if request.POST.get('q9') == 'YES' or request.POST.get(
                'q9') == 'NO' else request.POST.get('answer9'),
            q10=request.POST.get('q10') if request.POST.get('q10') == 'YES' or request.POST.get(
                'q10') == 'NO' else request.POST.get('answer10'),
            q11=request.POST.get('q11') if request.POST.get('q11') == 'YES' or request.POST.get(
                'q11') == 'NO' else request.POST.get('answer11'),
            q12=request.POST.get('q12') if request.POST.get('q12') == 'YES' or request.POST.get(
                'q12') == 'NO' else request.POST.get('answer12'),
            q13=request.POST.get('q13') if request.POST.get('q13') == 'YES' or request.POST.get(
                'q13') == 'NO' else request.POST.get('answer13'),
            q14=request.POST.get('q14') if request.POST.get('q14') == 'YES' or request.POST.get(
                'q14') == 'NO' else request.POST.get('answer14'),
            choose=request.POST.get('choose'),
            emp_id=request.session['emp_id']
        )
        # Save points for RITO
        gamify(17, request.session['emp_id'])

        return redirect('backend-dashboard')


@login_required
def dashboard(request):
    try:
        if request.user.is_staff:
            request.session['username'] = request.user.username
            emp = Empprofile.objects.filter(pi__user_id=request.user.id).first()
            if emp:
                request.session['full_name'] = "{}".format(emp.pi.user.get_fullname)
                request.session['pi_id'] = emp.pi_id
                request.session['fname'] = emp.pi.user.first_name
                request.session['emp_id'] = emp.id
                request.session['picture'] = str(emp.picture)

                request.session.modified = True

                # Save points for daily login activity
                gamify(1, request.session['emp_id'])
            context = {
                'incase_of_emergency': IncaseOfEmergency.objects.filter(pi_id=emp.pi_id).first(),
                'employee': Empprofile.objects.filter(id=emp.id).first(),
                'tab_title': 'Dashboard',
                'title': 'dashboard',
            }
            return render(request, 'dashboard.html', context)
    except EmptyResultSet as e:
        DjangoLoggedInUser.objects.filter(user_id=request.user.id).update(session_key=None)


@login_required
def main_dashboard(request):
    today = datetime.today()

    user = AuthUser.objects.filter(id=request.user.id).first()
    personal_info = Personalinfo.objects.filter(id=request.session['pi_id']).first()
    employee = Empprofile.objects.filter(id=request.session['emp_id']).first()

    check = User.objects.get(id=request.user.id)

    # get announcements and display in dashboard
    announcements = PortalAnnouncements.objects \
        .filter(is_active=True, announcement_type=None).order_by('is_active', 'is_urgent', 'datetime')
    for announcement in announcements:
        get_date_duration_from_now(announcement.datetime, announcement.id, 10)
    announcements = PortalAnnouncements.objects \
        .filter(is_active=True, announcement_type=None).order_by('is_active', 'is_urgent', 'datetime')

    # get links and display in widget
    links = PortalAnnouncements.objects \
        .filter(Q(is_active=True), ~Q(announcement_type=None)).order_by('is_active', 'is_urgent', 'datetime')

    # get details if it is user's birthday today
    if personal_info.dob:
        if today.strftime('%m-%d') == personal_info.dob.strftime('%m-%d'):
            my_bday = True
        else:
            my_bday = False
    else:
        my_bday = False

    # display modal if in case of emergency details are still not provided
    incase_of_emergency = IncaseOfEmergency.objects.filter(pi_id=request.session['pi_id']).first()

    update = False
    if incase_of_emergency is None or personal_info.mobile_no is None or personal_info.mobile_no == '' or \
            user.email is None or user.email == '':
        update = True

    check_password = False
    if check.check_password(employee.id_number):
        check_password = True

    config = PortalConfiguration.objects.filter(key_name='Health Checklist', key_acronym=1).first()
    checklist = False

    if config:
        now = today.strftime('%Y-%m-%d')
        checkif_weekdays = HrwsHealthchecklist.objects.filter(emp_id=request.session['emp_id']).last()

        if checkif_weekdays:
            check_date = today.strptime(str(checkif_weekdays.datetime_added), '%Y-%m-%d %H:%M:%S')
            check_date2 = check_date.strftime('%Y-%m-%d')
            if now == check_date2:
                checklist = False
            else:
                checklist = True
        else:
            checklist = True

    vquotes = Quotes.objects.filter(type_id=2).values_list('quotes', flat=True)

    vacancy = PositionVacancy.objects.filter(
        Q(status=1),
        Q(deadline__gte=date.today()) |
        Q(deadline=None)
    ).order_by('position__name')


    approved_trainings_qs = (
        LdsRso.objects.select_related('training', 'created_by', 'created_by__pi', 'created_by__pi__user')
        .filter(rrso_status=1, rso_status=1)
        .order_by('-date_approved', '-date_added')
    )

    trainings_paginator = Paginator(approved_trainings_qs, 5)
    trainings_page_number = request.GET.get('trainings_page')
    approved_trainings = trainings_paginator.get_page(trainings_page_number)


    context = {
        'vacancy': vacancy,
        'approved_trainings': approved_trainings,
        'announcements': announcements,
        'birthday_celebrants': Paginator(
            Empprofile.objects.filter(pi__dob__month=today.month, pi__dob__day__gte=today.day,
                                      pi__user__is_active=1).order_by(
                'pi__dob__month', 'pi__dob__day', 'pi__user__last_name'),
            18).page(1),
        'links': links,
        'my_bday': my_bday,
        'qrcode': ConvocationQRCode.objects.filter(emp_id=request.session['emp_id']).first(),
        'vquotes': vquotes[random.randint(0, len(vquotes) - 1)],
        'eligibility': Civilservice.objects.filter(Q(el__el_name__icontains='1080') & Q(pi_id=request.session['pi_id']),
                                                   Q(course_id=None)),
        'update': update,
        'month': today,
        'today': date.today(),
        'version_latest': Patches.objects.order_by('-release_date').first(),
        'version': Paginator(Patches.objects.order_by('-release_date'), 3).page(1),
        'checklist': checklist,
        'check_password': check_password
    }
    return render(request, 'dashboard/main.html', context)


@login_required
def vacancies(request):
    vacancy = PositionVacancy.objects.filter(
        Q(status=1),
        Q(deadline__gte=date.today()) |
        Q(deadline=None)
    ).order_by('position__name')

    context = {
        'vacancy': vacancy,
        'today': date.today(),
    }
    return render(request, 'dashboard/vacancies.html', context)


@login_required
def issuances_dashboard(request):
    context = {
        'issuances_type': DocsIssuancesType.objects.filter(status=1).order_by('name'),
    }
    return render(request, 'dashboard/issuances.html', context)


@login_required
def leaderboards_dashboard(request):
    data_top = GamifyPoints.objects.filter(emp__pi__user__is_superuser=False).values('emp_id') \
                   .annotate(all_points=Sum('activity__points')).order_by('-all_points')[:3]
    data_bottom = GamifyPoints.objects.filter(emp__pi__user__is_superuser=False).values('emp_id') \
                      .annotate(all_points=Sum('activity__points')).order_by('-all_points')[4:11]
    context = {
        'data_top': data_top,
        'data_bottom': data_bottom,
    }
    return render(request, 'dashboard/leaderboards.html', context)


@login_required
def issuances_dashboard_view(request, type_id):
    context = {
        'type': DocsIssuancesType.objects.filter(id=type_id).first(),
        'files': DocsIssuancesFiles.objects.filter(issuances_type_id=type_id).order_by('-year'),
        'type_id': type_id
    }
    return render(request, 'dashboard/issuances_view.html', context)


@login_required
def statistics_dashboard(request):
    context = {
        'male': Empprofile.objects.filter(Q(pi__gender=1) & ~Q(position__name='OJT')).count(),
        'female': Empprofile.objects.filter(Q(pi__gender=2) & ~Q(position__name='OJT')).count(),
        'total_employee': Empprofile.objects.filter(~Q(position__name='OJT')).count(),
        'total_active_employee': Empprofile.objects.filter(Q(pi__user__is_active=1) & ~Q(position__name='OJT')).count(),
        'total_inactive_employee': Empprofile.objects.filter(
            Q(pi__user__is_active=0) & ~Q(position__name='OJT')).count(),
        'helpdesk_pending': DeskServices.objects.filter(latest_status=0).count(),
        'helpdesk_resolved': DeskServices.objects.filter(latest_status=1).count(),
        'helpdesk_for_referral': DeskServices.objects.filter(latest_status=4).count(),
        'helpdesk_declined': DeskServices.objects.filter(latest_status=3).count(),
        'empstatus': Empstatus.objects.all(),
        'colors': ['#ed5565', '#00bfa5', '#f8ac59'],
    }
    return render(request, 'dashboard/statistics.html', context)


@login_required
def ticket_request(request):
    if request.method == "POST":
        with transaction.atomic():
            lasttrack = DeskServices.objects.order_by('-id').first()
            track_num = generate_serial_string(lasttrack.tracking_no) if lasttrack else \
                generate_serial_string(None, 'ER')

            emp = ""
            others = ""
            if request.POST.get('in_behalf_of'):
                if '[' in request.POST.get('in_behalf_of'):
                    id_number = re.split('\[|\]', request.POST.get('in_behalf_of'))
                    emp = Empprofile.objects.values('id').filter(id_number=id_number[1])
                    emp = emp.first()['id']
                else:
                    emp = None
                    others = request.POST.get('in_behalf_of')

            services = DeskServices(
                tracking_no=track_num,
                classification_id=request.POST.get('services'),
                purpose=request.POST.get('purpose'),
                description=request.POST.get('description'),
                requested_by_id=request.session['emp_id'] if not request.POST.get('in_behalf_of') else emp,
                year=request.POST.get('year') if request.POST.get('year') else None,
                semester=request.POST.get('semester') if request.POST.get('semester') else None,
                latest_status=0,
                others=others
            )
            services.save()

            DeskServicesTransaction.objects.create(
                services_id=services.id,
                status=0,
                emp_id=request.session['emp_id'] if not request.POST.get('in_behalf_of') else emp
            )

            attachment = request.FILES.getlist('attachment')
            if attachment:
                for row in attachment:
                    DeskServicesAttachment.objects.create(
                        services_id=services.id,
                        file=row
                    )

            return JsonResponse({'data': 'success',
                                 'msg': 'New request has been created. Please copy the generated tracking number for '
                                        'your reference',
                                 'tracking_no': track_num})
    context = {
        'title': 'profile',
        'sub_title': 'cert_requests',
        'tab_parent': 'Employee Data',
        'tab_title': 'Employee Requests',
        'classifications': DeskClassification.objects.filter(status=1).order_by('name')
    }
    return render(request, 'frontend/ticket_request/ticket_request.html', context)


@login_required
@csrf_exempt
def delete_ticket_request(request):
    if request.method == "POST":
        ticket = DeskServices.objects.filter(id=request.POST.get('id'))

        tracking_no = ticket.first().tracking_no
        ticket.delete()

        return JsonResponse({'data': 'success',
                             'msg': 'Ticket request with tracking no. {} have been successfully deleted'.format(
                                 tracking_no)})


@login_required
def assigned_employee(request):
    if request.method == "POST":
        id_number = re.split('\[|\]', request.POST.get('assigned_to'))
        emp = Empprofile.objects.filter(id_number=id_number[1]).first()
        services = DeskServices.objects.filter(id=request.POST.get('services_id'))
        if emp.pi.mobile_no:
            message = "Good day, {}! A HelpDesk request with tracking number {} has been assigned to you. - The " \
                      "Caraga PORTAL Team" \
                .format(emp.pi.user.first_name, services.first().tracking_no)
            t = threading.Thread(target=send_sms_notification,
                                 args=(message, emp.pi.mobile_no, request.session['emp_id'], emp.id))
            t.start()

        services.update(
            assigned_emp_id=emp.id
        )
        return JsonResponse({'data': 'success', 'msg': 'Successfully assigned a personnel.'})


@login_required
def get_help_desk_total(request):
    total_all = DeskServices.objects.all().count()
    total_all_assigned_to_me = DeskServices.objects.filter(assigned_emp_id=request.session['emp_id']).count()
    total_request = DeskServices.objects.filter(requested_by_id=request.session['emp_id']).count()

    total_pending_all = DeskServices.objects.filter(latest_status=0).count()
    total_resolved_all = DeskServices.objects.filter(latest_status=1).count()
    total_referred_all = DeskServices.objects.filter(latest_status=4).count()
    total_declined_all = DeskServices.objects.filter(latest_status=3).count()
    return JsonResponse({'total_all': total_all, 'total_all_assigned_to_me': total_all_assigned_to_me,
                         'total_request': total_request, 'total_pending_all': total_pending_all,
                         'total_resolved_all': total_resolved_all, 'total_referred_all': total_referred_all,
                         'total_declined_all': total_declined_all})


@login_required
def view_help_desk(request, pk):
    if request.method == "POST":
        with transaction.atomic():
            help_desk = DeskServices.objects.filter(id=pk).first()

            if request.POST.get('status') == 4:
                id_number = re.split('\[|\]', request.POST.get('refers_to'))
                emp = Empprofile.objects.values('id').filter(id_number=id_number[1]).first()
                emp = emp['id']

                DeskServices.objects.filter(id=pk).update(
                    assigned_emp_id=emp
                )

                services_transaction = DeskServicesTransaction(
                    services_id=pk,
                    status=request.POST.get('status'),
                    remarks=request.POST.get('action_taken'),
                    emp_id=request.session['emp_id']
                )
                services_transaction.save()

                send_notification(
                    "Good day {}! A HelpDesk request with tracking number {} has been referred to you. - The Caraga "
                    "PORTAL Team".format(
                        emp.pi.user.first_name, help_desk.tracking_no),
                    emp.pi.mobile_no,
                    request.session['emp_id'],
                    emp.id
                )

                DeskServices.objects.filter(id=pk).update(latest_status=4)
            else:
                if request.POST.get('status') == "1":
                    if help_desk.requested_by_id:
                        message = """Good day {}! Your request for {} has been resolved. Please fill in the client satisfaction measurement form. Thank you. - The My PORTAL Team""".format(
                            help_desk.requested_by.pi.user.first_name, help_desk.classification.name)
                        t = threading.Thread(target=send_sms_notification,
                                             args=(
                                             message, help_desk.requested_by.pi.mobile_no, request.session['emp_id'],
                                             request.session['emp_id']))
                        t.start()

                    remarks = request.POST.get(
                        'action_taken')

                    DeskServices.objects.filter(id=pk).update(latest_status=1)
                elif request.POST.get('status') == "3":
                    if help_desk.requested_by_id:
                        message = """Good day {}! We regret to inform you that your request for {} has been declined. If you
                            think there has been an error, please write a message on your HelpDesk request. - The Caraga
                            PORTAL Team""".format(help_desk.requested_by.pi.user.first_name,
                                                  help_desk.classification.name)
                        t = threading.Thread(target=send_sms_notification,
                                             args=(
                                             message, help_desk.requested_by.pi.mobile_no, request.session['emp_id'],
                                             request.session['emp_id']))
                        t.start()

                    DeskServices.objects.filter(id=pk).update(latest_status=3)
                    remarks = request.POST.get('action_taken')

                services_transaction = DeskServicesTransaction(
                    services_id=pk,
                    status=request.POST.get('status'),
                    remarks=remarks,
                    emp_id=request.session['emp_id']
                )

                services_transaction.save()

            attachment = request.FILES.getlist('attachment')

            for row in attachment:
                DeskServicesTransactionAttachment.objects.create(
                    transaction_id=services_transaction.id,
                    file=row
                )

        return JsonResponse({'data': 'success', 'msg': 'Selected action for request has been executed.'})
    context = {
        'total_message': DeskServicesAdInfo.objects.filter(services_id=pk, is_read=0).count(),
        'services': DeskServices.objects.filter(id=pk).first()
    }
    return render(request, 'frontend/ticket_request/view_request.html', context)


@login_required
def update_request(request, pk):
    if request.method == "POST":
        DeskServicesTransaction.objects.filter(id=pk).update(
            status=request.POST.get('status'),
            remarks=request.POST.get('action_taken')
        )

        if request.POST.get('do_not_delete') == '1':
            DeskServicesTransactionAttachment.objects.filter(transaction_id=pk).delete()

        attachment = request.FILES.getlist('attachment')

        for row in attachment:
            DeskServicesTransactionAttachment.objects.create(
                transaction_id=pk,
                file=row
            )

        return JsonResponse({'data': 'success', 'msg': 'You have successfully updated transaction status.'})
    context = {
        'transaction': DeskServicesTransaction.objects.filter(id=pk).first()
    }
    return render(request, 'frontend/ticket_request/update_request.html', context)


@login_required
def print_drf(request, pk):
    services = DeskServices.objects.filter(id=pk).first()
    control_number = services.tracking_no.replace("HD-20", "")
    context = {
        'control_number': control_number,
        'services': services,
        'primary_drf_type': [19, 18, 17, 16, 15],
        'tsof': TsObservationForm.objects.filter(classification_id=services.classification_id),
        'total_tsof': TsObservationFormTotal.objects.filter(classification_id=services.classification_id).first()
    }
    return render(request, 'frontend/ticket_request/print_drf.html', context)


@login_required
def load_ticket_message(request, pk):
    check = DeskServicesAdInfo.objects.filter(services_id=pk).order_by('-date_sent').first()

    if check:
        if check.emp_id != request.session['emp_id']:
            DeskServicesAdInfo.objects.filter(services_id=pk).update(is_read=1)
    context = {
        'services': DeskServices.objects.filter(id=pk).first()
    }
    return render(request, 'frontend/ticket_request/message.html', context)


@login_required
def add_additional_info(request, services_id):
    DeskServicesAdInfo.objects.create(
        services_id=services_id,
        description=request.POST.get('message'),
        emp_id=request.session['emp_id'],
        is_read=0
    )

    help_desk = DeskServices.objects.filter(id=services_id).first()
    send_notification(
        "{} sent a message to the Help Desk with tracking number: {}. - The My PORTAL Team".format(
            help_desk.requested_by.pi.user.get_fullname, help_desk.tracking_no),
        help_desk.requested_by.pi.mobile_no,
        request.session['emp_id'],
        help_desk.requested_by_id
    )
    return JsonResponse({'data': 'success', 'msg': 'Message sent!'})


@login_required
@csrf_exempt
def delete_additional_info(request):
    DeskServicesAdInfo.objects.filter(id=request.POST.get('id')).delete()
    return JsonResponse({'data': 'success', 'msg': 'Message deleted!'})


@login_required
def save_information(request):
    if request.method == "POST":
        Personalinfo.objects.filter(id=request.session['pi_id']).update(
            mobile_no=request.POST.get('mobile_no')
        )

        AuthUser.objects.filter(id=request.user.id).update(
            email=request.POST.get('email')
        )

        check = IncaseOfEmergency.objects.filter(pi_id=request.session['pi_id'])
        if check:
            check.update(
                contact_name=request.POST.get('contact_name'),
                contact_number=request.POST.get('contact_number'),
                is_others=1
            )
        else:
            IncaseOfEmergency.objects.create(
                pi_id=request.session['pi_id'],
                contact_name=request.POST.get('contact_name'),
                contact_number=request.POST.get('contact_number'),
                is_others=1
            )
        return redirect('backend-dashboard')


@login_required
def birthday_celebrants(request):
    today = datetime.now()
    page = request.GET.get('page', 1)

    data = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
            7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}

    context = {
        'birthday_celebrants': Paginator(
            Empprofile.objects.filter(pi__dob__month=today.month, pi__user__is_active=1)
            .order_by('pi__dob__month', 'pi__dob__day', 'pi__user__last_name'),
            25).page(page),
        'month': today,
        'month_num': today.month,
        'list_month': sorted(data.items()),
        'tab_title': 'Birthday Celebrants',
    }
    return render(request, 'birthday_celebrants.html', context)


@login_required
def birthday_celebrants_view(request, pk):
    page = request.GET.get('page', 1)
    context = {
        'month_num': pk,
        'birthday_celebrants': Paginator(
            Empprofile.objects.filter(pi__dob__month=pk, pi__user__is_active=1)
            .order_by('pi__dob__month', 'pi__dob__day', 'pi__user__last_name'),
            25).page(page),
        'tab_title': 'Birthday Celebrants'
    }
    return render(request, 'birthday_celebrants_layout.html', context)


@login_required
def monthly_birthdays(request):
    data = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
            7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}

    context = {
        'tab_title': 'Monthly Birthdays',
        'tab_parent': 'Employee Data',
        'title': 'employee',
        'sub_title': 'monthly_birthday',
        'month': datetime.now().month,
        'list_month': sorted(data.items()),
    }
    return render(request, 'backend/employee_data/birthdays/monthly_bdays.html', context)


@login_required
def print_birthdays(request, month):
    months_list = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
                   7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}

    employees = Empprofile.objects.filter(pi__dob__month=month, pi__user__is_active=1)\
        .order_by('pi__dob__day', 'pi__user__first_name')
    context = {
        'employees': employees,
        'month': months_list[month],
        'pagination': math.ceil(float(employees.count()) / 30),
    }
    return render(request, 'backend/employee_data/birthdays/print_bdays.html', context)


@login_required
def filter_shortcut_links(request):
    if request.method == "GET":
        query = request.GET.get('q', '')
        shortcut_links = PortalShortcutLinks.objects.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        ).order_by('name')[:5]

        data = [
            {
                'name': link.name.upper(),
                'description': link.description,
                'link': link.link
            }
            for link in shortcut_links
        ]

        return JsonResponse(data, safe=False)


@login_required
def filter_employee(request):
    if request.method == "GET":
        json = []
        employee = Empprofile.objects.filter(Q(Q(pi__user__first_name__icontains=request.GET.get('query', '')) |
                                               Q(pi__user__last_name__icontains=request.GET.get('query', ''))) & Q(
            pi__user__is_active=1)).order_by('pi__user__first_name')[:10]
        if employee:
            for row in employee:
                json.append("[{}] {}".format(row.id_number, row.pi.user.get_fullname.upper()))

            return JsonResponse(json, safe=False)
        else:
            return JsonResponse(json, safe=False)


@login_required
def filter_employee_by_division(request):
    if request.method == "GET":
        json = []
        me = Empprofile.objects.values('section__div_id').filter(id=request.session['emp_id']).first()
        employee = Empprofile.objects.filter(Q(Q(pi__user__first_name__icontains=request.GET.get('query', '')) |
                                               Q(pi__user__last_name__icontains=request.GET.get('query', ''))) &
                                             Q(section__div_id=me['section__div_id'])).order_by('pi__user__first_name')[
                   :10]
        if employee:
            for row in employee:
                json.append("[{}] {}".format(row.id_number, row.pi.user.get_fullname.upper()))

            return JsonResponse(json, safe=False)
        else:
            return JsonResponse(json, safe=False)


@login_required
def filter_employee_all(request):
    if request.method == "GET":
        json = []
        employee = Empprofile.objects.filter(Q(Q(id_number__icontains=request.GET.get('query', '')) |
                                               Q(pi__user__first_name__icontains=request.GET.get('query', '')) |
                                               Q(pi__user__last_name__icontains=request.GET.get('query',
                                                                                                '')))).order_by(
            'pi__user__first_name')[:10]
        if employee:
            for row in employee:
                json.append("[{}] {}".format(row.id_number, row.pi.user.get_fullname.upper()))

            return JsonResponse(json, safe=False)
        else:
            return JsonResponse(json, safe=False)


@login_required
def filter_leave_employee(request):
    if request.method == "GET":
        json = []
        employee = LeaveApplication.objects.filter(Q(emp__id_number__icontains=request.GET.get('query', '')) |
                                                   Q(tracking_no__icontains=request.GET.get('query', '')) |
                                                   Q(emp__pi__user__first_name__icontains=request.GET.get('query',
                                                                                                          '')) |
                                                   Q(emp__pi__user__last_name__icontains=request.GET.get('query',
                                                                                                         ''))).order_by(
            'emp__pi__user__first_name')[:10]
        if employee:
            for row in employee:
                latest_status = row.get_latest_tracker_status() if row.get_latest_tracker_status() == 'Pending' else ""

                if latest_status == 'Pending':
                    json.append("[{}] {} - [{}] {} ({})".format(
                        row.emp.id_number,
                        row.emp.pi.user.get_fullname.upper(),
                        row.tracking_no,
                        row.leavesubtype.name,
                        row.get_inclusive()
                    ))

            return JsonResponse(json, safe=False)
        else:
            return JsonResponse(json, safe=False)


@login_required
def filter_ctdo_employee(request):
    if request.method == "GET":
        json = []
        employee = CTDORequests.objects.filter(Q(emp__id_number__icontains=request.GET.get('query', '')) |
                                               Q(tracking_no__icontains=request.GET.get('query', '')) |
                                               Q(emp__pi__user__first_name__icontains=request.GET.get('query', '')) |
                                               Q(emp__pi__user__last_name__icontains=request.GET.get('query', ''))
                                               ).order_by('emp__pi__user__first_name')[:10]
        if employee:

            for row in employee:
                latest_status = row.get_latest_tracker_status() if row.get_latest_tracker_status() == 'Pending' else ""

                if latest_status == 'Pending':
                    json.append("[{}] {} - [{}] {} ({})".format(
                        row.emp.id_number,
                        row.emp.pi.user.get_fullname.upper(),
                        row.tracking_no,
                        "CTDO",
                        row.get_inclusive()
                    ))

            return JsonResponse(json, safe=False)
        else:
            return JsonResponse(json, safe=False)


@login_required
def filter_employee_by_permission(request, permission, except_me='false'):
    permissions = AuthUserUserPermissions.objects.filter(permission__codename__in=permission.split(',')) \
        .values_list('user_id', flat=True)
    if 'true' in except_me:
        employee_name = Empprofile.objects.filter(Q(pi__user__is_active=1), Q(pi__user__id__in=permissions),
                                                  ~Q(id=request.session['emp_id'])).annotate(
            fullname=Concat(Value('['), 'id_number', Value('] '), Upper('pi__user__first_name'), Value(' '),
                            Upper('pi__user__last_name'), Value(' '), Upper('pi__ext__name'))
        ).values_list('fullname', flat=True).order_by('pi__user__last_name')
    else:
        employee_name = Empprofile.objects.filter(pi__user__is_active=1, pi__user__id__in=permissions).annotate(
            fullname=Concat(Value('['), 'id_number', Value('] '), Upper('pi__user__first_name'), Value(' '),
                            Upper('pi__user__last_name'), Value(' '), Upper('pi__ext__name'))
        ).values_list('fullname', flat=True).order_by('pi__user__last_name')
    results = list(employee_name)
    data = json.dumps(results)
    return HttpResponse(data, 'application/json')


@login_required
def filter_incumbent(request):
    employee_name = Empprofile.objects.annotate(
        fullname=Concat(Upper('pi__user__first_name'), Value(' '), Upper(Substr('pi__user__middle_name', 1, 1)),
                        Value('. '),
                        Upper('pi__user__last_name'))
    ).values_list('fullname', flat=True).order_by('pi__user__last_name')
    results = list(employee_name)
    data = json.dumps(results)
    return HttpResponse(data, 'application/json')


@login_required
def upload_picture(request, pk):
    obj = get_object_or_404(Empprofile, pk=pk)
    form = UploadPictureForm(request.POST, request.FILES, instance=obj)
    if request.method == "POST":
        if form.is_valid():
            # Save points for update/upload profile photo
            gamify(6, request.session['emp_id'])

            form.save()
    return JsonResponse({'data': 'success', 'msg': 'Your profile image was updated.', 'picture': obj.picture.url})


@login_required
def upload_picture_admin(request, pk):
    obj = get_object_or_404(Empprofile, pk=pk)
    form = UploadPictureForm(request.POST, request.FILES, instance=obj)
    if request.method == "POST":
        if form.is_valid():
            form.save()
    return JsonResponse({'data': 'success', 'msg': 'User profile image was updated.', 'picture': obj.picture.url})


def page_not_found(request, exception, template_name="404.html"):
    return render(request, "404.html")


def testing(request):
    data = [1092, 1093, 1094, 313, 1022, 472, 1338, 1339, 1497, 1557, 1773, 1774, 1858, 1938, 1939,
            1983, 2082, 544, 545, 546, 824]

    context = {
        'details': Ritodetails.objects.filter(id__in=data)
    }
    return render(request, 'testing.html', context)


@login_required
@permission_required('admin.superadmin')
def feedback(request):
    page = request.GET.get('page', 1)
    search = request.GET.get('search', '')
    rows = request.GET.get('rows', 20)

    total_rate = Feedback.objects.aggregate(Sum('rate'))
    star1 = Feedback.objects.filter(rate=1).count()
    star2 = Feedback.objects.filter(rate=2).count()
    star3 = Feedback.objects.filter(rate=3).count()
    star4 = Feedback.objects.filter(rate=4).count()
    star5 = Feedback.objects.filter(rate=5).count()

    total_stars = star1 + star2 + star3 + star4 + star5
    in_stars = [{'star1': star1}, {'star2': star2}, {'star3': star3}, {'star4': star4}, {'star5': star5}]
    i = 0
    store = {}

    for row in range(5):
        i = i + 1
        count = in_stars[row]['star{}'.format(i)]
        percent = (count / total_stars) * 100

        store.update({'star{}'.format(i): "{:.2f}".format(percent)})

    if total_rate['rate__sum'] is not None:
        total_rate = total_rate['rate__sum']
    else:
        total_rate = 0

    context = {
        'feedback': Paginator(Feedback.objects.filter(Q(comment__icontains=search)).order_by('-date'), rows).page(page),
        'total_rate': total_rate,
        'star_percentage': store,
        'sub_title': 'feedback',
        'title': 'others',
        'average': '{:.2f}'.format(total_rate / total_stars),
        'shaded': int('{:.0f}'.format(total_rate / total_stars)),
        'total_stars': total_stars
    }
    return render(request, 'feedback.html', context)


@login_required
@permission_required('admin.superadmin')
def delete_feedback(request, pk):
    Feedback.objects.filter(id=pk).delete()
    messages.success(request, "Feedback successfully deleted..")
    return redirect('feedback')


@login_required
@permission_required('auth.text_blast')
def text_blast(request):
    if request.method == "POST":
        message = request.POST.get('message')
        cellphone_no = request.POST.getlist('cellphone_no[]')

        for number in cellphone_no:
            send_notification(message, number, request.session['emp_id'])

        return JsonResponse({'data': 'success', 'msg': 'Message successfully sent!'})

    context = {
        'title': 'others',
        'sub_title': 'text_blast',
    }
    return render(request, 'backend/text_blast.html', context)


@login_required
@permission_required('auth.dtr_logs')
def dtr_logs(request):
    if request.method == "POST":
        employee = Empprofile.objects.filter(id_number=request.POST.get('id_number')).first()
        if employee:
            WfhTime.objects.create(
                type_id=request.POST.get('type'),
                datetime=request.POST.get('datetime'),
                emp_id=employee.id,
                ip_address="192.168.34.5"
            )

        return JsonResponse({'data': 'success'})
    context = {
        'type': WfhType.objects.all(),
        'title': 'employee',
        'sub_title': 'monitoring',
        'sub_sub_title': 'dtr_logs'
    }
    return render(request, 'dtr/dtr_logs.html', context)


@login_required
@permission_required('auth.dtr_logs')
def update_dtr_logs(request):
    if request.method == "POST":
        WfhTime.objects.filter(id=request.POST.get('id')).update(
            type_id=request.POST.get('type'),
            datetime=request.POST.get('datetime'),
        )

        return JsonResponse({'data': 'success'})


@login_required
@csrf_exempt
@permission_required('auth.dtr_logs')
def delete_dtr_logs(request):
    if request.method == "POST":
        WfhTime.objects.filter(id=request.POST.get('id')).delete()
        return JsonResponse({'data': 'success'})


@login_required
def health_checklist_data(request):
    now = datetime.now()
    page = request.GET.get('page', 1)
    rows = request.GET.get('rows', 25)
    search = request.GET.get('search', '')
    inyear = request.GET.get('search_year', now.year)
    motx = request.GET.get('motx', '')
    data = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
            7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}
    if request.GET.get('search_year') and request.GET.get('motx'):
        check = Paginator(HrwsHealthchecklist.objects.all().order_by('-datetime_added').filter(
            Q(datetime_added__year=inyear) & Q(datetime_added__month=motx)), rows).page(page)
    else:
        check = Paginator(
            HrwsHealthchecklist.objects.all().order_by('-datetime_added')
            .filter(Q(emp__id_number__icontains=search) | Q(emp__pi__user__username__icontains=search) |
                    Q(emp__pi__user__first_name__icontains=search) | Q(emp__pi__user__last_name__icontains=search) |
                    Q(emp__pi__user__middle_name__icontains=search) | Q(temperature__icontains=search)),
            rows).page(page)
    context = {
        'checklists': check,
        'title': 'employee',
        'sub_title': 'monitoring',
        'sub_sub_title': 'checklist',
        'year': inyear,
        'data': sorted(data.items()),
    }
    return render(request, 'backend/health_checklist/health_checklist.html', context)


@csrf_exempt
def export_healthchecklist(request):
    today = datetime.today()
    year = today.strftime("%Y")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Health Checklist Records'
    wsrow = 2
    worksheet['A1'] = 'Email Address'
    worksheet['B1'] = 'Full Name'
    worksheet['C1'] = 'Sex'
    worksheet['D1'] = 'Age'
    worksheet['E1'] = 'Position'
    worksheet['F1'] = 'Office/Division/Section'
    worksheet['G1'] = 'Mobile Number'
    worksheet['H1'] = 'Complete Residential Address'
    worksheet['I1'] = 'Temperature'
    worksheet['J1'] = 'A. Are you experiencing/experienced for the last 24 hours: Sore Throat?'
    worksheet['K1'] = 'B. Are you experiencing/experienced for the last 24 hours: Body Pain?'
    worksheet['L1'] = 'C. Are you experiencing/experienced for the last 24 hours: Headache?'
    worksheet['M1'] = 'D. Are you experiencing/experienced for the last 24 hours: Cough?'
    worksheet['N1'] = 'E. Are you experiencing/experienced for the last 24 hours: Colds?'
    worksheet['O1'] = 'F. Are you experiencing/experienced for the last 24 hours: Fever?'
    worksheet['P1'] = 'G. Are you experiencing/experienced for the last 24 hours: Loss of smell? '
    worksheet['Q1'] = 'H. Are you experiencing/experienced for the last 24 hours: Loss of taste?'
    worksheet['R1'] = 'I. Are you experiencing/experienced for the last 24 hours: Diarrhea?'
    worksheet[
        'S1'] = 'J. Have you worked together or stayed in the same close environment of a confirmed COVID-19 case in ' \
                'the last 24 hours and/or if you are identified as close contact? '
    worksheet[
        'T1'] = 'K. Have you had any contact with anyone with fever, cough, colds, sore throat, body pains, headache, ' \
                'loss of smell and taste and diarrhea in the last 24 hours? '
    worksheet['U1'] = 'L. Have you travelled outside of the Philippines in the last 24 hours?'
    worksheet['V1'] = 'M. Have you travelled outside your household in the last 24 hours?'
    worksheet['W1'] = 'N. Have you visited any hospital or health facility in the last 24 hours?'
    worksheet['X1'] = 'Privacy Policy'
    worksheet['Y1'] = 'Date accomplished'

    worksheet['A1'].font = Font(bold=True)
    worksheet['B1'].font = Font(bold=True)
    worksheet['C1'].font = Font(bold=True)
    worksheet['D1'].font = Font(bold=True)
    worksheet['E1'].font = Font(bold=True)
    worksheet['F1'].font = Font(bold=True)
    worksheet['G1'].font = Font(bold=True)
    worksheet['H1'].font = Font(bold=True)
    worksheet['I1'].font = Font(bold=True)
    worksheet['J1'].font = Font(bold=True)
    worksheet['K1'].font = Font(bold=True)
    worksheet['L1'].font = Font(bold=True)
    worksheet['M1'].font = Font(bold=True)
    worksheet['N1'].font = Font(bold=True)
    worksheet['O1'].font = Font(bold=True)
    worksheet['P1'].font = Font(bold=True)
    worksheet['Q1'].font = Font(bold=True)
    worksheet['R1'].font = Font(bold=True)
    worksheet['S1'].font = Font(bold=True)
    worksheet['T1'].font = Font(bold=True)
    worksheet['U1'].font = Font(bold=True)
    worksheet['V1'].font = Font(bold=True)
    worksheet['W1'].font = Font(bold=True)
    worksheet['X1'].font = Font(bold=True)
    worksheet['Y1'].font = Font(bold=True)

    data_all = True
    if request.POST.get('year'):
        data_all = False
        by_year = HrwsHealthchecklist.objects.all().filter(datetime_added__year=request.POST.get('year')).order_by(
            '-datetime_added').values('temperature',
                                      'datetime_added',
                                      'q1',
                                      'q2',
                                      'q3',
                                      'q4',
                                      'q5',
                                      'q6',
                                      'q7',
                                      'q8',
                                      'q9',
                                      'q10',
                                      'q11',
                                      'q12',
                                      'q13',
                                      'q14',
                                      'choose',
                                      'emp__pi__user__middle_name',
                                      'emp__pi__user__first_name',
                                      'emp__pi__user__last_name',
                                      'emp__pi__ext__name',
                                      'emp__pi__gender',
                                      'emp__pi__dob',
                                      'emp__position__name',
                                      'emp__pi__mobile_no',
                                      'emp__pi_id',
                                      'emp__pi__user__email',
                                      'emp__section_id')
        if by_year:
            for indx, row in enumerate(by_year):
                lastname = row['emp__pi__user__last_name']
                middlename = row['emp__pi__user__middle_name']
                firstname = row['emp__pi__user__first_name']
                ext = row['emp__pi__ext__name']
                worksheet['A' + str(wsrow)] = row['emp__pi__user__email']
                worksheet['B' + str(wsrow)] = "{}, {} {} {}.".format(str(lastname.upper()), str(firstname.upper()),
                                                                     str(middlename.upper()),
                                                                     "" if ext is None else ext)
                worksheet['C' + str(wsrow)] = 'Male' if row['emp__pi__gender'] == 1 else 'Female'
                worksheet['D' + str(wsrow)] = get_age(row['emp__pi__dob'])
                worksheet['E' + str(wsrow)] = row['emp__position__name']
                worksheet['F' + str(wsrow)] = "{}-{}".format(get_division(row['emp__section_id']),
                                                             get_section(row['emp__section_id']))
                worksheet['G' + str(wsrow)] = row['emp__pi__mobile_no']
                worksheet['H' + str(wsrow)] = get_residentialadd(row['emp__pi_id'])
                worksheet['I' + str(wsrow)] = row['temperature']
                worksheet['J' + str(wsrow)] = row['q1']
                worksheet['K' + str(wsrow)] = row['q2']
                worksheet['L' + str(wsrow)] = row['q3']
                worksheet['M' + str(wsrow)] = row['q4']
                worksheet['N' + str(wsrow)] = row['q5']
                worksheet['O' + str(wsrow)] = row['q6']
                worksheet['P' + str(wsrow)] = row['q7']
                worksheet['Q' + str(wsrow)] = row['q8']
                worksheet['R' + str(wsrow)] = row['q9']
                worksheet['S' + str(wsrow)] = row['q10']
                worksheet['T' + str(wsrow)] = row['q11']
                worksheet['U' + str(wsrow)] = row['q12']
                worksheet['V' + str(wsrow)] = row['q13']
                worksheet['W' + str(wsrow)] = row['q14']
                worksheet['X' + str(wsrow)] = 'Agree' if row['choose'] == 1 else 'Disagree'
                worksheet['Y' + str(wsrow)] = row['datetime_added']
                wsrow = wsrow + 1
            response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
            response[
                'Content-Disposition'] = 'attachment; filename=Health Checklist Exported in year' + '(' + request.POST.get(
                'year') + ').xlsx'
            return response
        else:
            response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
            response[
                'Content-Disposition'] = 'attachment; filename=Health Checklist Exported in year' + '(' + request.POST.get(
                'year') + ').xlsx'
            return response

    elif request.POST.getlist('months[]'):
        data_all = False
        by_month = HrwsHealthchecklist.objects.all().filter(
            datetime_added__month__in=request.POST.getlist('months[]')).order_by('-datetime_added').values(
            'temperature',
            'datetime_added',
            'q1',
            'q2',
            'q3',
            'q4',
            'q5',
            'q6',
            'q7',
            'q8',
            'q9',
            'q10',
            'q11',
            'q12',
            'q13',
            'q14',
            'choose',
            'emp__pi__user__middle_name',
            'emp__pi__user__first_name',
            'emp__pi__user__last_name',
            'emp__pi__ext__name',
            'emp__pi__gender',
            'emp__pi__dob',
            'emp__position__name',
            'emp__pi__mobile_no',
            'emp__pi_id',
            'emp__pi__user__email',
            'emp__section_id')
        if by_month:
            for indx, row in enumerate(by_month):
                lastname = row['emp__pi__user__last_name']
                middlename = row['emp__pi__user__middle_name']
                firstname = row['emp__pi__user__first_name']
                ext = row['emp__pi__ext__name']
                worksheet['A' + str(wsrow)] = row['emp__pi__user__email']
                worksheet['B' + str(wsrow)] = "{}, {} {} {}.".format(str(lastname.upper()), str(firstname.upper()),
                                                                     str(middlename.upper()),
                                                                     "" if ext is None else ext)
                worksheet['C' + str(wsrow)] = 'Male' if row['emp__pi__gender'] == 1 else 'Female'
                worksheet['D' + str(wsrow)] = get_age(row['emp__pi__dob'])
                worksheet['E' + str(wsrow)] = row['emp__position__name']
                worksheet['F' + str(wsrow)] = "{}-{}".format(get_division(row['emp__section_id']),
                                                             get_section(row['emp__section_id']))
                worksheet['G' + str(wsrow)] = row['emp__pi__mobile_no']
                worksheet['H' + str(wsrow)] = get_residentialadd(row['emp__pi_id'])
                worksheet['I' + str(wsrow)] = row['temperature']
                worksheet['J' + str(wsrow)] = row['q1']
                worksheet['K' + str(wsrow)] = row['q2']
                worksheet['L' + str(wsrow)] = row['q3']
                worksheet['M' + str(wsrow)] = row['q4']
                worksheet['N' + str(wsrow)] = row['q5']
                worksheet['O' + str(wsrow)] = row['q6']
                worksheet['P' + str(wsrow)] = row['q7']
                worksheet['Q' + str(wsrow)] = row['q8']
                worksheet['R' + str(wsrow)] = row['q9']
                worksheet['S' + str(wsrow)] = row['q10']
                worksheet['T' + str(wsrow)] = row['q11']
                worksheet['U' + str(wsrow)] = row['q12']
                worksheet['V' + str(wsrow)] = row['q13']
                worksheet['W' + str(wsrow)] = row['q14']
                worksheet['X' + str(wsrow)] = 'Agree' if row['choose'] == 1 else 'Disagree'
                worksheet['Y' + str(wsrow)] = row['datetime_added']
                wsrow = wsrow + 1
            response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Health Checklist Exported by month.xlsx'
            return response
        else:
            response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Health Checklist Exported by month.xlsx'
            return response

    elif data_all:
        checklist = HrwsHealthchecklist.objects.all().order_by('-datetime_added').values('temperature',
                                                                                         'datetime_added',
                                                                                         'q1',
                                                                                         'q2',
                                                                                         'q3',
                                                                                         'q4',
                                                                                         'q5',
                                                                                         'q6',
                                                                                         'q7',
                                                                                         'q8',
                                                                                         'q9',
                                                                                         'q10',
                                                                                         'q11',
                                                                                         'q12',
                                                                                         'q13',
                                                                                         'q14',
                                                                                         'choose',
                                                                                         'emp__pi__user__middle_name',
                                                                                         'emp__pi__user__first_name',
                                                                                         'emp__pi__user__last_name',
                                                                                         'emp__pi__ext__name',
                                                                                         'emp__pi__gender',
                                                                                         'emp__pi__dob',
                                                                                         'emp__position__name',
                                                                                         'emp__pi__mobile_no',
                                                                                         'emp__pi_id',
                                                                                         'emp__pi__user__email',
                                                                                         'emp__section_id')
        if checklist:
            for indx, row in enumerate(checklist):
                lastname = row['emp__pi__user__last_name']
                middlename = row['emp__pi__user__middle_name']
                firstname = row['emp__pi__user__first_name']
                ext = row['emp__pi__ext__name']
                worksheet['A' + str(wsrow)] = row['emp__pi__user__email']
                worksheet['B' + str(wsrow)] = "{}, {} {} {}.".format(str(lastname.upper()), str(firstname.upper()),
                                                                     str(middlename.upper()),
                                                                     "" if ext is None else ext)
                worksheet['C' + str(wsrow)] = 'Male' if row['emp__pi__gender'] == 1 else 'Female'
                worksheet['D' + str(wsrow)] = get_age(row['emp__pi__dob'])
                worksheet['E' + str(wsrow)] = row['emp__position__name']
                worksheet['F' + str(wsrow)] = "{}-{}".format(get_division(row['emp__section_id']),
                                                             get_section(row['emp__section_id']))
                worksheet['G' + str(wsrow)] = row['emp__pi__mobile_no']
                worksheet['H' + str(wsrow)] = get_residentialadd(row['emp__pi_id'])
                worksheet['I' + str(wsrow)] = row['temperature']
                worksheet['J' + str(wsrow)] = row['q1']
                worksheet['K' + str(wsrow)] = row['q2']
                worksheet['L' + str(wsrow)] = row['q3']
                worksheet['M' + str(wsrow)] = row['q4']
                worksheet['N' + str(wsrow)] = row['q5']
                worksheet['O' + str(wsrow)] = row['q6']
                worksheet['P' + str(wsrow)] = row['q7']
                worksheet['Q' + str(wsrow)] = row['q8']
                worksheet['R' + str(wsrow)] = row['q9']
                worksheet['S' + str(wsrow)] = row['q10']
                worksheet['T' + str(wsrow)] = row['q11']
                worksheet['U' + str(wsrow)] = row['q12']
                worksheet['V' + str(wsrow)] = row['q13']
                worksheet['W' + str(wsrow)] = row['q14']
                worksheet['X' + str(wsrow)] = 'Agree' if row['choose'] == 1 else 'Disagree'
                worksheet['Y' + str(wsrow)] = row['datetime_added']
                wsrow = wsrow + 1
            response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Health Checklist.xlsx'
            return response


def export_qr(request):
    employee = Empprofile.objects.filter(id_number__in=[
        '16-09082',
        '16-12477',
    ])

    for row in employee:
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )

        incase_of_emergency = IncaseOfEmergency.objects.filter(pi_id=row.pi_id).first()
        if row.section_id and row.id_number and row.position_id and row.empstatus_id and row.pi.mobile_no and row.pi.user.email and incase_of_emergency:
            if row.pi.user.middle_name != '' and row.pi.user.middle_name is not None:
                qr.add_data(
                    "{} {} {}\n".format(row.pi.user.first_name.title(), row.pi.user.middle_name[0:1].upper() + '.',
                                        row.pi.user.last_name.title()))
            else:
                qr.add_data("{} {}".format(row.pi.user.first_name.title(), row.pi.user.last_name.title()))

            qr.add_data(row.id_number + '\n')
            qr.add_data(row.position.name + '\n')
            qr.add_data(row.section.sec_name + '\n')
            qr.add_data(row.empstatus.name + '\n')
            qr.add_data(row.pi.mobile_no + '\n')
            qr.add_data(row.pi.user.email + '\n\n')

            if incase_of_emergency:
                qr.add_data("Contact In Case of Emergency \n")
                qr.add_data(incase_of_emergency.contact_name + '\n')
                qr.add_data(incase_of_emergency.contact_number + '\n\n')

            qr.add_data("{}/profiles/{}".format(os.getenv('SERVER_URL'), row.id_number))
            qr.make(fit=True)
            img = qr.make_image(fill_color="black")
            img.save("qrcode/{}.png".format(row.id_number))
    return HttpResponse(request, "Success")


@login_required
def generate_training_qr(request):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    current_date = datetime.now().strftime("%b %d, %Y %I:%M %p")


    qr.add_data({
        "Employee ID": "10-3456789",
        "date": current_date,
        "Location": "Jollibee",
    })
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white").convert('RGBA')

    logo_path = finders.find('image/dswd.png')
    if logo_path:
        try:
            logo = Image.open(logo_path).convert('RGBA')

            qr_w, qr_h = img.size
            logo_size = int(min(qr_w, qr_h) * 0.22)
            logo = logo.resize((logo_size, logo_size), Image.LANCZOS)

            pos = ((qr_w - logo_size) // 2, (qr_h - logo_size) // 2)
            img.alpha_composite(logo, dest=pos)
        except Exception:
            pass

    out = img.convert('RGB')
    from io import BytesIO
    buf = BytesIO()
    out.save(buf, format='PNG')
    buf.seek(0)
    return HttpResponse(buf.getvalue(), content_type='image/png')


# New Function
def generate_training_qr_stub():
    pass


def generate_bulk_import_template(request):
    """
    Generate an Excel template for bulk importing employees with dropdown menus
    for employment status, position, and extension.
    """
    # Create a new workbook
    wb = Workbook()
    
    # Create the main data entry sheet
    ws_main = wb.active
    ws_main.title = "Data Entry"
    
    # Create reference sheets
    ws_empstatus = wb.create_sheet(title="Employment Status")
    ws_position = wb.create_sheet(title="Position")
    ws_extension = wb.create_sheet(title="Extension")
    ws_aoa = wb.create_sheet(title="Area of Assignment")
    ws_section = wb.create_sheet(title="Section")
    ws_division = wb.create_sheet(title="Division")
    
    # Set up headers for main sheet
    headers = ["ID NUMBER", "LAST NAME", "FIRST NAME", "MIDDLE NAME", "EXTENSION", 
               "SEX", "USERNAME", "EMPLOYMENT STATUS", "POSITION", "SALARY RATE", 
               "SALARY GRADE", "STEP INCREMENT", "AREA OF ASSIGNMENT", "SECTION", "DIVISION"]
    
    # Apply formatting to headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws_main.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Set up data validation for SEX column (F)
    sex_validation = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
    ws_main.add_data_validation(sex_validation)
    sex_validation.add('F2:F1000')  # Apply to SEX column
    
    # Set up Employment Status reference sheet
    ws_empstatus.append(["ID", "Name", "Acronym"])
    for cell in ws_empstatus[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all employment statuses
    empstatuses = Empstatus.objects.filter(status=1).order_by('name')
    
    for row_num, status in enumerate(empstatuses, 2):
        ws_empstatus.append([status.id, status.name, status.acronym])
    
    # Set up Position reference sheet
    ws_position.append(["ID", "Name", "Acronym"])
    for cell in ws_position[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all positions
    positions = Position.objects.filter(status=1).order_by('name')
    
    for row_num, position in enumerate(positions, 2):
        ws_position.append([position.id, position.name, position.acronym])
    
    # Set up Extension reference sheet
    ws_extension.append(["ID", "Name"])
    for cell in ws_extension[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all extensions
    extensions = ExtensionName.objects.filter(status=1).order_by('name')
    
    for row_num, extension in enumerate(extensions, 2):
        ws_extension.append([extension.id, extension.name])
    
    # Set up Area of Assignment reference sheet
    ws_aoa.append(["ID", "Name"])
    for cell in ws_aoa[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all areas of assignment
    aoas = Aoa.objects.filter(status=1).order_by('name')
    
    for row_num, aoa in enumerate(aoas, 2):
        ws_aoa.append([aoa.id, aoa.name])
    
    # Set up Section reference sheet
    ws_section.append(["ID", "Name", "Division"])
    for cell in ws_section[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all sections
    sections = Section.objects.order_by('sec_name')
    
    for row_num, section in enumerate(sections, 2):
        ws_section.append([section.id, section.sec_name, section.div.div_name if section.div else ""])
    
    # Set up Division reference sheet
    ws_division.append(["ID", "Name"])
    for cell in ws_division[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all divisions
    divisions = Division.objects.order_by('div_name')
    
    for row_num, division in enumerate(divisions, 2):
        ws_division.append([division.id, division.div_name])
    
    # Set up data validation for EMPLOYMENT STATUS column (H)
    empstatus_validation = DataValidation(type="list", formula1=f'=\'Employment Status\'!$C$2:$C${len(empstatuses)+1}', allow_blank=True)
    ws_main.add_data_validation(empstatus_validation)
    empstatus_validation.add('H2:H1000')  # Apply to EMPLOYMENT STATUS column
    
    # Set up data validation for POSITION column (I)
    position_validation = DataValidation(type="list", formula1=f'=\'Position\'!$B$2:$B${len(positions)+1}', allow_blank=True)
    ws_main.add_data_validation(position_validation)
    position_validation.add('I2:I1000')  # Apply to POSITION column
    
    # Set up data validation for EXTENSION column (E)
    extension_validation = DataValidation(type="list", formula1=f'=\'Extension\'!$B$2:$B${len(extensions)+1}', allow_blank=True)
    ws_main.add_data_validation(extension_validation)
    extension_validation.add('E2:E1000')  # Apply to EXTENSION column
    
    # Set up data validation for AREA OF ASSIGNMENT column (M)
    aoa_validation = DataValidation(type="list", formula1=f'=\'Area of Assignment\'!$B$2:$B${len(aoas)+1}', allow_blank=True)
    ws_main.add_data_validation(aoa_validation)
    aoa_validation.add('M2:M1000')  # Apply to AREA OF ASSIGNMENT column
    
    # Set up data validation for SECTION column (N)
    section_validation = DataValidation(type="list", formula1=f'=\'Section\'!$B$2:$B${len(sections)+1}', allow_blank=True)
    ws_main.add_data_validation(section_validation)
    section_validation.add('N2:N1000')  # Apply to SECTION column
    
    # Set up data validation for DIVISION column (O)
    division_validation = DataValidation(type="list", formula1=f'=\'Division\'!$B$2:$B${len(divisions)+1}', allow_blank=True)
    ws_main.add_data_validation(division_validation)
    division_validation.add('O2:O1000')  # Apply to DIVISION column
    
    # Add instructions to the main sheet
    ws_main.insert_rows(1)
    instructions = ws_main.cell(row=1, column=1, value="Instructions: Use the dropdown lists for EXTENSION, SEX, EMPLOYMENT STATUS, POSITION, AREA OF ASSIGNMENT, SECTION, and DIVISION columns. For reference, see the other sheets.")
    instructions.font = Font(bold=True, color="FF0000")
    ws_main.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    
    # Add example row
    example_row = ["10-2304001", "DELA CRUZ", "JUAN", "SANTOS", "", "M", "jsdcruz", "REG", "Administrative Officer V", "29000", "18", "1", "Field Office", "HRPPMS", "Admin"]
    for col_num, value in enumerate(example_row, 1):
        cell = ws_main.cell(row=2, column=col_num, value=value)
        cell.font = Font(italic=True, color="808080")
    
    # Adjust column widths
    for ws in [ws_empstatus, ws_position, ws_extension, ws_aoa, ws_section, ws_division]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Create response
    response = HttpResponse(
        content=save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Bulk Import Employees.xlsx"'
    return response


@login_required
@permission_required('admin.superadmin')
def portal_updater(request):
    if request.method == "POST":
        results = subprocess.call(['sh', '/opt/apps/portal/git-updater.sh'])
        return JsonResponse({'data': results})

    return render(request, 'portal_updater.html')


def reset_ad_password(username, id_number):
    user = searchSamAccountName(username)

    if user["status"]:
        enc_pwd = '"{}"'.format(str(id_number)).encode('utf-16-le')
        user["connection"].modify(user["userDN"], {'unicodePwd': [(MODIFY_REPLACE, [enc_pwd])]})

        # Perform a search to retrieve the userAccountControl attribute
        user["connection"].search(search_base=user["userDN"],
                                  search_filter='(objectclass=person)',
                                  attributes=['userAccountControl'])

        # Retrieve the userAccountControl attribute from the response
        uac = user["connection"].response[0]['attributes']['userAccountControl']

        # Set the DONT_EXPIRE_PASSWORD flag
        new_uac = uac | 0x10000  # 0x10000 = 65536 = ADS_UF_DONT_EXPIRE_PASSWD

        # Ensure the LOCKOUT flag is not set
        new_uac &= ~0x10  # 0x10 = 16 = ADS_UF_LOCKOUT

        # Ensure the ACCOUNTDISABLE flag is not set (to enable the account)
        new_uac &= ~0x2  # 0x2 = 2 = ADS_UF_ACCOUNTDISABLE

        # Update userAccountControl
        user["connection"].modify(user["userDN"], {'userAccountControl': [(MODIFY_REPLACE, [new_uac])]})


@login_required
@permission_required('auth.support_team')
def reset_portal_account(request):
    if request.method == "POST":
        try:
            id_number = request.POST.get('id_number').split(',')

            check = Empprofile.objects.filter(id_number__in=[row.strip() for row in id_number])

            for row in check:
                user = User.objects.get(id=row.pi.user_id)
                user.set_password(row.id_number)
                user.save()

                PortalSuccessLogs.objects.create(
                    logs="The password has been successfully reset for the user with this ID number: {} and username: {}.".format(
                        row.id_number, row.pi.user.username),
                    date_created=datetime.now(),
                    emp_id=request.session['emp_id'],
                    type='password_reset'
                )

                t = threading.Thread(target=reset_ad_password, args=(row.pi.user.username, row.id_number))
                t.start()

            return JsonResponse({'data': 'success', 'msg': 'Password updated successfully.'})
        except Exception as e:
            PortalErrorLogs.objects.create(
                logs="My PORTAL Scripts: {}".format(e),
                date_created=datetime.now(),
                emp_id=request.session['emp_id']
            )

    return render(request, 'backend/scripts/password_reset.html')


def test_ad_credential(username, password):
    USER = "{}@ENTDSWD.LOCAL".format(username)
    BASEDN = global_variables.BASE_DN
    s = Server(global_variables.AD_SERVER, get_info=ALL,
               use_ssl=False)

    # Try to connect using the provided username and password
    try:
        c = Connection(s,
                       user=USER,
                       password=password, auto_bind=True)

        SEARCHFILTER = '(&(userPrincipalName=' + USER + ')(objectClass=person))'
        USER_DN = ""
        USER_CN = ""
        c.search(search_base=BASEDN,
                 search_filter=SEARCHFILTER,
                 search_scope=ldap3.SUBTREE,
                 attributes=['cn', 'givenName', 'userPrincipalName'],
                 paged_size=5)

        for entry in c.response:
            if entry.get("dn") and entry.get("attributes"):
                if entry.get("attributes").get("userPrincipalName"):
                    if entry.get("attributes").get("userPrincipalName") == USER:
                        USER_DN = entry.get("dn")
                        USER_CN = entry.get("attributes").get("cn")

        if USER_DN and USER_CN:
            return {"status": True, "connection": c, "userDN": USER_DN}
        else:
            return {"status": False}
    except ldap3.core.exceptions.LDAPBindError:
        return {"status": False}


@login_required
@permission_required('auth.support_team')
def test_user_login(request):
    if request.method == "POST":
        status = test_ad_credential(request.POST.get('username'), request.POST.get('password'))

        msg = ''
        alert = ''
        if status['status']:
            msg = "The user's credentials have been successfully authenticated in both Active Directory and Global Protect."
            alert = 'success'
        else:
            msg = "The user's credentials are unable to authenticate in both Active Directory and Global Protect."
            alert = 'warning'

        return JsonResponse({'data': 'success', 'msg': msg, 'alert': alert})

    return render(request, 'backend/scripts/test_account_login.html')
