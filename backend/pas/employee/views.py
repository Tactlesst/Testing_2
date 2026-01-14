import re
import io
import csv
import json
import random
import string

from django.db import transaction
from django.db.models.functions import Upper, Concat
from django.utils.safestring import mark_safe
from django.utils.timezone import now
from ldap3 import MODIFY_REPLACE
from datetime import datetime, date, timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.hashers import make_password
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.paginator import Paginator

from django.db.models import Value, Q, Max
from django.http import JsonResponse, HttpResponse, Http404
from django.shortcuts import render, get_object_or_404, redirect

from django.urls import reverse_lazy
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import UpdateView
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook

from backend.forms import DivisionsForm, SectionsForm, SchoolForm, DegreeForm, HonorsForm, EligibilityForm, \
    PositionForm, EmpstatusForm, OrganizationForm, TrainingtitleForm, NonacadForm, HobbiesForm
from backend.libraries.pas.forms import UploadPictureForm
from backend.models import Empprofile, AuthUser, Personalinfo, Fundsource, Aoa, Project, Position, Empstatus, \
    ExtensionName, Salarygrade, Stepinc, Division, Section, AuthPermission, Indiginous,Religion,Ethnicity, \
    AuthUserUserPermissions, HrppmsModeaccession, HrppmsModeseparation, DtrPin, PayrollIncharge, PortalErrorLogs
from backend.views import AjaxableResponseMixin, send_notification
from frontend.models import Address, Deductionnumbers, Familybackground, Educationbackground, Civilservice, \
    Workexperience, Voluntary, Skills, Training, Recognition, Membership, Additional, Reference, Children, \
    PdsUpdateTracking, PisConfig, Ritopeople, Workhistory, Civilstatus, Province, Bloodtype, Countries, City, Brgy, \
    Educationlevel, Honors, Degree, School, Eligibility, Organization, Trainingtitle, Trainingtype, Nonacad, Hobbies, \
    IncaseOfEmergency, PortalConfiguration, Course
from portal.active_directory import searchSamAccountName, create_ad_account
from .functions import get_civil_status, get_solo_parent, get_section, get_division, get_aoa, get_age, get_senior, \
    get_residentialadd, get_permanentadd, entry_first_indswd, get_civilservice, get_ra, get_let, getlevelof_elig, \
    geteducation_status, get_masters, get_firstdegree, get_lastdegree, get_otherdegree, filledup, get_bloodtype
from ..payroll.models import PasEmpPayrollIncharge


@login_required
@permission_required("auth.staffing")
def bulk_import_employees(request):
    if request.method == "POST":
        datasource = []
        errorcounter = okcounter = 0
        filename = request.FILES.get('bulk-import')
        if not filename.name.endswith('.xlsx'):
            raise Http404("File you provided is invalid. Please upload an Excel (.xlsx) file.")
        else:
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(filename)
                worksheet = workbook.active
                
                # Skip the first row (header) and the instruction row
                for row in worksheet.iter_rows(min_row=3, values_only=True):
                    # Skip empty rows
                    if all(cell is None or cell == '' for cell in row):
                        continue
                        
                    datasource.append({
                        'id': str(row[0]) if row[0] else '',
                        'ln': str(row[1]) if row[1] else '',
                        'fn': str(row[2]) if row[2] else '',
                        'mn': str(row[3]) if row[3] else '',
                        'ext': str(row[4]) if row[4] else '',
                        'sex': str(row[5]) if row[5] else '',
                        'un': str(row[6]) if row[6] else '',
                        'es': str(row[7]) if row[7] else '',
                        'pos': str(row[8]) if row[8] else '',
                        'sr': str(row[9]) if row[9] else '',
                        'sg': str(row[10]) if row[10] else '',
                        'si': str(row[11]) if row[11] else '',
                        'aoa': str(row[12]) if row[12] else '',
                        'section': str(row[13]) if row[13] else '',
                        'division': str(row[14]) if row[14] else ''
                    })
            except Exception as e:
                raise Http404(e)

            hasatleastoneerror = False
            firsterror = ''
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Bulk Import Errors'
            wsrow = 2
            worksheet['A1'] = 'ID NUMBER'
            worksheet['B1'] = 'LAST NAME'
            worksheet['C1'] = 'FIRST NAME'
            worksheet['D1'] = 'MIDDLE NAME'
            worksheet['E1'] = 'EXTENSION'
            worksheet['F1'] = 'SEX'
            worksheet['G1'] = 'USERNAME'
            worksheet['H1'] = 'EMPLOYMENT STATUS'
            worksheet['I1'] = 'POSITION'
            worksheet['J1'] = 'SALARY RATE'
            worksheet['K1'] = 'SALARY GRADE'
            worksheet['L1'] = 'STEP INCREMENT'
            worksheet['M1'] = 'AREA OF ASSIGNMENT'
            worksheet['N1'] = 'SECTION'
            worksheet['O1'] = 'DIVISION'
            worksheet['P1'] = 'ERROR / DISCREPANCY'

            worksheet['A1'].font = Font(bold=True)
            worksheet['B1'].font = Font(bold=True)
            worksheet['C1'].font = Font(bold=True)
            worksheet['D1'].font = Font(bold=True)
            worksheet['E1'].font = Font(bold=True)
            worksheet['F1'].font = Font(bold=True)
            worksheet['G1'].font = Font(bold=True)
            worksheet['H1'].font = Font(bold=True)
            worksheet['I1'].font = Font(bold=True)
            worksheet['J1'].font = Font(bold=True)
            worksheet['K1'].font = Font(bold=True)
            worksheet['L1'].font = Font(bold=True)
            worksheet['M1'].font = Font(bold=True)
            worksheet['N1'].font = Font(bold=True)
            worksheet['O1'].font = Font(bold=True)
            worksheet['P1'].font = Font(bold=True)

            for emp in datasource:
                haserror = False

                # Check every entry here
                check_id = check_user = check_ext = check_es = check_pos = check_aoa = check_section = check_division = False
                q1 = Empprofile.objects.filter(id_number=emp['id']).first()
                q2 = AuthUser.objects.filter(username=emp['un']).first()
                q3 = ExtensionName.objects.filter(name=emp['ext']).first()
                q4 = Empstatus.objects.filter(acronym=emp['es']).first()
                q5 = Position.objects.filter(Q(acronym=emp['pos']) | Q(name=emp['pos'])).first()
                q6 = Aoa.objects.filter(name=emp['aoa']).first()
                q7 = Section.objects.filter(sec_name=emp['section']).first()
                q8 = Division.objects.filter(div_name=emp['division']).first()

                if q1 is None:
                    check_id = True
                if q2 is None and not emp['un'] == '':
                    check_user = True
                if q3 is not None or emp['ext'] == '':
                    check_ext = True
                if q4 is not None and not emp['es'] == '':
                    check_es = True
                if q5 is not None and not emp['pos'] == '':
                    check_pos = True
                if q6 is not None or emp['aoa'] == '':
                    check_aoa = True
                if q7 is not None or emp['section'] == '':
                    check_section = True
                if q8 is not None or emp['division'] == '':
                    check_division = True

                if not len(emp['id']) == 9 and not check_id:
                    haserror = True
                    firsterror = 'ID Number'
                elif not check_user:
                    haserror = True
                    firsterror = 'Username'
                elif emp['ln'] == '' or emp['ln'] is None:
                    haserror = True
                    firsterror = 'Last Name'
                elif emp['fn'] == '' or emp['fn'] is None:
                    haserror = True
                    firsterror = 'First Name'
                elif not check_ext:
                    haserror = True
                    firsterror = 'Extension'
                elif not check_es:
                    haserror = True
                    firsterror = 'Employment Status'
                elif not check_pos:
                    haserror = True
                    firsterror = 'Position'
                elif not isinstance(float(emp['sr']), float):
                    haserror = True
                    firsterror = 'Salary Rate'
                elif not isinstance(int(emp['sg']), int) and not int(emp['sg']) == 0:
                    haserror = True
                    firsterror = 'Salary Grade'
                elif not isinstance(int(emp['si']), int) and int(emp['si']) == 0:
                    haserror = True
                    firsterror = 'Step Increment'
                elif emp['sex'] != "M" and emp['sex'] != "F":
                    haserror = True
                    firsterror = 'Sex'
                elif not check_aoa:
                    haserror = True
                    firsterror = 'Area of Assignment'
                elif not check_section:
                    haserror = True
                    firsterror = 'Section'
                elif not check_division:
                    haserror = True
                    firsterror = 'Division'
                else:
                    haserror = False

                if haserror:
                    worksheet['A' + str(wsrow)] = emp['id']
                    worksheet['B' + str(wsrow)] = emp['ln']
                    worksheet['C' + str(wsrow)] = emp['fn']
                    worksheet['D' + str(wsrow)] = emp['mn']
                    worksheet['E' + str(wsrow)] = emp['ext']
                    worksheet['F' + str(wsrow)] = emp['sex']
                    worksheet['G' + str(wsrow)] = emp['un']
                    worksheet['H' + str(wsrow)] = emp['es']
                    worksheet['I' + str(wsrow)] = emp['pos']
                    worksheet['J' + str(wsrow)] = emp['sr']
                    worksheet['K' + str(wsrow)] = emp['sg']
                    worksheet['L' + str(wsrow)] = emp['si']
                    worksheet['M' + str(wsrow)] = emp['aoa']
                    worksheet['N' + str(wsrow)] = emp['section']
                    worksheet['O' + str(wsrow)] = emp['division']
                    worksheet['P' + str(wsrow)] = firsterror
                    wsrow = wsrow + 1
                    haserror = False
                    hasatleastoneerror = True
                    errorcounter = errorcounter + 1
                else:
                    okcounter = okcounter + 1
                    user = AuthUser(
                        first_name=emp['fn'],
                        middle_name=emp['mn'],
                        last_name=emp['ln'],
                        username=emp['un'],
                        password=make_password('dswd123$'),
                        is_superuser=0,
                        is_staff=1,
                        is_active=1,
                        date_joined=datetime.now())
                    user.save()

                    pi = Personalinfo(
                        user_id=user.id,
                        gender=1 if emp['sex'] == "M" else 2)
                    pi.save()

                    # Create employee profile with additional fields
                    employee = Empprofile(
                        pi_id=pi.id,
                        id_number=emp['id'],
                        empstatus_id=q4.id,
                        position_id=q5.id,
                        salary_rate=float(emp['sr'].replace(',', '')),
                        salary_grade=emp['sg'],
                        step_inc=emp['si']
                    )
                    
                    # Add optional fields if provided
                    if q6:
                        employee.aoa_id = q6.id
                    if q7:
                        employee.section_id = q7.id
                    
                    employee.save()
                    
            if hasatleastoneerror:
                now = str(datetime.timestamp(datetime.now()))
                response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename=Upload' + ' (' + now + ').xlsx'
                return response

        messages.success(request, mark_safe("Bulk import of employees was successful!<br><br>Uploaded: {} <br>Errors: {}".format(okcounter, errorcounter)))
        return redirect('backend-employees')



@login_required
@permission_required("auth.employee_list")
def employees(request):
    page = request.GET.get('page', 1)
    rows = request.GET.get('rows', 25)
    search = request.GET.get('search', '')
    data = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
                            7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}
    gender_sex = [
        {'id': 1, 'name': 'Male'},
        {'id': 2, 'name': 'Female'}
    ]
    context = {
        'employee': Paginator(Empprofile.objects.order_by('pi__user__last_name').filter(Q(id_number__icontains=search) |
                                                                                        Q(pi__user__username__icontains=search) |
                                                                                        Q(pi__user__first_name__icontains=search) |
                                                                                        Q(pi__user__last_name__icontains=search) |
                                                                                        Q(pi__user__middle_name__icontains=search)),rows).page(page),
        'fundsource': Fundsource.objects.filter(status=1).order_by('name'),
        'aoa': Aoa.objects.filter(status=1).order_by('name'),
        'project': Project.objects.filter(status=1).order_by('name'),
        'position': Position.objects.filter(status=1).order_by('name'),
        'empstatus': Empstatus.objects.filter(status=1).order_by('name'),
        'ext': ExtensionName.objects.filter(status=1).order_by('name'),
        'sg': Salarygrade.objects.all(),
        'stepinc': Stepinc.objects.all(),
        'rows': rows,
        'tab_title': 'Employees List',
        'title': 'employee',
        'sub_title': 'all_employee',
        'division': Division.objects.all().order_by('div_name'),
        'section': Section.objects.all().order_by('sec_name'),
        'data': sorted(data.items()),
        'bt': Bloodtype.objects.all().order_by('name'),
        'gender': gender_sex
    }

    return render(request, 'backend/employee_data/staff/employees.html', context)

def generate_id_number():
    # Get the current year and month
    now = datetime.now()
    current_year = now.year % 100  # Get the last two digits of the current year
    current_month = now.month  # Get the current month

    # Format the month in two digits
    month_str = f"{current_month:02}"

    # Define the prefix based on the current year and month
    id_prefix = f"10-{current_year:02}{month_str}"

    # Find the maximum ID number that starts with the prefix
    max_id = Empprofile.objects.filter(
        id_number__startswith=id_prefix
    ).aggregate(Max('id_number'))['id_number__max']

    # Determine the next applicant number
    applicant_number = 1
    if max_id:
        last_applicant_number = int(max_id[-3:])  # Extract the last 3 digits
        applicant_number = last_applicant_number + 1

    # Format the new ID number with a three-digit applicant number
    id_number = f"{id_prefix}{applicant_number:03}"

    return id_number

@login_required
@permission_required("auth.register_employee")
def register_employee(request):
    if request.method == 'POST':
        with transaction.atomic():
            check_user = False
            # user = searchSamAccountName(request.POST.get('username'))
            if AuthUser.objects.filter(username=request.POST.get('username')):
                return JsonResponse({'msg': 'The username you entered is already in used. Please try another one.'})
            # elif user["status"]:
            #    return JsonResponse({'error': 'The username you entered is already in used. Please try another one.'})
            elif Empprofile.objects.filter(id_number=request.POST.get('id_number')):
                return JsonResponse({'msg': 'The id number you have inputted is already in used. Please try another one.'})
            else:
                check_user = True

            if check_user:
                if request.POST.get('id_number'):
                    id_number = request.POST.get('id_number')
                else:
                    id_number = generate_id_number()

                user = AuthUser(
                    first_name=request.POST.get('first_name'),
                    middle_name=request.POST.get('middle_name'),
                    last_name=request.POST.get('last_name'),
                    username=request.POST.get('username'),
                    password=make_password(id_number),
                    email=request.POST.get('email'),
                    is_superuser=0,
                    is_staff=1,
                    is_active=1,
                    date_joined=datetime.now())

                user.save()

                pi = Personalinfo(
                    user_id=user.id,
                    gender=request.POST.get('gender'),
                    mobile_no=request.POST.get('contact_number')
                )

                pi.save()

                emp = Empprofile(
                    pi_id=pi.id,
                    id_number=request.POST.get('id_number') if request.POST.get('id_number') else id_number,
                    aoa_id=request.POST.get('aoa'),
                    fundsource_id=request.POST.get('fundsource'),
                    empstatus_id=request.POST.get('empstatus'),
                    position_id=request.POST.get('position'),
                    salary_rate=float(request.POST.get('sr').replace(',', '')),
                    salary_grade=request.POST.get('sg'),
                    step_inc=request.POST.get('si'),
                    section_id=request.POST.get('section')
                )

                emp.save()

                # create_ad_account(id_number,
                #                  request.POST.get('first_name'),
                #                  request.POST.get('middle_name'),
                #                  request.POST.get('last_name'),
                #                  user.get_fullname,
                #                  request.POST.get('username'),
                #                  request.POST.get('email'),
                #                  request.POST.get('contact_number'),
                #                  emp.position.name,
                #                  emp.section.sec_name,
                #                  emp.id)

                # send_notification(
                #    "Hi {}! This is an automated message to inform you that you have been registered to Caraga PORTAL. Your username is {} and your password is {}. Thank you and God bless! - The My PORTAL Team".
                #    format(request.POST.get('first_name').title(), request.POST.get('username'),request.POST.get('password') if request.POST.get('password') else id_number),
                #    request.POST.get('contact_number'),
                #    request.session['emp_id'],
                #    emp.id
                # )

            return JsonResponse({'data': 'success', 'msg': 'The employee with the ID Number "{}" was added successfully'.format(request.POST.get('id_number') if request.POST.get('id_number') else id_number)})
        return JsonResponse({'error': True, 'msg': 'Unauthorized transaction'})


@login_required
@csrf_exempt
@permission_required("auth.edit_employee")
def ad_account_creation(request):
    if request.method == "POST":
        try:
            user = searchSamAccountName(request.POST.get('username'))
            if user['status']:
                return JsonResponse({'error': True, 'msg': 'The username you entered is already in used. Please try another one.'})
            else:
                result = create_ad_account(str(request.POST.get('id_number').strip()),
                                  str(request.POST.get('first_name')),
                                  str(request.POST.get('initials')),
                                  str(request.POST.get('last_name')),
                                  str(request.POST.get('display_name')),
                                  str(request.POST.get('username')),
                                  str(request.POST.get('email')),
                                  str(request.POST.get('contact_number')),
                                  str(request.POST.get('title')),
                                  str(request.POST.get('department')),
                                  str(request.POST.get('emp_id')))

                return JsonResponse({'data': 'success', 'msg': result})
        except Exception as e:
            PortalErrorLogs.objects.create(
                logs="AD Account Creation: {}".format(e),
                date_created=datetime.now(),
                emp_id=request.session['emp_id']
            )


@login_required
@permission_required("admin.superadmin")
def reset_password(request, pk):
    emp = Empprofile.objects.filter(id=pk).first()
    if request.method == "POST":
        try:
            # user = searchSamAccountName(request.POST.get('username'))
            #
            # if user["status"]:
            #     # Password Generator with 8 random numbers and letters
            #     chars = "ABCDEFGHJKMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz123456789"
            #     random_password = ''.join(random.choice(chars) for _ in range(8))
            #     enc_pwd = '"{}"'.format(str(random_password)).encode('utf-16-le')
            #     user["connection"].modify(user["userDN"], {'unicodePwd': [(MODIFY_REPLACE, [enc_pwd])]})
            #
            #     # Perform a search to retrieve the userAccountControl attribute
            #     user["connection"].search(search_base=user["userDN"],
            #                               search_filter='(objectclass=person)',
            #                               attributes=['userAccountControl'])
            #
            #     # Retrieve the userAccountControl attribute from the response
            #     uac = user["connection"].response[0]['attributes']['userAccountControl']
            #
            #     # Set the DONT_EXPIRE_PASSWORD flag
            #     new_uac = uac | 0x10000  # 0x10000 = 65536 = ADS_UF_DONT_EXPIRE_PASSWD
            #
            #     # Ensure the LOCKOUT flag is not set
            #     new_uac &= ~0x10  # 0x10 = 16 = ADS_UF_LOCKOUT
            #
            #     # Ensure the ACCOUNTDISABLE flag is not set (to enable the account)
            #     new_uac &= ~0x2  # 0x2 = 2 = ADS_UF_ACCOUNTDISABLE
            #
            #     # Update userAccountControl
            #     user["connection"].modify(user["userDN"], {'userAccountControl': [(MODIFY_REPLACE, [new_uac])]})
            #
            #     AuthUser.objects.filter(id=request.POST.get('user_id')).update(
            #         password=make_password(str(random_password))
            #     )
            #
            #     message = "Good day, {}! This is an automated message to inform you that your password has been resetted. Your username is {} and your new password is {}. - The My PORTAL Team".format(
            #         emp.pi.user.first_name, emp.pi.user.username, random_password)
            #
            #     send_notification(message, emp.pi.mobile_no, request.session['emp_id'], emp.id)
            #     return JsonResponse({'data': 'success'})
            # else:
            #     return JsonResponse({'error': "Cannot change password for user '{}' or username does not exist in Active Directory.".format(request.POST.get('username'))})

            AuthUser.objects.filter(id=request.POST.get('user_id')).update(
                password=make_password(emp.id_number)
            )

            message = "Good day, {}! This is an automated message to inform you that your password has been resetted. Your username is {} and your new password is {}. - The CDO PORTAL Team".format(
                emp.pi.user.first_name, emp.pi.user.username, emp.id_number)

            send_notification(message, emp.pi.mobile_no, request.session['emp_id'], emp.id)
            return JsonResponse({'data': 'success'})
        except Exception as e:
            PortalErrorLogs.objects.create(
                logs="Password Reset: {}".format(e),
                date_created=datetime.now(),
                emp_id=request.session['emp_id']
            )

    context = {
        'emp': emp
    }
    return render(request, 'backend/employee_data/staff/reset_password.html', context)


@login_required
@permission_required('auth.staffing')
@csrf_exempt
def export_data(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'DATA BY EXPORTING'
    wsrow = 2

    if request.POST.getlist('months[]'):
        worksheet['A1'] = 'DIVISION'
        worksheet['B1'] = 'SECTION/UNIT'
        worksheet['C1'] = 'OFFICE LOCATION/OFFICIAL STATION'
        worksheet['D1'] = 'POSITION TITLE'
        worksheet['E1'] = 'COMPLETE NAME (CRUZ, JUAN PEREZ)'
        worksheet['F1'] = 'LASTNAME'
        worksheet['G1'] = 'FIRST NAME'
        worksheet['H1'] = 'MIDDLE NAME'
        worksheet['I1'] = 'EXT.'
        worksheet['J1'] = 'GENDER'
        worksheet['K1'] = 'DATE OF BIRTH'
        worksheet['L1'] = 'AGE'
        worksheet['M1'] = 'EMAIL'

        worksheet['A1'].font = Font(bold=True)
        worksheet['B1'].font = Font(bold=True)
        worksheet['C1'].font = Font(bold=True)
        worksheet['D1'].font = Font(bold=True)
        worksheet['E1'].font = Font(bold=True)
        worksheet['F1'].font = Font(bold=True)
        worksheet['G1'].font = Font(bold=True)
        worksheet['H1'].font = Font(bold=True)
        worksheet['I1'].font = Font(bold=True)
        worksheet['J1'].font = Font(bold=True)
        worksheet['K1'].font = Font(bold=True)
        worksheet['L1'].font = Font(bold=True)
        worksheet['M1'].font = Font(bold=True)

        bday = \
            Empprofile.objects.filter(pi__dob__month__in=request.POST.getlist('months[]'), pi__user__is_active=1).order_by('pi__user__last_name').values(
            'section_id',
            'pi__user__last_name',
            'pi__user__first_name',
            'pi__user__middle_name',
            'pi__ext__name',
            'pi__gender',
            'empstatus__acronym',
            'position__name',
            'pi__dob',
            'aoa_id',
            'pi_id',
            'pi__mobile_no',
            'pi__user__email',
            )
        if bday:
            for (indx, row) in enumerate(bday):
                lastname = row['pi__user__last_name']
                middlename = row['pi__user__middle_name']
                firstname = row['pi__user__first_name']
                ext = row['pi__ext__name']
                worksheet['A' + str(wsrow)] = \
                    get_division(row['section_id'])
                worksheet['B' + str(wsrow)] = (''
                         if get_aoa(row['aoa_id'
                        ]) else get_section(row['section_id']))
                worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                worksheet['D' + str(wsrow)] = row['position__name']
                worksheet['E' + str(wsrow)] = \
                    '{}, {} {} {}.'.format(str(lastname.upper()),
                        str(firstname.upper()),
                        str(middlename.upper()), ('' if ext
                        is None else ext))
                worksheet['F' + str(wsrow)] = str(lastname.upper())
                worksheet['G' + str(wsrow)] = str(firstname.upper())
                worksheet['H' + str(wsrow)] = str(middlename.upper())
                worksheet['I' + str(wsrow)] = ext
                worksheet['J' + str(wsrow)] = ('Male'
                         if row['pi__gender'] == 1 else 'Female')
                worksheet['K' + str(wsrow)] = row['pi__dob']
                worksheet['L' + str(wsrow)] = get_age(row['pi__dob'])
                worksheet['M' + str(wsrow)] = row['pi__user__email']
                wsrow = wsrow + 1
            response = \
                HttpResponse(content=save_virtual_workbook(workbook),
                             content_type='application/ms-excel')
            response['Content-Disposition'] = \
                'attachment; filename=Staffing Birthdays By Month.xlsx'
            return response
    elif request.POST.getlist('bt[]'):
        worksheet['A1'] = 'DIVISION'
        worksheet['B1'] = 'SECTION/UNIT'
        worksheet['C1'] = 'OFFICE LOCATION/OFFICIAL STATION'
        worksheet['D1'] = 'POSITION TITLE'
        worksheet['E1'] = 'COMPLETE NAME (CRUZ, JUAN PEREZ)'
        worksheet['F1'] = 'LASTNAME'
        worksheet['G1'] = 'FIRST NAME'
        worksheet['H1'] = 'MIDDLE NAME'
        worksheet['I1'] = 'EXT.'
        worksheet['J1'] = 'GENDER'
        worksheet['K1'] = 'DATE OF BIRTH'
        worksheet['L1'] = 'AGE'
        worksheet['M1'] = 'EMAIL'
        worksheet['N1'] = 'BLOODTYPE'

        worksheet['A1'].font = Font(bold=True)
        worksheet['B1'].font = Font(bold=True)
        worksheet['C1'].font = Font(bold=True)
        worksheet['D1'].font = Font(bold=True)
        worksheet['E1'].font = Font(bold=True)
        worksheet['F1'].font = Font(bold=True)
        worksheet['G1'].font = Font(bold=True)
        worksheet['H1'].font = Font(bold=True)
        worksheet['I1'].font = Font(bold=True)
        worksheet['J1'].font = Font(bold=True)
        worksheet['K1'].font = Font(bold=True)
        worksheet['L1'].font = Font(bold=True)
        worksheet['M1'].font = Font(bold=True)
        worksheet['N1'].font = Font(bold=True)

        bt = \
            Empprofile.objects.filter(pi__bt_id__in=request.POST.getlist('bt[]'), pi__user__is_active=1).order_by('pi__user__last_name').values(
            'section_id',
            'pi__user__last_name',
            'pi__user__first_name',
            'pi__user__middle_name',
            'pi__ext__name',
            'pi__gender',
            'empstatus__acronym',
            'position__name',
            'pi__dob',
            'aoa_id',
            'pi_id',
            'pi__mobile_no',
            'pi__user__email',
            'pi__bt_id',
            )
        if bt:
            for (indx, row) in enumerate(bt):
                lastname = row['pi__user__last_name']
                middlename = row['pi__user__middle_name']
                firstname = row['pi__user__first_name']
                ext = row['pi__ext__name']
                worksheet['A' + str(wsrow)] = \
                    get_division(row['section_id'])
                worksheet['B' + str(wsrow)] = (''
                         if get_aoa(row['aoa_id'
                        ]) else get_section(row['section_id']))
                worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                worksheet['D' + str(wsrow)] = row['position__name']
                worksheet['E' + str(wsrow)] = \
                    '{}, {} {} {}.'.format(str(lastname.upper()),
                        str(firstname.upper()),
                        str(middlename.upper()), ('' if ext
                        is None else ext))
                worksheet['F' + str(wsrow)] = str(lastname.upper())
                worksheet['G' + str(wsrow)] = str(firstname.upper())
                worksheet['H' + str(wsrow)] = str(middlename.upper())
                worksheet['I' + str(wsrow)] = ext
                worksheet['J' + str(wsrow)] = ('Male'
                         if row['pi__gender'] == 1 else 'Female')
                worksheet['K' + str(wsrow)] = row['pi__dob']
                worksheet['L' + str(wsrow)] = get_age(row['pi__dob'])
                worksheet['M' + str(wsrow)] = row['pi__user__email']
                worksheet['N' + str(wsrow)] = \
                    get_bloodtype(row['pi__bt_id'])
                wsrow = wsrow + 1
            response = \
                HttpResponse(content=save_virtual_workbook(workbook),
                             content_type='application/ms-excel')
            response['Content-Disposition'] = \
                'attachment; filename=Staffing Bloodtypes.xlsx'
            return response
    else:
        worksheet['A1'] = 'DIVISION'
        worksheet['B1'] = 'SECTION/UNIT'
        worksheet['C1'] = 'OFFICE LOCATION/OFFICIAL STATION'
        worksheet['D1'] = 'POSITION TITLE'
        worksheet['E1'] = 'COMPLETE NAME (DELA CRUZ, JUAN)'
        worksheet['F1'] = 'LASTNAME'
        worksheet['G1'] = 'FIRST NAME'
        worksheet['H1'] = 'MIDDLE NAME'
        worksheet['I1'] = 'EXT.'
        worksheet['J1'] = 'GENDER'
        worksheet['K1'] = 'DATE OF BIRTH'
        worksheet['L1'] = 'AGE'
        worksheet['M1'] = 'EMAIL'
        worksheet['N1'] = 'BLOODTYPE'

        worksheet['A1'].font = Font(bold=True)
        worksheet['B1'].font = Font(bold=True)
        worksheet['C1'].font = Font(bold=True)
        worksheet['D1'].font = Font(bold=True)
        worksheet['E1'].font = Font(bold=True)
        worksheet['F1'].font = Font(bold=True)
        worksheet['G1'].font = Font(bold=True)
        worksheet['H1'].font = Font(bold=True)
        worksheet['I1'].font = Font(bold=True)
        worksheet['J1'].font = Font(bold=True)
        worksheet['K1'].font = Font(bold=True)
        worksheet['L1'].font = Font(bold=True)
        worksheet['M1'].font = Font(bold=True)
        worksheet['N1'].font = Font(bold=True)

        position_id = request.GET.get('position_id')
        section_id = request.GET.get('section_id')

        data = Empprofile.objects.filter(pi__user__is_active=1)

        if position_id:
            data = data.filter(position_id=position_id)
        elif section_id:
            data = data.filter(section_id=section_id)

        data = data.order_by(
                'pi__user__last_name').values(
                'section_id',
                'pi__user__last_name',
                'pi__user__first_name',
                'pi__user__middle_name',
                'pi__ext__name',
                'pi__gender',
                'empstatus__acronym',
                'position__name',
                'pi__dob',
                'aoa_id',
                'pi_id',
                'pi__mobile_no',
                'pi__user__email',
                'pi__bt_id',
            )

        if position_id:
            data = data.filter(position_id=position_id)

        if data:
            for (indx, row) in enumerate(data):
                lastname = row['pi__user__last_name']
                middlename = row['pi__user__middle_name']
                firstname = row['pi__user__first_name']
                ext = row['pi__ext__name']
                worksheet['A' + str(wsrow)] = get_division(row['section_id'])
                worksheet['B' + str(wsrow)] = ('' if get_aoa(row['aoa_id']) else get_section(row['section_id']))
                worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                worksheet['D' + str(wsrow)] = row['position__name']
                worksheet['E' + str(wsrow)] = '{}, {} {} {}.'.format(
                    str(lastname.upper()),
                    str(firstname.upper()),
                    str(middlename.upper()),
                    ('' if ext is None else ext)
                )
                worksheet['F' + str(wsrow)] = str(lastname.upper())
                worksheet['G' + str(wsrow)] = str(firstname.upper())
                worksheet['H' + str(wsrow)] = str(middlename.upper())
                worksheet['I' + str(wsrow)] = ext
                worksheet['J' + str(wsrow)] = ('Male' if row['pi__gender'] == 1 else 'Female')
                worksheet['K' + str(wsrow)] = row['pi__dob']
                worksheet['L' + str(wsrow)] = get_age(row['pi__dob'])
                worksheet['M' + str(wsrow)] = row['pi__user__email']
                worksheet['N' + str(wsrow)] = get_bloodtype(row['pi__bt_id'])
                wsrow = wsrow + 1

            file_type = request.GET.get('type', 'excel')

            if file_type == 'csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename=Staffing.csv'
                writer = csv.writer(response)

                for row in worksheet.rows:
                    writer.writerow([cell.value for cell in row])
            else:
                response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename=Staffing.xlsx'

            return response


@login_required
@permission_required("auth.staffing")
@csrf_exempt
def export_employees(request):
    today = datetime.today()
    year = today.strftime("%Y")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'STAFFING ' + year
    wsrow = 2
    worksheet['A1'] = 'DIVISION'
    worksheet['B1'] = 'SECTION/UNIT'
    worksheet['C1'] = 'OFFICE LOCATION/OFFICIAL STATION'
    worksheet['D1'] = 'ITEM NUMBER'
    worksheet['E1'] = 'DATE OF CREATION OF POSITION'
    worksheet['F1'] = 'POSITION TITLE'
    worksheet['G1'] = 'PARENTHETICAL TITLE'
    worksheet['H1'] = 'POSITION LEVEL'
    worksheet['I1'] = 'SG'
    worksheet['J1'] = 'MONTHLY SALARY'
    worksheet['K1'] = 'DESIGNATION (AS APPROPRIATE) BASED ON SO/RSO'
    worksheet['L1'] = 'DATE OF DESIGNATION'
    worksheet['M1'] = 'SPECIAL ORDER NO.'
    worksheet['N1'] = 'OFFICE/BUREAU/SERVICE/PROGRAM(Plantilla Assignment based on PSIPOP)'
    worksheet['O1'] = 'FUND SOURCE FOR CONTRACTUAL, CONTRACT OF SERVICE AND JOB ORDER (BASED ON CREATION)'
    worksheet['P1'] = 'CLASSIFICATION OF EMPLOYMENT (PERMANENT, COTERMINOUS, CASUAL, CONTRACTUAL, CONTRACT OF SERVICE, JOB ORDER)'
    worksheet['Q1'] = 'STATUS'
    worksheet['R1'] = 'MODE OF ACCESSION (FOR PERMANENT AND COTERMINOUS ONLY)'
    worksheet['S1'] = 'DATE FILLED UP'
    worksheet['T1'] = 'COMPLETE NAME (CRUZ, JUAN PEREZ)'
    worksheet['U1'] = 'LASTNAME'
    worksheet['V1'] = 'FIRST NAME'
    worksheet['W1'] = 'MIDDLE NAME'
    worksheet['X1'] = 'EXT.'
    worksheet['Y1'] = 'DATE OF ORIGINAL APPOINTMENT'
    worksheet['Z1'] = 'DATE OF LAST PROMOTION'
    worksheet['AA1'] = 'ENTRY DATE IN DSWD (FIRST DAY IN SERVICE)'
    worksheet['AB1'] = 'ELIGIBILITY \n (CS and other Eligibilities)'
    worksheet['AC1'] = 'ELIGIBILITY \n (License - RA 1080)'
    worksheet['AD1'] = 'LICENSE \n (RA 1080-LET, RN,RS,ETC)'
    worksheet['AE1'] = 'HIGHEST LEVEL OF ELIGIBILITY (1ST AND 2ND LEVEL ONLY)'
    worksheet['AF1'] = 'HIGHEST EDUCATION COMPLETED'
    worksheet['AG1'] = 'DEGREE AND COURSE \n (1st Course/Vocational)' 
    worksheet['AH1'] = 'DEGREE AND COURSE \n (2nd Course)'
    worksheet['AI1'] = 'OTHER COURSE/S'
    worksheet['AJ1'] = 'MASTERS OR DOCTORAL DEGREE (Specify)'
    worksheet['AK1'] = 'GENDER'
    worksheet['AL1'] = 'DATE OF BIRTH'
    worksheet['AM1'] = 'AGE'
    worksheet['AN1'] = 'CIVIL STATUS'
    worksheet['AO1'] = 'RESIDENTIAL ADDRESS'
    worksheet['AP1'] = 'PERMANENT ADDRESS'
    worksheet['AQ1'] = 'INDICATE WHETHER SOLO PARENT'
    worksheet['AR1'] = 'INDICATE WHETHER SENIOR CITIZEN'
    worksheet['AS1'] = 'INDICATE WHETHER PWD'
    worksheet['AT1'] = 'TYPE OF DISABILITY'
    worksheet['AU1'] = 'INDICATE WHETHER MEMBER OF INDIGINOUS GROUP'
    worksheet['AV1'] = 'INDIGENOUS GROUP'
    worksheet['AW1'] = 'CITIZENSHIP'
    worksheet['AX1'] = 'ACTIVE CONTACT NOS.'
    worksheet['AY1'] = 'ACTIVE AND WORKING EMAIL ADDRESS'
    worksheet['AZ1'] = 'FORMER INCUMBENT'
    worksheet['BA1'] = 'MODE OF SEPARATION'
    worksheet['BB1'] = 'DATE VACATED'
    worksheet['BC1'] = 'REMARKS / STATUS OF VACANT POSITION'

    worksheet['A1'].font = Font(bold=True)
    worksheet['B1'].font = Font(bold=True)
    worksheet['C1'].font = Font(bold=True)
    worksheet['D1'].font = Font(bold=True)
    worksheet['E1'].font = Font(bold=True)
    worksheet['F1'].font = Font(bold=True)
    worksheet['G1'].font = Font(bold=True)
    worksheet['H1'].font = Font(bold=True)
    worksheet['I1'].font = Font(bold=True)
    worksheet['J1'].font = Font(bold=True)
    worksheet['K1'].font = Font(bold=True)
    worksheet['L1'].font = Font(bold=True)
    worksheet['M1'].font = Font(bold=True)
    worksheet['N1'].font = Font(bold=True)
    worksheet['O1'].font = Font(bold=True)
    worksheet['P1'].font = Font(bold=True)
    worksheet['Q1'].font = Font(bold=True)
    worksheet['R1'].font = Font(bold=True)
    worksheet['S1'].font = Font(bold=True)
    worksheet['T1'].font = Font(bold=True)
    worksheet['U1'].font = Font(bold=True)
    worksheet['V1'].font = Font(bold=True)
    worksheet['W1'].font = Font(bold=True)
    worksheet['X1'].font = Font(bold=True)
    worksheet['Y1'].font = Font(bold=True)
    worksheet['Z1'].font = Font(bold=True)

    worksheet['AA1'].font = Font(bold=True)
    worksheet['AB1'].font = Font(bold=True)
    worksheet['AC1'].font = Font(bold=True)
    worksheet['AD1'].font = Font(bold=True)
    worksheet['AE1'].font = Font(bold=True)
    worksheet['AF1'].font = Font(bold=True)
    worksheet['AG1'].font = Font(bold=True)
    worksheet['AH1'].font = Font(bold=True)
    worksheet['AI1'].font = Font(bold=True)
    worksheet['AJ1'].font = Font(bold=True)
    worksheet['AK1'].font = Font(bold=True)
    worksheet['AL1'].font = Font(bold=True)
    worksheet['AM1'].font = Font(bold=True)
    worksheet['AN1'].font = Font(bold=True)
    worksheet['AO1'].font = Font(bold=True)
    worksheet['AP1'].font = Font(bold=True)
    worksheet['AQ1'].font = Font(bold=True)
    worksheet['AR1'].font = Font(bold=True)
    worksheet['AS1'].font = Font(bold=True)
    worksheet['AT1'].font = Font(bold=True)
    worksheet['AU1'].font = Font(bold=True)
    worksheet['AV1'].font = Font(bold=True)
    worksheet['AW1'].font = Font(bold=True)
    worksheet['AX1'].font = Font(bold=True)
    worksheet['AY1'].font = Font(bold=True)
    worksheet['AZ1'].font = Font(bold=True)
    worksheet['BA1'].font = Font(bold=True)
    worksheet['BB1'].font = Font(bold=True)
    worksheet['BC1'].font = Font(bold=True)

    if request.POST.getlist('section[]'):
        sections = Empprofile.objects.filter(section_id__in=request.POST.getlist('section[]'), pi__user__is_active=1).order_by('pi__user__last_name').values('section_id',
                                                                               'pi__user__last_name',
                                                                               'pi__user__first_name',
                                                                               'pi__user__middle_name',
                                                                               'pi__ext__name',
                                                                               'pi__gender',
                                                                               'salary_grade',
                                                                               'salary_rate',
                                                                               'empstatus__acronym',
                                                                               'position__name',
                                                                               'fundsource__name',
                                                                               'pi__cs_id',
                                                                               'pi__dob',
                                                                               'aoa_id',
                                                                               'pi_id',
                                                                               'pi__mobile_no',
                                                                               'pi__user__email',
                                                                               'pi__isfilipino',
                                                                               'pi__user__is_active',
                                                                               'item_number',
                                                                               'former_incumbent',
                                                                               'mode_sep__name',
                                                                               'mode_access__name',
                                                                               'date_vacated',
                                                                               'dateofcreation_pos',
                                                                               'designation',
                                                                               'dateof_designation',
                                                                               'dateof_orig_appointment',
                                                                               'dateof_last_promotion',
                                                                               'remarks_vacated',
                                                                               'pi__type_of_disability',
                                                                               'pi__type_of_indi__name',
                                                                               'specialorder_no',
                                                                               'plantilla_psipop')

        if sections:
            for indx, row in enumerate(sections):
                if row['pi__user__is_active'] == 1:
                    lastname = row['pi__user__last_name']
                    middlename = row['pi__user__middle_name']
                    firstname = row['pi__user__first_name']
                    ext = row['pi__ext__name']
                    worksheet['A' + str(wsrow)] = get_division(row['section_id'])
                    worksheet['B' + str(wsrow)] = '' if get_aoa(row['aoa_id']) else get_section(row['section_id'])
                    worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                    worksheet['D' + str(wsrow)] = row['item_number']
                    worksheet['E' + str(wsrow)] = row['dateofcreation_pos']
                    worksheet['F' + str(wsrow)] = row['position__name']
                    worksheet['G' + str(wsrow)] = ''
                    worksheet['H' + str(wsrow)] = ''
                    worksheet['I' + str(wsrow)] = row['salary_grade']
                    worksheet['J' + str(wsrow)] = row['salary_rate']
                    worksheet['K' + str(wsrow)] = row['designation']
                    worksheet['L' + str(wsrow)] = row['dateof_designation']
                    worksheet['M' + str(wsrow)] = row['specialorder_no']
                    worksheet['N' + str(wsrow)] = row['plantilla_psipop']
                    worksheet['O' + str(wsrow)] = row['fundsource__name']
                    worksheet['P' + str(wsrow)] = row['empstatus__acronym']
                    worksheet['Q' + str(wsrow)] = 'Filled'
                    worksheet['R' + str(wsrow)] = row['mode_access__name']
                    worksheet['S' + str(wsrow)] = filledup(row['pi_id'])
                    worksheet['T' + str(wsrow)] = "{}, {} {} {}.".format(str(lastname.upper()), str(firstname.upper()), str(middlename.upper()), "" if ext is None else ext)
                    worksheet['U' + str(wsrow)] = str(lastname.upper())
                    worksheet['V' + str(wsrow)] = str(firstname.upper())
                    worksheet['W' + str(wsrow)] = str(middlename.upper())
                    worksheet['X' + str(wsrow)] = ext
                    worksheet['Y' + str(wsrow)] = row['dateof_orig_appointment']
                    worksheet['Z' + str(wsrow)] = row['dateof_last_promotion']
                    worksheet['AA' + str(wsrow)] = entry_first_indswd(row['pi_id'])
                    worksheet['AB' + str(wsrow)] = get_civilservice(row['pi_id'])
                    worksheet['AC' + str(wsrow)] = get_ra(row['pi_id'])
                    worksheet['AD' + str(wsrow)] = get_let(row['pi_id'])
                    worksheet['AE' + str(wsrow)] = getlevelof_elig(row['pi_id'])
                    worksheet['AF' + str(wsrow)] = geteducation_status(row['pi_id'])
                    worksheet['AG' + str(wsrow)] = get_firstdegree(row['pi_id'])
                    worksheet['AH' + str(wsrow)] = get_lastdegree(row['pi_id'])
                    worksheet['AI' + str(wsrow)] = get_otherdegree(row['pi_id'])
                    worksheet['AJ' + str(wsrow)] = get_masters(row['pi_id'])
                    worksheet['AK' + str(wsrow)] = 'Male' if row['pi__gender'] == 1 else 'Female'
                    worksheet['AL' + str(wsrow)] = row['pi__dob']
                    worksheet['AM' + str(wsrow)] = get_age(row['pi__dob'])
                    worksheet['AN' + str(wsrow)] = get_civil_status(row['pi__cs_id'])
                    worksheet['AO' + str(wsrow)] = get_residentialadd(row['pi_id'])
                    worksheet['AP' + str(wsrow)] = get_permanentadd(row['pi_id'])
                    worksheet['AQ' + str(wsrow)] = get_solo_parent(row['pi_id'])
                    worksheet['AR' + str(wsrow)] = get_senior(row['pi__dob'])
                    worksheet['AS' + str(wsrow)] = 'PWD' if row['pi__type_of_disability'] else ''
                    worksheet['AT' + str(wsrow)] = row['pi__type_of_disability']
                    worksheet['AU' + str(wsrow)] = 'INDIGINOUS' if row['pi__type_of_indi__name'] else ''
                    worksheet['AV' + str(wsrow)] = row['pi__type_of_indi__name']
                    worksheet['AW' + str(wsrow)] = 'Filipino' if row['pi__isfilipino'] == 1 else ""
                    worksheet['AX' + str(wsrow)] = row['pi__mobile_no']
                    worksheet['AY' + str(wsrow)] = row['pi__user__email']
                    worksheet['AZ' + str(wsrow)] = row['former_incumbent']
                    worksheet['BA' + str(wsrow)] = row['mode_sep__name']
                    worksheet['BB' + str(wsrow)] = row['date_vacated']
                    worksheet['BC' + str(wsrow)] = row['remarks_vacated']
                    wsrow = wsrow + 1
                else:
                    worksheet['A' + str(wsrow)] = get_division(row['section_id'])
                    worksheet['B' + str(wsrow)] = '' if get_aoa(row['aoa_id']) else get_section(row['section_id'])
                    worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                    worksheet['D' + str(wsrow)] = row['item_number']
                    worksheet['E' + str(wsrow)] = row['dateofcreation_pos']
                    worksheet['F' + str(wsrow)] = row['position__name']
                    worksheet['G' + str(wsrow)] = ''
                    worksheet['H' + str(wsrow)] = ''
                    worksheet['I' + str(wsrow)] = row['salary_grade']
                    worksheet['J' + str(wsrow)] = row['salary_rate']
                    worksheet['K' + str(wsrow)] = row['designation']
                    worksheet['L' + str(wsrow)] = row['dateof_designation']
                    worksheet['M' + str(wsrow)] = ''
                    worksheet['N' + str(wsrow)] = ''
                    worksheet['O' + str(wsrow)] = row['fundsource__name']
                    worksheet['P' + str(wsrow)] = row['empstatus__acronym']
                    worksheet['Q' + str(wsrow)] = 'Unfilled'
                    worksheet['R' + str(wsrow)] = ''
                    worksheet['S' + str(wsrow)] = ''
                    worksheet['T' + str(wsrow)] = ''
                    worksheet['U' + str(wsrow)] = ''
                    worksheet['V' + str(wsrow)] = ''
                    worksheet['W' + str(wsrow)] = ''
                    worksheet['X' + str(wsrow)] = ''
                    worksheet['Y' + str(wsrow)] = ''
                    worksheet['Z' + str(wsrow)] = ''
                    worksheet['AA' + str(wsrow)] = ''
                    worksheet['AB' + str(wsrow)] = ''
                    worksheet['AC' + str(wsrow)] = ''
                    worksheet['AD' + str(wsrow)] = ''
                    worksheet['AE' + str(wsrow)] = ''
                    worksheet['AF' + str(wsrow)] = ''
                    worksheet['AG' + str(wsrow)] = ''
                    worksheet['AH' + str(wsrow)] = ''
                    worksheet['AI' + str(wsrow)] = ''
                    worksheet['AJ' + str(wsrow)] = ''
                    worksheet['AK' + str(wsrow)] = ''
                    worksheet['AL' + str(wsrow)] = ''
                    worksheet['AM' + str(wsrow)] = ''
                    worksheet['AN' + str(wsrow)] = ''
                    worksheet['AO' + str(wsrow)] = ''
                    worksheet['AP' + str(wsrow)] = ''
                    worksheet['AQ' + str(wsrow)] = ''
                    worksheet['AR' + str(wsrow)] = ''
                    worksheet['AS' + str(wsrow)] = ''
                    worksheet['AT' + str(wsrow)] = ''
                    worksheet['AU' + str(wsrow)] = ''
                    worksheet['AV' + str(wsrow)] = ''
                    worksheet['AW' + str(wsrow)] = ''
                    worksheet['AX' + str(wsrow)] = ''
                    worksheet['AY' + str(wsrow)] = ''
                    worksheet['AZ' + str(wsrow)] = row['former_incumbent']
                    worksheet['BA' + str(wsrow)] = row['mode_sep__name']
                    worksheet['BB' + str(wsrow)] = row['date_vacated']
                    worksheet['BC' + str(wsrow)] = row['remarks_vacated']
                    wsrow = wsrow + 1
            response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Staffing by sections exporting' + ' (' + year + ').xlsx'
            return response
    elif request.POST.getlist('division[]'):
        divisions = Empprofile.objects.all().filter(section__div_id__in=request.POST.getlist('division[]'), pi__user__is_active=1).order_by('pi__user__last_name').values('section_id',
                                                                               'pi__user__last_name',
                                                                               'pi__user__first_name',
                                                                               'pi__user__middle_name',
                                                                               'pi__ext__name',
                                                                               'pi__gender',
                                                                               'salary_grade',
                                                                               'salary_rate',
                                                                               'empstatus__acronym',
                                                                               'position__name',
                                                                               'fundsource__name',
                                                                               'pi__cs_id',
                                                                               'pi__dob',
                                                                               'aoa_id',
                                                                               'pi_id',
                                                                               'pi__mobile_no',
                                                                               'pi__user__email',
                                                                               'pi__isfilipino',
                                                                               'pi__user__is_active',
                                                                               'item_number',
                                                                               'former_incumbent',
                                                                               'mode_sep__name',
                                                                               'mode_access__name',
                                                                               'date_vacated',
                                                                               'dateofcreation_pos',
                                                                               'designation',
                                                                               'dateof_designation',
                                                                               'dateof_orig_appointment',
                                                                               'dateof_last_promotion',
                                                                               'remarks_vacated',
                                                                               'pi__type_of_disability',
                                                                               'pi__type_of_indi__name',
                                                                               'specialorder_no',
                                                                               'plantilla_psipop')

        if divisions:
            for indx, row in enumerate(divisions):
                if row['pi__user__is_active'] == 1:
                    lastname = row['pi__user__last_name']
                    middlename = row['pi__user__middle_name']
                    firstname = row['pi__user__first_name']
                    ext = row['pi__ext__name']
                    worksheet['A' + str(wsrow)] = get_division(row['section_id'])
                    worksheet['B' + str(wsrow)] = '' if get_aoa(row['aoa_id']) else get_section(row['section_id'])
                    worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                    worksheet['D' + str(wsrow)] = row['item_number']
                    worksheet['E' + str(wsrow)] = row['dateofcreation_pos']
                    worksheet['F' + str(wsrow)] = row['position__name']
                    worksheet['G' + str(wsrow)] = ''
                    worksheet['H' + str(wsrow)] = ''
                    worksheet['I' + str(wsrow)] = row['salary_grade']
                    worksheet['J' + str(wsrow)] = row['salary_rate']
                    worksheet['K' + str(wsrow)] = row['designation']
                    worksheet['L' + str(wsrow)] = row['dateof_designation']
                    worksheet['M' + str(wsrow)] = row['specialorder_no']
                    worksheet['N' + str(wsrow)] = row['plantilla_psipop']
                    worksheet['O' + str(wsrow)] = row['fundsource__name']
                    worksheet['P' + str(wsrow)] = row['empstatus__acronym']
                    worksheet['Q' + str(wsrow)] = 'Filled'
                    worksheet['R' + str(wsrow)] = row['mode_access__name']
                    worksheet['S' + str(wsrow)] = filledup(row['pi_id'])
                    worksheet['T' + str(wsrow)] = "{}, {} {} {}.".format(str(lastname.upper()), str(firstname.upper()), str(middlename.upper()), "" if ext is None else ext)
                    worksheet['U' + str(wsrow)] = str(lastname.upper())
                    worksheet['V' + str(wsrow)] = str(firstname.upper())
                    worksheet['W' + str(wsrow)] = str(middlename.upper())
                    worksheet['X' + str(wsrow)] = ext
                    worksheet['Y' + str(wsrow)] = row['dateof_orig_appointment']
                    worksheet['Z' + str(wsrow)] = row['dateof_last_promotion']
                    worksheet['AA' + str(wsrow)] = entry_first_indswd(row['pi_id'])
                    worksheet['AB' + str(wsrow)] = get_civilservice(row['pi_id'])
                    worksheet['AC' + str(wsrow)] = get_ra(row['pi_id'])
                    worksheet['AD' + str(wsrow)] = get_let(row['pi_id'])
                    worksheet['AE' + str(wsrow)] = getlevelof_elig(row['pi_id'])
                    worksheet['AF' + str(wsrow)] = geteducation_status(row['pi_id'])
                    worksheet['AG' + str(wsrow)] = get_firstdegree(row['pi_id'])
                    worksheet['AH' + str(wsrow)] = get_lastdegree(row['pi_id'])
                    worksheet['AI' + str(wsrow)] = get_otherdegree(row['pi_id'])
                    worksheet['AJ' + str(wsrow)] = get_masters(row['pi_id'])
                    worksheet['AK' + str(wsrow)] = 'Male' if row['pi__gender'] == 1 else 'Female'
                    worksheet['AL' + str(wsrow)] = row['pi__dob']
                    worksheet['AM' + str(wsrow)] = get_age(row['pi__dob'])
                    worksheet['AN' + str(wsrow)] = get_civil_status(row['pi__cs_id'])
                    worksheet['AO' + str(wsrow)] = get_residentialadd(row['pi_id'])
                    worksheet['AP' + str(wsrow)] = get_permanentadd(row['pi_id'])
                    worksheet['AQ' + str(wsrow)] = get_solo_parent(row['pi_id'])
                    worksheet['AR' + str(wsrow)] = get_senior(row['pi__dob'])
                    worksheet['AS' + str(wsrow)] = 'PWD' if row['pi__type_of_disability'] else ''
                    worksheet['AT' + str(wsrow)] = row['pi__type_of_disability']
                    worksheet['AU' + str(wsrow)] = 'INDIGINOUS' if row['pi__type_of_indi__name'] else ''
                    worksheet['AV' + str(wsrow)] = row['pi__type_of_indi__name']
                    worksheet['AW' + str(wsrow)] = 'Filipino' if row['pi__isfilipino'] == 1 else ""
                    worksheet['AX' + str(wsrow)] = row['pi__mobile_no']
                    worksheet['AY' + str(wsrow)] = row['pi__user__email']
                    worksheet['AZ' + str(wsrow)] = row['former_incumbent']
                    worksheet['BA' + str(wsrow)] = row['mode_sep__name']
                    worksheet['BB' + str(wsrow)] = row['date_vacated']
                    worksheet['BC' + str(wsrow)] = row['remarks_vacated']
                    wsrow = wsrow + 1
                else:
                    worksheet['A' + str(wsrow)] = get_division(row['section_id'])
                    worksheet['B' + str(wsrow)] = '' if get_aoa(row['aoa_id']) else get_section(row['section_id'])
                    worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                    worksheet['D' + str(wsrow)] = row['item_number']
                    worksheet['E' + str(wsrow)] = row['dateofcreation_pos']
                    worksheet['F' + str(wsrow)] = row['position__name']
                    worksheet['G' + str(wsrow)] = ''
                    worksheet['H' + str(wsrow)] = ''
                    worksheet['I' + str(wsrow)] = row['salary_grade']
                    worksheet['J' + str(wsrow)] = row['salary_rate']
                    worksheet['K' + str(wsrow)] = row['designation']
                    worksheet['L' + str(wsrow)] = row['dateof_designation']
                    worksheet['M' + str(wsrow)] = ''
                    worksheet['N' + str(wsrow)] = ''
                    worksheet['O' + str(wsrow)] = row['fundsource__name']
                    worksheet['P' + str(wsrow)] = row['empstatus__acronym']
                    worksheet['Q' + str(wsrow)] = 'Unfilled'
                    worksheet['R' + str(wsrow)] = ''
                    worksheet['S' + str(wsrow)] = ''
                    worksheet['T' + str(wsrow)] = ''
                    worksheet['U' + str(wsrow)] = ''
                    worksheet['V' + str(wsrow)] = ''
                    worksheet['W' + str(wsrow)] = ''
                    worksheet['X' + str(wsrow)] = ''
                    worksheet['Y' + str(wsrow)] = ''
                    worksheet['Z' + str(wsrow)] = ''
                    worksheet['AA' + str(wsrow)] = ''
                    worksheet['AB' + str(wsrow)] = ''
                    worksheet['AC' + str(wsrow)] = ''
                    worksheet['AD' + str(wsrow)] = ''
                    worksheet['AE' + str(wsrow)] = ''
                    worksheet['AF' + str(wsrow)] = ''
                    worksheet['AG' + str(wsrow)] = ''
                    worksheet['AH' + str(wsrow)] = ''
                    worksheet['AI' + str(wsrow)] = ''
                    worksheet['AJ' + str(wsrow)] = ''
                    worksheet['AK' + str(wsrow)] = ''
                    worksheet['AL' + str(wsrow)] = ''
                    worksheet['AM' + str(wsrow)] = ''
                    worksheet['AN' + str(wsrow)] = ''
                    worksheet['AO' + str(wsrow)] = ''
                    worksheet['AP' + str(wsrow)] = ''
                    worksheet['AQ' + str(wsrow)] = ''
                    worksheet['AR' + str(wsrow)] = ''
                    worksheet['AS' + str(wsrow)] = ''
                    worksheet['AT' + str(wsrow)] = ''
                    worksheet['AU' + str(wsrow)] = ''
                    worksheet['AV' + str(wsrow)] = ''
                    worksheet['AW' + str(wsrow)] = ''
                    worksheet['AX' + str(wsrow)] = ''
                    worksheet['AY' + str(wsrow)] = ''
                    worksheet['AZ' + str(wsrow)] = row['former_incumbent']
                    worksheet['BA' + str(wsrow)] = row['mode_sep__name']
                    worksheet['BB' + str(wsrow)] = row['date_vacated']
                    worksheet['BC' + str(wsrow)] = row['remarks_vacated']
                    wsrow = wsrow + 1
            response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Staffing by divisions exporting' + ' (' + year + ').xlsx'
            return response
    else:
        data = Empprofile.objects.all().filter(~Q(fundsource__name__icontains='Savings')).order_by('pi__user__last_name').values('section_id',
                                                                               'pi__user__last_name',
                                                                               'pi__user__first_name',
                                                                               'pi__user__middle_name',
                                                                               'pi__ext__name',
                                                                               'pi__gender',
                                                                               'salary_grade',
                                                                               'salary_rate',
                                                                               'empstatus__acronym',
                                                                               'position__name',
                                                                               'fundsource__name',
                                                                               'pi__cs_id',
                                                                               'pi__dob',
                                                                               'aoa_id',
                                                                               'pi_id',
                                                                               'pi__mobile_no',
                                                                               'pi__user__email',
                                                                               'pi__isfilipino',
                                                                               'pi__user__is_active',
                                                                               'item_number',
                                                                               'former_incumbent',
                                                                               'mode_sep__name',
                                                                               'mode_access__name',
                                                                               'date_vacated',
                                                                               'dateofcreation_pos',
                                                                               'designation',
                                                                               'dateof_designation',
                                                                               'dateof_orig_appointment',
                                                                               'dateof_last_promotion',
                                                                               'remarks_vacated',
                                                                               'pi__type_of_disability',
                                                                               'pi__type_of_indi__name',
                                                                               'specialorder_no',
                                                                               'plantilla_psipop')

        if data:
            for indx, row in enumerate(data):
                if row['pi__user__is_active'] == 1:
                    lastname = row['pi__user__last_name']
                    middlename = row['pi__user__middle_name']
                    firstname = row['pi__user__first_name']
                    ext = row['pi__ext__name']
                    worksheet['A' + str(wsrow)] = get_division(row['section_id'])
                    worksheet['B' + str(wsrow)] = '' if get_aoa(row['aoa_id']) else get_section(row['section_id'])
                    worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                    worksheet['D' + str(wsrow)] = row['item_number']
                    worksheet['E' + str(wsrow)] = row['dateofcreation_pos']
                    worksheet['F' + str(wsrow)] = row['position__name']
                    worksheet['G' + str(wsrow)] = ''
                    worksheet['H' + str(wsrow)] = ''
                    worksheet['I' + str(wsrow)] = row['salary_grade']
                    worksheet['J' + str(wsrow)] = row['salary_rate']
                    worksheet['K' + str(wsrow)] = row['designation']
                    worksheet['L' + str(wsrow)] = row['dateof_designation']
                    worksheet['M' + str(wsrow)] = row['specialorder_no']
                    worksheet['N' + str(wsrow)] = row['plantilla_psipop']
                    worksheet['O' + str(wsrow)] = row['fundsource__name']
                    worksheet['P' + str(wsrow)] = row['empstatus__acronym']
                    worksheet['Q' + str(wsrow)] = 'Filled'
                    worksheet['R' + str(wsrow)] = row['mode_access__name']
                    worksheet['S' + str(wsrow)] = filledup(row['pi_id'])
                    worksheet['T' + str(wsrow)] = "{}, {} {} {}.".format(str(lastname.upper()), str(firstname.upper()), str(middlename.upper()), "" if ext is None else ext)
                    worksheet['U' + str(wsrow)] = str(lastname.upper())
                    worksheet['V' + str(wsrow)] = str(firstname.upper())
                    worksheet['W' + str(wsrow)] = str(middlename.upper())
                    worksheet['X' + str(wsrow)] = ext
                    worksheet['Y' + str(wsrow)] = row['dateof_orig_appointment']
                    worksheet['Z' + str(wsrow)] = row['dateof_last_promotion']
                    worksheet['AA' + str(wsrow)] = entry_first_indswd(row['pi_id'])
                    worksheet['AB' + str(wsrow)] = get_civilservice(row['pi_id'])
                    worksheet['AC' + str(wsrow)] = get_ra(row['pi_id'])
                    worksheet['AD' + str(wsrow)] = get_let(row['pi_id'])
                    worksheet['AE' + str(wsrow)] = getlevelof_elig(row['pi_id'])
                    worksheet['AF' + str(wsrow)] = geteducation_status(row['pi_id'])
                    worksheet['AG' + str(wsrow)] = get_firstdegree(row['pi_id'])
                    worksheet['AH' + str(wsrow)] = get_lastdegree(row['pi_id'])
                    worksheet['AI' + str(wsrow)] = get_otherdegree(row['pi_id'])
                    worksheet['AJ' + str(wsrow)] = get_masters(row['pi_id'])
                    worksheet['AK' + str(wsrow)] = 'Male' if row['pi__gender'] == 1 else 'Female'
                    worksheet['AL' + str(wsrow)] = row['pi__dob']
                    worksheet['AM' + str(wsrow)] = get_age(row['pi__dob'])
                    worksheet['AN' + str(wsrow)] = get_civil_status(row['pi__cs_id'])
                    worksheet['AO' + str(wsrow)] = get_residentialadd(row['pi_id'])
                    worksheet['AP' + str(wsrow)] = get_permanentadd(row['pi_id'])
                    worksheet['AQ' + str(wsrow)] = get_solo_parent(row['pi_id'])
                    worksheet['AR' + str(wsrow)] = get_senior(row['pi__dob'])
                    worksheet['AS' + str(wsrow)] = 'PWD' if row['pi__type_of_disability'] else ''
                    worksheet['AT' + str(wsrow)] = row['pi__type_of_disability']
                    worksheet['AU' + str(wsrow)] = 'INDIGINOUS' if row['pi__type_of_indi__name'] else ''
                    worksheet['AV' + str(wsrow)] = row['pi__type_of_indi__name']
                    worksheet['AW' + str(wsrow)] = 'Filipino' if row['pi__isfilipino'] == 1 else ""
                    worksheet['AX' + str(wsrow)] = row['pi__mobile_no']
                    worksheet['AY' + str(wsrow)] = row['pi__user__email']
                    worksheet['AZ' + str(wsrow)] = row['former_incumbent']
                    worksheet['BA' + str(wsrow)] = row['mode_sep__name']
                    worksheet['BB' + str(wsrow)] = row['date_vacated']
                    worksheet['BC' + str(wsrow)] = row['remarks_vacated']
                    wsrow = wsrow + 1
                else:
                    worksheet['A' + str(wsrow)] = get_division(row['section_id'])
                    worksheet['B' + str(wsrow)] = '' if get_aoa(row['aoa_id']) else get_section(row['section_id'])
                    worksheet['C' + str(wsrow)] = get_aoa(row['aoa_id'])
                    worksheet['D' + str(wsrow)] = row['item_number']
                    worksheet['E' + str(wsrow)] = row['dateofcreation_pos']
                    worksheet['F' + str(wsrow)] = row['position__name']
                    worksheet['G' + str(wsrow)] = ''
                    worksheet['H' + str(wsrow)] = ''
                    worksheet['I' + str(wsrow)] = row['salary_grade']
                    worksheet['J' + str(wsrow)] = row['salary_rate']
                    worksheet['K' + str(wsrow)] = row['designation']
                    worksheet['L' + str(wsrow)] = row['dateof_designation']
                    worksheet['M' + str(wsrow)] = ''
                    worksheet['N' + str(wsrow)] = ''
                    worksheet['O' + str(wsrow)] = row['fundsource__name']
                    worksheet['P' + str(wsrow)] = row['empstatus__acronym']
                    worksheet['Q' + str(wsrow)] = 'Unfilled'
                    worksheet['R' + str(wsrow)] = ''
                    worksheet['S' + str(wsrow)] = ''
                    worksheet['T' + str(wsrow)] = ''
                    worksheet['U' + str(wsrow)] = ''
                    worksheet['V' + str(wsrow)] = ''
                    worksheet['W' + str(wsrow)] = ''
                    worksheet['X' + str(wsrow)] = ''
                    worksheet['Y' + str(wsrow)] = ''
                    worksheet['Z' + str(wsrow)] = ''
                    worksheet['AA' + str(wsrow)] = ''
                    worksheet['AB' + str(wsrow)] = ''
                    worksheet['AC' + str(wsrow)] = ''
                    worksheet['AD' + str(wsrow)] = ''
                    worksheet['AE' + str(wsrow)] = ''
                    worksheet['AF' + str(wsrow)] = ''
                    worksheet['AG' + str(wsrow)] = ''
                    worksheet['AH' + str(wsrow)] = ''
                    worksheet['AI' + str(wsrow)] = ''
                    worksheet['AJ' + str(wsrow)] = ''
                    worksheet['AK' + str(wsrow)] = ''
                    worksheet['AL' + str(wsrow)] = ''
                    worksheet['AM' + str(wsrow)] = ''
                    worksheet['AN' + str(wsrow)] = ''
                    worksheet['AO' + str(wsrow)] = ''
                    worksheet['AP' + str(wsrow)] = ''
                    worksheet['AQ' + str(wsrow)] = ''
                    worksheet['AR' + str(wsrow)] = ''
                    worksheet['AS' + str(wsrow)] = ''
                    worksheet['AT' + str(wsrow)] = ''
                    worksheet['AU' + str(wsrow)] = ''
                    worksheet['AV' + str(wsrow)] = ''
                    worksheet['AW' + str(wsrow)] = ''
                    worksheet['AX' + str(wsrow)] = ''
                    worksheet['AY' + str(wsrow)] = ''
                    worksheet['AZ' + str(wsrow)] = row['former_incumbent']
                    worksheet['BA' + str(wsrow)] = row['mode_sep__name']
                    worksheet['BB' + str(wsrow)] = row['date_vacated']
                    worksheet['BC' + str(wsrow)] = row['remarks_vacated']
                    wsrow = wsrow + 1

            response = HttpResponse(save_virtual_workbook(workbook), content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Staffing Export Data in' + ' (' + year + ').xlsx'
            return response


@login_required
@permission_required("auth.employee_list")
def view_pds_pageone(request, pk):
    pi = Personalinfo.objects.filter(id=pk).first()
    context = {
        'pds': Personalinfo.objects.filter(id=pk).first(),
        'address': Address.objects.filter(pi_id=pi.id).first(),
        'numbers': Deductionnumbers.objects.filter(pi_id=pi.id).first(),
        'family': Familybackground.objects.filter(pi_id=pi.id).first(),
        'eb': Educationbackground.objects.filter(pi_id=pi.id).order_by('level_id')
    }
    return render(request, 'backend/employee_data/staff/view_pds_pageone.html', context)


@login_required
@permission_required("auth.employee_list")
def view_pds_pagetwo(request, pk):
    context = {
        'civil': Civilservice.objects.filter(pi_id=pk),
        'work': Workexperience.objects.filter(pi_id=pk).order_by('-we_to'),
        'pk': pk
    }
    return render(request, 'backend/employee_data/staff/view_pds_pagetwo.html', context)


@login_required
@permission_required("auth.employee_list")
def view_pds_pagethree(request, pk):
    context = {
        'voluntary': Voluntary.objects.filter(pi_id=pk),
        'training': Training.objects.filter(pi_id=pk).order_by('-tr_to')[0:21],
        'skills': Skills.objects.filter(pi_id=pk),
        'nonacad': Recognition.objects.filter(pi_id=pk),
        'membership': Membership.objects.filter(pi_id=pk),
        'pk': pk
    }
    return render(request, 'backend/employee_data/staff/view_pds_pagethree.html', context)


@login_required
@permission_required("auth.employee_list")
def view_pds_pagefour(request, pk):
    additional = Additional.objects.filter(pi_id=pk)
    data = [dict(question=row.question, answers=row.answers) for row in additional]
    context = {
        'reference': Reference.objects.filter(pi_id=pk),
        'additional': data,
        'date': datetime.now(),
        'employee': Empprofile.objects.filter(pi_id=pk).first(),
        'pk': pk
    }
    return render(request, 'backend/employee_data/staff/view_pds_pagefour.html', context)


@login_required
@permission_required("auth.employee_list")
def view_pds_pagefive(request, pk):
    date_today = date.today()
    fiveyear_interval_date = datetime.now() - timedelta(days=5 * 365)
    context = {
        'work': Workexperience.objects.filter(pi_id=pk).order_by('-we_to')[30:],
        'training': Training.objects.filter(pi_id=pk, tr_to__range=[fiveyear_interval_date.strftime("%Y-%m-%d"),
                                                                    date_today]).order_by('-tr_to')[21:],
        'date': datetime.now(),
        'pk': pk
    }
    return render(request, 'backend/employee_data/staff/view_pds_pagefive.html', context)


@csrf_exempt
@login_required
def show_children(request, pk):
    child = Children.objects.all().filter(pi_id=pk)
    data = [dict(child_fullname=row.child_fullname, child_dob=row.child_dob) for row in child]
    return JsonResponse({'data': data})


@csrf_exempt
@permission_required("auth.employee_list")
def remove_children(request, pk):
    Children.objects.filter(pi_id=pk, id=request.POST.get('id')).delete()
    return JsonResponse({'data': 'success'})


@login_required
@permission_required("auth.employee_list")
def view_profile(request, pk):
    obj = get_object_or_404(Empprofile, pk=pk)
    upload_form = UploadPictureForm(instance=obj)

    context = {
        'employee': Empprofile.objects.filter(id=pk).first(),
        'education': Educationbackground.objects.filter(Q(pi_id=pk) & Q(level_id=3)).first(),
        'skills': Skills.objects.filter(Q(pi_id=pk)),
        'upload_form': upload_form,
    }
    return render(request, 'backend/employee_data/staff/view_profile.html', context)


@login_required
@permission_required("auth.edit_employee")
def edit_employee(request, id):
    if request.method == "POST":
        employee = Empprofile.objects.filter(id=id).first()

        Empprofile.objects.filter(id=id).update(
            id_number=request.POST.get('id_number')
        )

        Personalinfo.objects.filter(id=employee.pi_id).update(
            ext_id=request.POST.get('ext_id'),
            mobile_no=request.POST.get('contact_number')
        )

        AuthUser.objects.filter(id=employee.pi.user_id).update(
            first_name=request.POST.get('first_name'),
            last_name=request.POST.get('last_name'),
            middle_name=request.POST.get('middle_name'),
            username=request.POST.get('username'),
            email=request.POST.get('email'),
            is_active=1 if request.POST.get('is_active') else 0
        )

        return JsonResponse({'data': 'success', 'msg': 'You have successfully updated the employee information.'})

    context = {
        'employee': Empprofile.objects.filter(id=id).first(),
        'extension_name': ExtensionName.objects.filter(status=1)
    }
    return render(request, 'backend/employee_data/staff/edit_employee.html', context)


@login_required
@permission_required("admin.superadmin")
def view_ad_account(request, pk):
    context = {
        'employee': Empprofile.objects.filter(id=pk).first(),
    }
    return render(request, 'backend/employee_data/staff/view_ad_account.html', context)


# @login_required
# @permission_required("auth.employee_list")
# def activities(request, pk):
#     emp = Empprofile.objects.filter(pi_id=pk).first()
#     context = {
#         'travel_history': Ritopeople.objects.filter(name_id=emp.id, detail__rito__status=3).order_by(
#             '-detail__inclusive_to')[:3],
#         'last_update': PdsUpdateTracking.objects.filter(pi_id=pk),
#         'pis_config': PisConfig.objects.all(),
#         'employee': emp
#     }
#     return render(request, 'backend/employee_data/staff/activities.html', context)



@login_required
@permission_required("auth.employee_list")
def activities(request, pk):
    emp = Empprofile.objects.filter(pi_id=pk).first()

    last_update_dict = {
        update.pis_config_id: update.date for update in PdsUpdateTracking.objects.filter(pi_id=pk)
    }

    pis_config_list = PisConfig.objects.all()
    for config in pis_config_list:
        config.last_update_date = last_update_dict.get(config.id, None) 

    context = {
        'travel_history': Ritopeople.objects.filter(name_id=emp.id, detail__rito__status=3).order_by(
            '-detail__inclusive_to')[:3],
        'pis_config': pis_config_list,
        'employee': emp
    }
    return render(request, 'backend/employee_data/staff/activities.html', context)


# @login_required
# @permission_required("auth.edit_employee")
# def validate_pds(request):
#     check = PdsUpdateTracking.objects.filter(pi_id=request.POST.get('pi_id'))

#     config = PortalConfiguration.objects.filter(key_name='PDS Days Validation').first()

#     valid = True
#     empty_fields = []
#     days_before = []
#     for row in check:
#         today = datetime.today()
#         if row.date:
#             updated_date = row.date
#             days_ago = updated_date - today
#             if abs(days_ago.days) > int(config.key_acronym):
#                 valid = False
#                 days_before.append(row.pis_config.name)
#         else:
#             valid = False
#             empty_fields.append(row.pis_config.name)

#     if valid:
#         msg = 'Good day {}! You have successfully updated your ePDS. Please present this notification message to the Personnel Administration Section for clearance purposes. Thank you!'.format(check.first().pi.user.first_name.title())
#         if check.first().pi.mobile_no:
#             send_notification(msg, check.first().pi.mobile_no, request.session['emp_id'], check.first().id)
#         return JsonResponse({'data': 'success', 'msg': msg})
#     else:
#         empty_fields_msg = "<ul style='list-style-type:none;'>"
#         for row in empty_fields:
#             empty_fields_msg += "<li>{}</li>".format(str(row))

#         for row in days_before:
#             empty_fields_msg += "<li>{}</li>".format(str(row))

#         empty_fields_msg += "</ul>"

#         return JsonResponse({'error': True, 'msg': '<strong>ePDS was not updated in the last {} days </strong><br><br>{}'\
#                             .format(int(config.key_acronym), empty_fields_msg)})




@login_required
@permission_required("auth.edit_employee")
def validate_pds(request):
    check = PdsUpdateTracking.objects.filter(pi_id=request.POST.get('pi_id'))

    config = PortalConfiguration.objects.filter(key_name='PDS Days Validation').first()
    
    if not config or not config.key_acronym.isdigit():  
        return JsonResponse({'error': True, 'msg': "Invalid configuration for PDS validation."})
    
    valid = True
    empty_fields = []
    days_before = []
    
    today = datetime.now(timezone.utc)  # Ensure timezone-aware datetime

    for row in check:
        if row.date:
            updated_date = row.date
            days_ago = (today - updated_date).days  # Correct the subtraction order
            
            if days_ago > int(config.key_acronym):  
                valid = False
                days_before.append(row.pis_config.name)
        else:
            valid = False
            empty_fields.append(row.pis_config.name)

    if valid:
        msg = 'Good day {}! You have successfully updated your ePDS. Please present this notification message to the Personnel Administration Section for clearance purposes. Thank you!'.format(check.first().pi.user.first_name.title())
        if check.first().pi.mobile_no:
            send_notification(msg, check.first().pi.mobile_no, request.session['emp_id'], check.first().id)
        return JsonResponse({'data': 'success', 'msg': msg})
    else:
        empty_fields_msg = "<ul style='list-style-type:none;'>"
        for row in empty_fields + days_before:  
            empty_fields_msg += "<li>{}</li>".format(str(row))
        empty_fields_msg += "</ul>"

        return JsonResponse({'error': True, 'msg': '<strong>ePDS was not updated in the last {} days </strong><br><br>{}'.format(int(config.key_acronym), empty_fields_msg)})



@login_required
@permission_required("auth.travel_history")
def backend_travel_history(request, pk):
    page = request.GET.get('page', 1)
    rows = request.GET.get('rows', 20)
    context = {
        'travel_history': Paginator(Ritopeople.objects.filter(name_id=pk, detail__rito__status=3).order_by(
            '-detail__inclusive_to'), rows).page(page),
        'employee': Empprofile.objects.filter(id=pk).first(),
        'title': 'employee',
        'sub_title': 'all_employee'
    }
    return render(request, 'backend/employee_data/staff/travel_history.html', context)


def generate_itemnumber(request,pk):
    if request.method == "GET":
        emp = Empprofile.objects.filter(id=pk).annotate(
        fullname=Concat(Value('FOCARAGA-'), Upper('empstatus__acronym'), Value('-'))
        ).values_list('fullname', flat=True)
        results = list(emp)
        data = json.dumps(results)
        return HttpResponse(data, 'application/json')


@login_required
@permission_required("auth.employee_list")

def about(request, pk, pi_id):
    if request.method == "POST":
        with transaction.atomic():
            Empprofile.objects.filter(id=pk).update(
                item_number=request.POST.get('item_number'),
                account_number=request.POST.get('account_number'),
                position_id=request.POST.get('position'),
                empstatus_id=request.POST.get('empstatus'),
                project_id=request.POST.get('project'),
                section_id=request.POST.get('section'),
                fundsource_id=request.POST.get('fund_source'),
                aoa_id=request.POST.get('aoa'),
                salary_rate=request.POST.get('salary_rate'),
                salary_grade=request.POST.get('salary_grade'),
                step_inc=request.POST.get('step_inc'),
                dateofcreation_pos=request.POST.get('docp') if request.POST.get('docp') else None,
                designation=request.POST.get('designation'),
                dateof_designation=request.POST.get('dod') if request.POST.get('dod') else None,
                mode_access_id=request.POST.get('mode_access'),
                specialorder_no=request.POST.get('special_order_no'),
                plantilla_psipop=request.POST.get('plantilla'),
                former_incumbent=request.POST.get('former_incumbent'),
                mode_sep_id=request.POST.get('mode_sep'),
                date_vacated=request.POST.get('date_vacated') if request.POST.get('date_vacated') else None,
                remarks_vacated=request.POST.get('remarks_vacated')
            )

            Personalinfo.objects.filter(id=pi_id).update(
                type_of_disability=request.POST.get('tod'),
                type_of_indi_id=request.POST.get('tig')
            )

            return JsonResponse({'data': 'success'})

    context = {
        'employee': Empprofile.objects.filter(id=pk).first(),
        'division': Division.objects.all(),
        'section': Section.objects.all(),
        'fundsource': Fundsource.objects.filter(status=1).order_by('name'),
        'aoa': Aoa.objects.filter(status=1).order_by('name'),
        'project': Project.objects.filter(status=1).order_by('name'),
        'position': Position.objects.filter(status=1).order_by('name'),
        'empstatus': Empstatus.objects.filter(status=1).order_by('name'),
        'mode_separation': HrppmsModeseparation.objects.filter(status=1).order_by('name'),
        'mode_accession': HrppmsModeaccession.objects.filter(status=1).order_by('name'),
        'sg': Salarygrade.objects.all(),
        'stepinc': Stepinc.objects.all(),
        'personal_info': Personalinfo.objects.filter(id=pi_id).first(),
        'employee': Empprofile.objects.filter(id=pk).first(),
        'indi': Indiginous.objects.all().order_by('name'),
        'religion': Religion.objects.filter(status=1).order_by('name'),
        'ethnicity': Ethnicity.objects.filter(status=1).order_by('name')
    }
    return render(request, 'backend/employee_data/staff/about.html', context)


@login_required
@permission_required("auth.employee_list")
def settings(request, pk):
    context = {
        'dtr': DtrPin.objects.filter(emp_id=pk).first(),
        'employee': Empprofile.objects.filter(id=pk).first(),
        'auth_permission': AuthPermission.objects.all().order_by('name'),
        'your_payroll_incharge': PasEmpPayrollIncharge.objects.filter(emp_id=pk).first(),
        'payroll_incharge': PayrollIncharge.objects.filter(status=1).order_by('id')
    }
    return render(request, 'backend/employee_data/staff/settings.html', context)


@login_required
@permission_required("auth.edit_employee")
def add_payroll_incharge(request):
    if request.method == "POST":
        check = PasEmpPayrollIncharge.objects.filter(emp_id=request.POST.get('employee_pk'))
        if not check:
            PasEmpPayrollIncharge.objects.create(
                emp_id=request.POST.get('employee_pk'),
                payroll_incharge_id=request.POST.get('payroll_incharge')
            )
        else:
            check.update(
                payroll_incharge_id=request.POST.get('payroll_incharge')
            )

        return JsonResponse({'data': 'success', 'msg': 'You have successfully assigned employee to a payroll incharge'})


@login_required
@permission_required("auth.dtr_logs")
def registered_dtr_pin(request):
    if request.method == "POST":
        check = DtrPin.objects.filter(emp_id=request.POST.get('emp_id'))
        if check:
            check.update(
                pin_id=request.POST.get('serial_number')
            )
        else:
            DtrPin.objects.create(
                emp_id=request.POST.get('emp_id'),
                pin_id=request.POST.get('serial_number')
            )

        return JsonResponse({'data': 'success'})


@login_required
@permission_required("admin.superadmin")
def permission(request):
    perm = request.POST.getlist('permission[]')
    data = [str(row.permission_id) for row in
            AuthUserUserPermissions.objects.filter(user_id=request.POST.get('user_id'))]
    universe = [str(row.id) for row in AuthPermission.objects.all()]
    comparison = list(set(perm) - set(data))
    comparison_two = list(set(list(set(universe) - set(comparison))) - set(perm))

    for row in comparison:
        AuthUserUserPermissions.objects.create(permission_id=row, user_id=request.POST.get('user_id'))

    for row in comparison_two:
        AuthUserUserPermissions.objects.filter(permission_id=row, user_id=request.POST.get('user_id')).delete()

    return JsonResponse({'data': 'success'})


# @login_required
# @permission_required("auth.employee_list")
# def work_history(request, pi_id):
#     employee = Empprofile.objects.filter(pi_id=pi_id).first()
#     data = Workexperience.objects.filter(Q(pi_id=pi_id) & (
#             Q(company__icontains='DSWD') | Q(
#         company__icontains='Department of Social Welfare and Development')))
#     for row in data:
#         position = Position.objects.filter(name=row.position_name)
#         check = Workhistory.objects.filter(we_id=row.id).first()
#         if check is None:
#             Workhistory.objects.create(
#                 emp_id=employee.id,
#                 we_id=row.id,
#             )

#             Workexperience.objects.filter(id=row.id).update(
#                 position_id=position.first().id if position else None
#             )
#         else:
#             Workexperience.objects.filter(id=row.id).update(
#                 position_id=position.first().id if position else None
#             )

#     context = {
#         'employee_id': employee.id,
#         'work_experience': Workexperience.objects.filter(Q(pi_id=pi_id) & (Q(company__icontains='DSWD') | Q(
#             company__icontains='Department of Social Welfare and Development'))).order_by('-we_to')
#     }
#     return render(request, 'backend/employee_data/staff/workhistory.html', context)


@login_required
@permission_required("auth.employee_list")
def work_history(request, pi_id):
    employee = Empprofile.objects.filter(pi_id=pi_id).first()
    data = Workexperience.objects.filter(Q(pi_id=pi_id) & (
            Q(company__icontains='DSWD') | Q(
        company__icontains='Department of Social Welfare and Development')))
    for row in data:
        position = Position.objects.filter(name__icontains=row.position_name)
        check = Workhistory.objects.filter(we_id=row.id).first()
        if check is None:
            Workhistory.objects.create(
                emp_id=employee.id,
                we_id=row.id,
            )

            Workexperience.objects.filter(id=row.id).update(
                position_id=position.first().id if position else None
            )
        else:
            Workexperience.objects.filter(id=row.id).update(
                position_id=position.first().id if position else None
            )

    context = {
        'employee_id': employee.id,
        'work_experience': Workexperience.objects.filter(Q(pi_id=pi_id) & (Q(company__icontains='DSWD') | Q(
            company__icontains='Department of Social Welfare and Development'))).order_by('-we_to')
    }
    return render(request, 'backend/employee_data/staff/workhistory.html', context)



@login_required
@permission_required("auth.employee_list")
def edit_workhistory(request, pk):
    if request.method == "POST":
        check = Workhistory.objects.filter(we_id=request.POST.get('we_id')).first()
        if check:
            Workhistory.objects.filter(we_id=request.POST.get('we_id')).update(
                aoa_id=request.POST.get('aoa'),
                fundsource_id=request.POST.get('fundsource'),
                project_id=request.POST.get('project'),
                section_id=request.POST.get('section'),
                emp_id=request.POST.get('emp_id'),
                we_id=request.POST.get('we_id'),
                item_number=request.POST.get('item_number'),
                former_incumbent=request.POST.get('former_incumbent'),
                date_vacated=request.POST.get('date_vacated') if request.POST.get('date_vacated') else None,
                datecreation_pos=request.POST.get('dateofcreation') if request.POST.get('dateofcreation') else None,
                dateof_orig_appointment=request.POST.get('dateoforig') if request.POST.get('dateoforig') else None,
                dateof_last_promotion=request.POST.get('dateoflast') if request.POST.get('dateoflast') else None,
                mode_access_id=request.POST.get('modeaccess'),
                mode_sep_id=request.POST.get('modesep'),
                specialorder_no=request.POST.get('specialorder'),
                plantilla_psipop=request.POST.get('plantilla'),
                remarks_vacated=request.POST.get('remarks_vacated'),
                designation=request.POST.get('designation'),
                dateof_designation=request.POST.get('dateof_designation') if request.POST.get('dateof_designation') else None,
            )

            Workexperience.objects.filter(id=request.POST.get('we_id')).update(
                we_from=request.POST.get('date_of_appointment'),
                we_to=request.POST.get('end_of_appointment'),
                salary_rate=request.POST.get('salary_rate'),
                position_id=request.POST.get('position')
            )
        else:
            Workhistory.objects.create(
                aoa_id=request.POST.get('aoa'),
                fundsource_id=request.POST.get('fundsource'),
                project_id=request.POST.get('project'),
                section_id=request.POST.get('section'),
                emp_id=request.POST.get('emp_id'),
                we_id=request.POST.get('we_id'),
                item_number=request.POST.get('item_number'),
                former_incumbent=request.POST.get('former_incumbent'),
                date_vacated=request.POST.get('date_vacated'),
                datecreation_pos=request.POST.get('dateofcreation'),
                dateof_orig_appointment=request.POST.get('dateoforig'),
                dateof_last_promotion=request.POST.get('dateoflast'),
                mode_access_id=request.POST.get('modeaccess'),
                mode_sep_id=request.POST.get('modesep'),
                specialorder_no=request.POST.get('specialorder'),
                plantilla_psipop=request.POST.get('plantilla'),
                remarks_vacated=request.POST.get('remarks_vacated'),
                designation=request.POST.get('designation'),
                dateof_designation=request.POST.get('dateof_designation'),
            )

        return JsonResponse({'data': 'success'})

    work_experience = Workexperience.objects.filter(id=pk).first()
    emp = Empprofile.objects.filter(pi_id=work_experience.pi_id).first()
    context = {
        'division': Division.objects.all(),
        'section': Section.objects.all(),
        'fundsource': Fundsource.objects.all().order_by('name'),
        'aoa': Aoa.objects.all().order_by('name'),
        'project': Project.objects.all().order_by('name'),
        'position': Position.objects.all().order_by('name'),
        'empstatus': Empstatus.objects.all().order_by('name'),
        'work_experience': work_experience,
        'employee_id': emp.id,
        'work_history': Workhistory.objects.filter(we_id=pk).first(),
        'access': HrppmsModeaccession.objects.all().order_by('name'),
        'sep': HrppmsModeseparation.objects.all().order_by('name'),
    }
    return render(request, 'backend/employee_data/staff/edit_workhistory.html', context)




@csrf_exempt
@login_required
@permission_required("auth.employee_list")
def verify_work_experience(request, we_id):
    if request.method == "POST":
        w_experience = Workexperience.objects.filter(id=we_id).first()
        if w_experience:
            w_experience.status = True
            w_experience.save()
            print(f'Work experience {we_id} status: {w_experience.status}')

            return JsonResponse({'status': 'status', 'status': w_experience.status})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


@csrf_exempt
@login_required
@permission_required("auth.employee_list")
def verify_work_experience_group(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            verified_count = 0

            for we_id in ids:
                work_experience = Workexperience.objects.filter(id=we_id).first()
                if work_experience and not work_experience.status:
                    work_experience.status = True  
                    work_experience.save()
                    verified_count += 1

            return JsonResponse({"verified_count": verified_count}, status=200)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)
        except Exception as e:
            print(f"Error: {str(e)}")  
            return JsonResponse({"error": str(e)}, status=400)
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@login_required
@permission_required("auth.employee_list")
def unverify_work_experience(request, we_id):
    if request.method == "POST":
        w_experience = Workexperience.objects.filter(id=we_id).first()
        if w_experience:
            w_experience.status = False
            w_experience.save()
            print(f'Work experience {we_id} status: {w_experience.status}')

            return JsonResponse({'status': 'status', 'status': w_experience.status})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


@csrf_exempt
@login_required
@permission_required("auth.employee_list")
def unverify_work_experience_group(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            unverified_count = 0

            for we_id in ids:
                work_experience = Workexperience.objects.filter(id=we_id).first()
                if work_experience and work_experience.status:
                    work_experience.status = False
                    work_experience.save()
                    unverified_count += 1

            return JsonResponse({"unverified_count": unverified_count}, status=200)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)
        except Exception as e:
            print(f"Error: {str(e)}")
            return JsonResponse({"error": str(e)}, status=400)
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@login_required
@permission_required("auth.employee_list")
def delete_workexperience_admin(request):
    Workexperience.objects.filter(id=request.POST.get('id')).delete()
    return JsonResponse({'data': 'success'})


@csrf_exempt
@login_required
@permission_required("auth.employee_list")
def job_movement(request, pi_id):
    employee = get_object_or_404(Empprofile.objects.select_related('position', 'empstatus', 'section'), pi_id=pi_id)
    positions = Position.objects.all().order_by('name')
    employment_statuses = Empstatus.objects.all()
    stepinc = Stepinc.objects.all()
    fundsources = Fundsource.objects.all()  
    aoa = Aoa.objects.all()
    sections = Section.objects.all().order_by('sec_name') 

    latest_experience = Workexperience.objects.filter(pi=employee.pi, we_to__isnull=True).order_by('-we_from').first()

    if request.method == "POST":
        new_position_id = request.POST.get("new_position_id")
        emp_status_id = request.POST.get("emp_status")  
        govt_service = request.POST.get("govt_service")
        we_from = request.POST.get("we_from")  
        salary_rate = request.POST.get("salary_rate")
        sg_step = request.POST.get("sg_step")
        reason = request.POST.get("reason")
        step_inc_id = request.POST.get("step_inc")
        fundsource_id = request.POST.get("fundsource")  
        aoa_id = request.POST.get("aoa")
        section_id = request.POST.get("section") 

        step_inc = Stepinc.objects.get(id=int(step_inc_id)) if step_inc_id else None
        fundsource = Fundsource.objects.get(id=int(fundsource_id)) if fundsource_id else None  
        aoa = Aoa.objects.get(id=int(aoa_id)) if aoa_id else None
        section = Section.objects.get(id=int(section_id)) if section_id else None  

        if not new_position_id or not emp_status_id or not we_from:
            return JsonResponse({"status": "error", "message": "Position, Employment Status, and Starting Date are required!"}, status=400)

        new_position = get_object_or_404(Position, id=new_position_id)
        emp_status = get_object_or_404(Empstatus, id=int(emp_status_id))  

        we_from_date = datetime.strptime(we_from, "%Y-%m-%d").date()
        we_to_date = we_from_date - timedelta(days=1)  

        if latest_experience:
            latest_experience.we_to = we_to_date  
            latest_experience.save() 

        work_experience = Workexperience.objects.create(
            pi=employee.pi,
            position=new_position,  
            position_name=new_position.name,  
            empstatus=emp_status,  
            govt_service=govt_service,  
            we_from=we_from_date,
            salary_rate=salary_rate,
            sg_step=sg_step,
            company="Department of Social Welfare and Development",
            reason=reason,
            step_inc=step_inc
        )

        Workhistory.objects.create(
            emp=employee,
            we=work_experience,
            fundsource=fundsource,  
            designation=new_position.name,
            dateof_designation=we_from_date
        )

        employee.position = new_position
        employee.empstatus = emp_status
        employee.fundsource = fundsource
        employee.salary_rate = salary_rate
        employee.salary_grade = sg_step
        employee.step_inc = step_inc.id if step_inc else None
        employee.aoa = aoa if aoa else None
        employee.section = section if section else None 
        employee.save()

        return JsonResponse({"success": True, "message": "Position, section, fund source, and status updated successfully!"})

    context = {
        'employee': employee,
        'stepinc': stepinc,
        'positions': positions,
        'aoa': Aoa.objects.filter(status=1).order_by('name'),
        'employment_statuses': employment_statuses,
        'latest_experience': latest_experience,
        'salary_grades': range(1, 34),
        'fundsources': fundsources,  
        'section': Section.objects.all().order_by('sec_name'),  
    }

    return render(request, 'backend/employee_data/staff/job_movement.html', context)



@login_required
@permission_required("auth.edit_employee")
def update_user(request, pk):
    user = searchSamAccountName(request.POST.get('username'))
    if user["status"]:
        AuthUser.objects.filter(id=pk).update(username=request.POST.get('username'), email=request.POST.get('email'))
        if request.POST.get('initials'):
            user["connection"].modify(user["userDN"], {
                                                        'employeeID': [(MODIFY_REPLACE, [request.POST.get('id_number')])],
                                                        'employeeNumber': [(MODIFY_REPLACE, [request.POST.get('id_number')])],
                                                        'givenName': [(MODIFY_REPLACE, [request.POST.get('first_name')])],
                                                        'sn': [(MODIFY_REPLACE, [request.POST.get('last_name')])],
                                                        'initials': [(MODIFY_REPLACE, [request.POST.get('initials')])],
                                                        'displayName': [(MODIFY_REPLACE, [request.POST.get('display_name')])],
                                                        'telephoneNumber': [
                                                           (MODIFY_REPLACE, [request.POST.get('telephone_number')])],
                                                        'mail': [(MODIFY_REPLACE, [request.POST.get('email')])],
                                                        'title': [(MODIFY_REPLACE, [request.POST.get('job_title')])],
                                                        'department': [(MODIFY_REPLACE, [request.POST.get('department')])],
                                                       })
        else:
            user["connection"].modify(user["userDN"],
                                      {
                                        'employeeID': [(MODIFY_REPLACE, [request.POST.get('id_number')])],
                                        'employeeNumber': [(MODIFY_REPLACE, [request.POST.get('id_number')])],
                                        'givenName': [(MODIFY_REPLACE, [request.POST.get('first_name')])],
                                        'sn': [(MODIFY_REPLACE, [request.POST.get('last_name')])],
                                        'displayName': [(MODIFY_REPLACE, [request.POST.get('display_name')])],
                                        'telephoneNumber': [
                                           (MODIFY_REPLACE, [request.POST.get('telephone_number')])],
                                        'mail': [(MODIFY_REPLACE, [request.POST.get('email')])],
                                        'title': [(MODIFY_REPLACE, [request.POST.get('job_title')])],
                                        'department': [(MODIFY_REPLACE, [request.POST.get('department')])],
                                       })
        return JsonResponse({'data': 'success'})
    else:
        return JsonResponse({'error': True, 'msg': "Failed to update credentials or username does not exist."})


@login_required
@permission_required("admin.superadmin")
def divisions(request):
    form = DivisionsForm(initial={'div_chief_id': None})
    if request.method == "POST":
        form = DivisionsForm(request.POST, initial={'div_chief_id': None})
        if form.is_valid():
            divi = form.save(commit=False)
            if request.POST.get('div_chief_id') == '':
                divi.div_chief_id = None
            else:
                div_chief_id = re.split('\[|\]', request.POST.get('div_chief_id'))
                name = Empprofile.objects.values('id').filter(id_number=div_chief_id[1]).first()
                divi.div_chief_id = name['id']
            messages.success(request, 'The division {} was added successfully.'.format(form.cleaned_data['div_name']))
            divi.save()
            return JsonResponse({'error': False})
        else:
            return JsonResponse({'error': True, 'errors': form.errors})
    page = request.GET.get('page', 1)
    rows = request.GET.get('rows', 20)
    search = request.GET.get('search', '')
    context = {
        'form': form,
        'divi': Paginator(Division.objects.filter(Q(div_name__icontains=search)).order_by('div_name'), rows).page(page),
        'rows': rows,
        'title': 'libraries',
        'sub_title': 'employees',
        'sub_sub_title': 'division'
    }
    return render(request, 'backend/pas/employee/divisions.html', context)


class DiviUpdate(LoginRequiredMixin, AjaxableResponseMixin, PermissionRequiredMixin, UpdateView):
    template_name = 'backend/pas/employee/divisions_update.html'
    model = Division
    form_class = DivisionsForm
    success_url = reverse_lazy('divisions')
    permission_required = 'admin.superadmin'

    def form_valid(self, form):
        if form.instance.div_chief_id == '':
            form.instance.div_chief_id = None
        else:
            div_chief_id = re.split('\[|\]', form.instance.div_chief_id)
            name = Empprofile.objects.values('id').filter(id_number=div_chief_id[1]).first()
            form.instance.div_chief_id = name['id']
        return super().form_valid(form)


@login_required
@permission_required("admin.superadmin")
def sections(request):
    form = SectionsForm(initial={'sec_head_id': None})
    if request.method == "POST":
        form = SectionsForm(request.POST, initial={'sec_head_id': None})
        if form.is_valid():
            divi = form.save(commit=False)
            if request.POST.get('sec_head_id') == '':
                divi.sec_head_id = None
            else:
                sec_head_id = re.split('\[|\]', request.POST.get('sec_head_id'))
                name = Empprofile.objects.values('id').filter(id_number=sec_head_id[1]).first()
                divi.sec_head_id = name['id']
            messages.success(request, 'The section {} was added successfully.'.format(form.cleaned_data['sec_name']))
            divi.save()
            return JsonResponse({'error': False})
        else:
            return JsonResponse({'error': True, 'errors': form.errors})

    page = request.GET.get('page', 1)
    rows = request.GET.get('rows', 20)
    search = request.GET.get('search', '')
    context = {
        'form': form,
        'divi': Paginator(Section.objects.filter(Q(sec_name__icontains=search)).order_by('sec_name'), rows).page(page),
        'rows': rows,
        'title': 'libraries',
        'sub_title': 'employees',
        'sub_sub_title': 'section'
    }
    return render(request, 'backend/pas/employee/sections.html', context)


class SecUpdate(LoginRequiredMixin, AjaxableResponseMixin, PermissionRequiredMixin, UpdateView):
    template_name = 'backend/pas/employee/sections_update.html'
    model = Section
    form_class = SectionsForm
    success_url = reverse_lazy('sections')
    permission_required = 'admin.superadmin'

    def form_valid(self, form):
        if form.instance.sec_head_id == '':
            form.instance.sec_head_id = None
        else:
            sec_head_id = re.split('\[|\]', form.instance.sec_head_id)
            name = Empprofile.objects.values('id').filter(id_number=sec_head_id[1]).first()
            form.instance.sec_head_id = name['id']
        return super().form_valid(form)


# PDS Administrator Updating
@login_required
@permission_required("auth.edit_employee")
def update_personal_information(request, pi_id):
    if request.method == "POST":
        AuthUser.objects.filter(id=request.POST.get('user_id')).update(
            first_name=request.POST.get('fname').upper(),
            last_name=request.POST.get('surname').upper(),
            middle_name=request.POST.get('mname').upper(),
            email=request.POST.get('email'))

        Personalinfo.objects.filter(id=pi_id).update(
            ext_id=request.POST.get('ext'),
            dob=request.POST.get('dob'),
            pob=request.POST.get('pob').upper(),
            gender=request.POST.get('gender'),
            cs_id=request.POST.get('cs'),
            height=request.POST.get('height').upper(),
            weight=request.POST.get('weight').upper(),
            bt_id=request.POST.get('bt'),
            isfilipino=request.POST.get('isfilipino'),
            country_id=request.POST.get('countries'),
            mobile_no=request.POST.get('mobile_no'),
            isdualcitizenship=1 if request.POST.get('isdualcit') == 'on' else 0,
            dc_bybirth=1 if request.POST.get('isbybirth') == 'on' else 0,
            dc_bynaturalization=1 if request.POST.get('isbynaturalization') == 'on' else 0,
            telephone_no=request.POST.get('telephone_no').upper())

        check_address = Address.objects.filter(pi_id=pi_id)

        if check_address:
            Address.objects.filter(pi_id=pi_id).update(
                ra_house_no=request.POST.get('ra_house_no').upper(),
                ra_street=request.POST.get('ra_street').upper(),
                ra_village=request.POST.get('ra_village').upper(),
                ra_prov_code=None if request.POST.get('ra_prov_code') == '' else request.POST.get('ra_prov_code'),
                ra_city=request.POST.get('ra_city'),
                ra_brgy=request.POST.get('ra_brgy'),
                ra_zipcode=request.POST.get('ra_zipcode').upper(),
                pa_house_no=request.POST.get('pa_house_no').upper(),
                pa_street=request.POST.get('pa_street').upper(),
                pa_village=request.POST.get('pa_village').upper(),
                pa_prov_code=request.POST.get('pa_prov_code'),
                pa_city=request.POST.get('pa_city'),
                pa_brgy=request.POST.get('pa_brgy'),
                pa_zipcode=request.POST.get('pa_zipcode').upper()
            )
        else:
            Address.objects.create(
                ra_house_no=request.POST.get('ra_house_no').upper(),
                ra_street=request.POST.get('ra_street').upper(),
                ra_village=request.POST.get('ra_village').upper(),
                ra_prov_code=None if request.POST.get('ra_prov_code') == '' else request.POST.get('ra_prov_code'),
                ra_city=request.POST.get('ra_city'),
                ra_brgy=request.POST.get('ra_brgy'),
                ra_zipcode=request.POST.get('ra_zipcode').upper(),
                pa_house_no=request.POST.get('pa_house_no').upper(),
                pa_street=request.POST.get('pa_street').upper(),
                pa_village=request.POST.get('pa_village').upper(),
                pa_prov_code=request.POST.get('pa_prov_code'),
                pa_city=request.POST.get('pa_city'),
                pa_brgy=request.POST.get('pa_brgy'),
                pa_zipcode=request.POST.get('pa_zipcode').upper(),
                pi_id=pi_id
            )

        check_num = Deductionnumbers.objects.filter(pi_id=pi_id)

        if check_num:
            Deductionnumbers.objects.filter(pi_id=pi_id).update(
                gsis_no=request.POST.get('gsis_no').upper(),
                pagibig_no=request.POST.get('pagibig_no').upper(),
                philhealth_no=request.POST.get('philhealth_no').upper(),
                sss_no=request.POST.get('sss_no').upper(),
                tin_no=request.POST.get('tin_no').upper())
        else:
            Deductionnumbers.objects.create(
                gsis_no=request.POST.get('gsis_no').upper(),
                pagibig_no=request.POST.get('pagibig_no').upper(),
                philhealth_no=request.POST.get('philhealth_no').upper(),
                sss_no=request.POST.get('sss_no').upper(),
                tin_no=request.POST.get('tin_no').upper(),
                pi_id=pi_id)

        check_upd_tracking = PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id = 1)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=1)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=1,
                pi_id=pi_id
            )

        return JsonResponse({'data': 'success'})

    address = Address.objects.filter(pi_id=pi_id).first()
    if address:
        context = {
            'pds': Personalinfo.objects.filter(id=pi_id).first(),
            'ext': ExtensionName.objects.all().order_by('name'),
            'prov': Province.objects.all().order_by('name'),
            'bt': Bloodtype.objects.all().order_by('name'),
            'countries': Countries.objects.all().order_by('name'),
            'cs': Civilstatus.objects.all().order_by('name'),
            'numbers': Deductionnumbers.objects.filter(pi_id=pi_id).first(),
            'pa_city': City.objects.filter(prov_code=address.pa_prov_code).order_by('name'),
            'ra_city': City.objects.filter(prov_code=address.ra_prov_code).order_by('name'),
            'pa_brgy': Brgy.objects.filter(city_code=address.pa_city).order_by('name'),
            'ra_brgy': Brgy.objects.filter(city_code=address.ra_city).order_by('name'),
            'address': Address.objects.filter(pi_id=pi_id).first(),
            'shortcuts': PisConfig.objects.all()
        }
    else:
        context = {
            'pds': Personalinfo.objects.filter(id=pi_id).first(),
            'ext': ExtensionName.objects.all().order_by('name'),
            'prov': Province.objects.all().order_by('name'),
            'bt': Bloodtype.objects.all().order_by('name'),
            'countries': Countries.objects.all().order_by('name'),
            'cs': Civilstatus.objects.all().order_by('name'),
            'numbers': Deductionnumbers.objects.filter(pi_id=pi_id).first(),
            'address': Address.objects.filter(pi_id=pi_id).first(),
            'shortcuts': PisConfig.objects.all()
        }
    return render(request, 'backend/employee_data/staff/pds/personal_information.html', context)


@csrf_exempt
@login_required
@permission_required("auth.edit_employee")
def update_family_background(request, pi_id):
    if request.method == 'POST':
        childrens = request.POST.getlist('children[]')
        birth = request.POST.getlist('birth[]')

        data = dict(zip(childrens, birth))
        count = Children.objects.filter(pi_id=pi_id)

        store = [row.id for row in count]
        if count:
            y = 1
            x = 0
            for row in data.items():
                if y > len(count):
                    Children.objects.create(
                        child_fullname=row[0],
                        child_dob=row[1],
                        pi_id=pi_id)
                else:
                    Children.objects.filter(id=store[x]).update(
                        child_fullname=row[0],
                        child_dob=row[1],
                        pi_id=pi_id)
                    y += 1
                    x += 1
        else:
            for row in data.items():
                Children.objects.create(
                    child_fullname=row[0],
                    child_dob=row[1],
                    pi_id=pi_id)

        check_family = Familybackground.objects.filter(pi_id=pi_id)
        if check_family:
            Familybackground.objects.filter(pi_id=pi_id).update(
                sp_surname=request.POST.get('sp_surname'),
                sp_fname=request.POST.get('sp_fname'),
                sp_mname=request.POST.get('sp_mname'),
                sp_ext_id=request.POST.get('sp_ext'),
                sp_occupation=request.POST.get('sp_occupation'),
                sp_employer=request.POST.get('sp_employer'),
                sp_business=request.POST.get('sp_business'),
                sp_telephone=request.POST.get('sp_telephone'),
                f_surname=request.POST.get('f_surname'),
                f_fname=request.POST.get('f_fname'),
                f_mname=request.POST.get('f_mname'),
                f_ext_id=request.POST.get('f_ext'),
                m_surname=request.POST.get('m_surname'),
                m_fname=request.POST.get('m_fname'),
                m_mname=request.POST.get('m_mname'))

        else:
            Familybackground.objects.create(
                sp_surname=request.POST.get('sp_surname'),
                sp_fname=request.POST.get('sp_fname'),
                sp_mname=request.POST.get('sp_mname'),
                sp_ext_id=request.POST.get('sp_ext'),
                sp_occupation=request.POST.get('sp_occupation'),
                sp_employer=request.POST.get('sp_employer'),
                sp_business=request.POST.get('sp_business'),
                sp_telephone=request.POST.get('sp_telephone'),
                f_surname=request.POST.get('f_surname'),
                f_fname=request.POST.get('f_fname'),
                f_mname=request.POST.get('f_mname'),
                f_ext_id=request.POST.get('f_ext'),
                m_surname=request.POST.get('m_surname'),
                m_fname=request.POST.get('m_fname'),
                m_mname=request.POST.get('m_mname'),
                pi_id=pi_id)

        exists = IncaseOfEmergency.objects.filter(pi_id=pi_id)
        if exists:
            exists.update(
                contact_name=request.POST.get('ioe').title() if request.POST.get(
                    'ioe') != "Others" else request.POST.get('ioe_others').title(),
                contact_number=request.POST.get('ioe_no'),
                is_others=0 if request.POST.get('ioe') != "Others" else 1
            )
        else:
            IncaseOfEmergency.objects.create(
                pi_id=pi_id,
                contact_name=request.POST.get('ioe').title() if request.POST.get(
                    'ioe') != "Others" else request.POST.get('ioe_others').title(),
                contact_number=request.POST.get('ioe_no'),
                is_others=0 if request.POST.get('ioe') != "Others" else 1
            )

        check_upd_tracking = PdsUpdateTracking.objects.filter(
            Q(pi_id=request.session['pi_id']) & Q(pis_config_id=2)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=2)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=2,
                pi_id=pi_id
            )
        return JsonResponse({'data': 'success'})

    context = {
        'ext': ExtensionName.objects.all().order_by('id'),
        'family': Familybackground.objects.filter(pi_id=pi_id).first(),
        'children': Children.objects.filter(pi_id=pi_id),
        'incase_of_emergency': IncaseOfEmergency.objects.filter(pi_id=pi_id).first(),
        'pi_id': pi_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, 'backend/employee_data/staff/pds/family_background.html', context)


@login_required
@permission_required("auth.edit_employee")
def update_educational_background(request, pi_id):
    if request.method == "POST":
        level = request.POST.getlist('level[]')
        school = request.POST.getlist('school[]')
        degree = request.POST.getlist('degree[]')
        period_from = request.POST.getlist('from[]')
        period_to = request.POST.getlist('to[]')
        unit = request.POST.getlist('unit[]')
        year = request.POST.getlist('year[]')
        hon = request.POST.getlist('hon[]')

        data = [
            {'level': l, 'school': s, 'degree': d, 'period_from': fr, 'period_to': to, 'unit': u, 'year': y, 'hon': h}
            for l, s, d, fr, to, u, y, h in zip(level, school, degree, period_from, period_to, unit, year, hon)]

        check = Educationbackground.objects.filter(pi_id=pi_id)

        store = [row.id for row in check]
        if check:
            y = 1
            x = 0
            for row in data:
                if y > len(check):
                    Educationbackground.objects.create(
                        level_id=row['level'],
                        school_id=row['school'],
                        degree_id=row['degree'],
                        period_from=row['period_from'],
                        period_to=row['period_to'],
                        units_earned=row['unit'],
                        year_graduated=row['year'],
                        hon_id=row['hon'],
                        pi_id=pi_id)
                else:
                    Educationbackground.objects.filter(id=store[x]).update(
                        level_id=row['level'],
                        school_id=row['school'],
                        degree_id=row['degree'],
                        period_from=row['period_from'],
                        period_to=row['period_to'],
                        units_earned=row['unit'],
                        year_graduated=row['year'],
                        hon_id=row['hon'],
                        pi_id=pi_id)
                    y += 1
                    x += 1
        else:
            for row in data:
                Educationbackground.objects.create(
                    level_id=row['level'],
                    school_id=row['school'],
                    degree_id=row['degree'],
                    period_from=row['period_from'],
                    period_to=row['period_to'],
                    units_earned=row['unit'],
                    year_graduated=row['year'],
                    hon_id=row['hon'],
                    pi_id=pi_id)

        check_upd_tracking = PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=3)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=3)).update(date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=3,
                pi_id=pi_id
            )

        return JsonResponse({'data': 'success'})

    context = {
        'level': Educationlevel.objects.all(),
        'hon': Honors.objects.filter((Q(hon_status=0) & Q(pi_id=pi_id)) | Q(hon_status=1)).order_by(
            'hon_name'),
        'course': Degree.objects.filter(
            (Q(deg_status=0) & Q(pi_id=pi_id)) | Q(deg_status=1)).order_by('degree_name'),
        'school': School.objects.filter(
            (Q(school_status=0) & Q(pi_id=pi_id)) | Q(school_status=1)).order_by(
            'school_name'),
        'educ': Educationbackground.objects.filter(pi_id=pi_id).order_by('level_id', 'period_to'),
        'form_school': SchoolForm(),
        'form_degree': DegreeForm(),
        'form_honor': HonorsForm(),
        'pi_id': pi_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, 'backend/employee_data/staff/pds/educational_background.html', context)


@login_required
@permission_required("auth.edit_employee")
def add_school(request, pi_id):
    if request.method == "POST":
        form = SchoolForm(request.POST)

        if form.is_valid():
            check = School.objects.filter(Q(school_name=form.cleaned_data['school_name']) & Q(pi_id=pi_id))
            if check:
                return JsonResponse({'msg': 'This school is already added.'})
            else:
                school = form.save(commit=False)
                school.school_status = 1 if request.user.id == 1 else 0
                school.pi_id = pi_id
                school.save()
                return JsonResponse({'error': False, 'school_id': school.id, 'school_name': school.school_name})
        else:
            return JsonResponse({'error': True, 'errors': form.errors})


@login_required
@permission_required("auth.edit_employee")
def add_degree(request, pi_id):
    if request.method == "POST":
        form = DegreeForm(request.POST)

        if form.is_valid():
            check = Degree.objects.filter(Q(degree_name=form.cleaned_data['degree_name']) & Q(pi_id=pi_id))
            if check:
                return JsonResponse({'msg': 'This degree is already added.'})
            else:
                degree = form.save(commit=False)
                degree.deg_status = 1 if request.user.id == 1 else 0
                degree.pi_id = pi_id
                degree.save()
                return JsonResponse({'error': False, 'degree_id': degree.id, 'degree_name': degree.degree_name})
        else:
            return JsonResponse({'error': True, 'errors': form.errors})


@login_required
@permission_required("auth.edit_employee")
def add_honor(request, pi_id):
    if request.method == "POST":
        form = HonorsForm(request.POST)

        if form.is_valid():
            check = Honors.objects.filter(Q(hon_name=form.cleaned_data['hon_name']) & Q(pi_id=pi_id))
            if check:
                return JsonResponse({'msg': 'This scholarship/academic honor is already added.'})
            else:
                hon = form.save(commit=False)
                hon.hon_status = 1 if request.user.id == 1 else 0
                hon.pi_id = pi_id
                hon.save()
                return JsonResponse({'error': False, 'hon_id': hon.id, 'hon_name': hon.hon_name})
        else:
            return JsonResponse({'error': True, 'errors': form.errors})


@login_required
@permission_required("auth.edit_employee")
def update_civil_service(request, pi_id):
    if request.method == "POST":
        el = request.POST.getlist('el[]')
        course = request.POST.getlist('course[]')
        rating = request.POST.getlist('rating[]')
        date_exam = request.POST.getlist('date_exam[]')
        place = request.POST.getlist('place[]')
        number = request.POST.getlist('number[]')
        date_val = request.POST.getlist('date_val[]')

        data = [{'el': e, 'course': c, 'rating': r, 'date_exam': de, 'place': p, 'number': n, 'date_val': dv}
                for e, c, r, de, p, n, dv in zip(el, course, rating, date_exam, place, number, date_val)]

        check = Civilservice.objects.filter(pi_id=pi_id)

        store = [row.id for row in check]

        if check:
            y = 1
            x = 0
            for row in data:
                if y > len(check):
                    Civilservice.objects.create(
                        el_id=row['el'],
                        cs_rating=row['rating'],
                        cs_dateexam=row['date_exam'],
                        cs_place=row['place'],
                        cs_number=row['number'],
                        cs_date_val=row['date_val'],
                        pi_id=pi_id,
                        course_id=row['course']
                    )
                else:
                    Civilservice.objects.filter(id=store[x]).update(
                        el_id=row['el'],
                        cs_rating=row['rating'],
                        cs_dateexam=row['date_exam'],
                        cs_place=row['place'],
                        cs_number=row['number'],
                        cs_date_val=row['date_val'],
                        pi_id=pi_id,
                        course_id=row['course']
                    )
                    y += 1
                    x += 1
        else:
            for row in data:
                Civilservice.objects.create(
                    el_id=row['el'],
                    cs_rating=row['rating'],
                    cs_dateexam=row['date_exam'],
                    cs_place=row['place'],
                    cs_number=row['number'],
                    cs_date_val=row['date_val'],
                    pi_id=pi_id,
                    course_id=row['course']
                )

        check_upd_tracking = PdsUpdateTracking.objects.filter(
            Q(pi_id=pi_id) & Q(pis_config_id=4)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=4)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=4,
                pi_id=pi_id
            )
        return JsonResponse({'data': 'success'})

    context = {
        'el': Eligibility.objects.filter(
            (Q(el_status=0) & Q(pi_id=pi_id)) | Q(el_status=1)).order_by('el_name'),
        'civil': Civilservice.objects.filter(pi_id=pi_id),
        'course': Course.objects.filter(status=1),
        'form': EligibilityForm(),
        'pi_id': pi_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, 'backend/employee_data/staff/pds/civil_service.html', context)


@login_required
@permission_required("auth.edit_employee")
def add_eligibility(request, pi_id):
    if request.method == "POST":
        form = EligibilityForm(request.POST)

        if form.is_valid():
            check = Eligibility.objects.filter(Q(el_name=form.cleaned_data['el_name']) & Q(pi_id=request.session['pi_id']))
            if check:
                return JsonResponse({'msg': 'This eligibility is already added.'})
            else:
                el = form.save(commit=False)
                el.el_status = 1 if request.user.id == 1 else 0
                el.pi_id = request.session['pi_id']
                el.save()
                return JsonResponse({'error': False, 'el_id': el.id, 'el_name': el.el_name})
        else:
            return JsonResponse({'error': True, 'errors': form.errors})


@login_required
@permission_required("auth.edit_employee")
def update_work_experience(request, pi_id):
    if request.method == "POST":
        inc_from = request.POST.getlist('inc_from[]')
        inc_to = request.POST.getlist('inc_to[]')
        position = request.POST.getlist('position[]')
        company = request.POST.getlist('company[]')
        salary = request.POST.getlist('salary[]')
        sg_step = request.POST.getlist('sg_step[]')
        app_status = request.POST.getlist('as[]')
        gs = request.POST.getlist('gs[]')

        data = [{'inc_from': ic_fr, 'inc_to': ic_to, 'position': p, 'company': c, 'salary': s, 'sg_step': sg,
                 'app_status': aps, 'gs': g}
                for ic_fr, ic_to, p, c, s, sg, aps, g in
                zip(inc_from, inc_to, position, company, salary, sg_step, app_status, gs)]

        check = Workexperience.objects.filter(pi_id=pi_id)
        store = [row.id for row in check]

        if check:
            y = 1
            x = 0
            for row in data:
                if y > len(check):
                    Workexperience.objects.create(
                        we_from=row['inc_from'],
                        we_to=row['inc_to'],
                        position_name=row['position'],
                        company=row['company'],
                        salary_rate=row['salary'].replace(',', ''),
                        sg_step=row['sg_step'],
                        empstatus_id=row['app_status'],
                        govt_service=row['gs'],
                        pi_id=pi_id)
                else:
                    Workexperience.objects.filter(id=store[x]).update(
                        we_from=row['inc_from'],
                        we_to=row['inc_to'],
                        position_name=row['position'],
                        company=row['company'],
                        salary_rate=row['salary'].replace(',', ''),
                        sg_step=row['sg_step'],
                        empstatus_id=row['app_status'],
                        govt_service=row['gs'],
                        pi_id=pi_id)

                    y += 1
                    x += 1
        else:
            for row in data:
                Workexperience.objects.create(
                    we_from=row['inc_from'],
                    we_to=row['inc_to'],
                    position_name=row['position'],
                    company=row['company'],
                    salary_rate=row['salary'].replace(',', ''),
                    sg_step=row['sg_step'],
                    empstatus_id=row['app_status'],
                    govt_service=row['gs'],
                    pi_id=pi_id)

        check_upd_tracking = PdsUpdateTracking.objects.filter(
            Q(pi_id=pi_id) & Q(pis_config_id=5)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=5)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=5,
                pi_id=pi_id
            )
        return JsonResponse({'data': 'success'})

    personal_info = Personalinfo.objects.filter(id=pi_id).first()
    context = {
        'position': Position.objects.filter((Q(status=0) & Q(upload_by_id=personal_info.user_id)) | Q(status=1)).order_by(
            'name'),
        'app_status': Empstatus.objects.filter((Q(status=0) & Q(upload_by_id=personal_info.user_id)) | Q(status=1)).order_by(
            'name'),
        'work': Workexperience.objects.filter(pi_id=pi_id).order_by('-we_to'),
        'form_pt': PositionForm(),
        'form_empstatus': EmpstatusForm(),
        'pi_id': pi_id,
        'user_id': personal_info.user_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, "backend/employee_data/staff/pds/work_experience.html", context)


@login_required
@permission_required("auth.edit_employee")
def add_position(request, user_id):
    if request.method == "POST":
        form = PositionForm(request.POST)

        if form.is_valid():
            check = Position.objects.filter(Q(name=form.cleaned_data['name']) & Q(upload_by_id=user_id))
            if check:
                return JsonResponse({'msg': 'This position title is already added.'})
            else:
                pt = form.save(commit=False)
                pt.status = 1 if user_id == 1 else 0
                pt.upload_by_id = user_id
                pt.save()
                return JsonResponse({'error': False, 'pos_id': pt.id, 'name': pt.name})
            
            
@login_required
@permission_required("auth.edit_employee")
def update_voluntary_work(request, pi_id):
    if request.method == "POST":
        org = request.POST.getlist('org[]')
        inc_from = request.POST.getlist('inc_from[]')
        inc_to = request.POST.getlist('inc_to[]')
        num_hours = request.POST.getlist('num_hours[]')
        now = request.POST.getlist('now[]')

        data = [{'org': o, 'inc_from': ic_fr, 'inc_to': ic_to, 'num_hours': nh, 'now': n}
                for o, ic_fr, ic_to, nh, n in zip(org, inc_from, inc_to, num_hours, now)]

        check = Voluntary.objects.filter(pi_id=pi_id)
        store = [row.id for row in check]

        if check:
            y = 1
            x = 0
            for row in data:
                if y > len(check):
                    Voluntary.objects.create(
                        organization=row['org'],
                        vol_from=row['inc_from'],
                        vol_to=row['inc_to'],
                        vol_hours=row['num_hours'],
                        now=row['now'],
                        pi_id=pi_id)
                else:
                    Voluntary.objects.filter(id=store[x]).update(
                        organization=row['org'],
                        vol_from=row['inc_from'],
                        vol_to=row['inc_to'],
                        vol_hours=row['num_hours'],
                        now=row['now'],
                        pi_id=pi_id)
                    y += 1
                    x += 1
        else:
            for row in data:
                Voluntary.objects.create(
                    organization=row['org'],
                    vol_from=row['inc_from'],
                    vol_to=row['inc_to'],
                    vol_hours=row['num_hours'],
                    now=row['now'],
                    pi_id=pi_id)

        check_upd_tracking = PdsUpdateTracking.objects.filter(
            Q(pi_id=pi_id) & Q(pis_config_id=6)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=6)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=6,
                pi_id=pi_id
            )

        return JsonResponse({'data': 'success'})

    context = {
        'org': Organization.objects.filter(
            (Q(org_status=0) & Q(pi_id=pi_id)) | Q(org_status=1)).order_by('org_name'),
        'voluntary': Voluntary.objects.filter(pi_id=pi_id),
        'form_org': OrganizationForm(),
        'pi_id': pi_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, "backend/employee_data/staff/pds/voluntary_work.html", context)


@login_required
@permission_required("auth.edit_employee")
def add_org(request, pi_id):
    if request.method == "POST":
        form = OrganizationForm(request.POST)

        if form.is_valid():
            check = Organization.objects.filter(Q(org_name=form.cleaned_data['org_name']) & Q(pi_id=pi_id))
            if check:
                return JsonResponse({'msg': 'This organization name is already added.'})
            else:
                org = form.save(commit=False)
                org.org_status = 1 if request.user.id == 1 else 0
                org.pi_id = pi_id
                org.save()
                return JsonResponse({'error': False, 'org_id': org.id, 'org_name': org.org_name})


@login_required
@permission_required("auth.edit_employee")
def update_trainings(request, pi_id):
    if request.method == "POST":
        training = request.POST.getlist('training[]')
        inc_from = request.POST.getlist('inc_from[]')
        inc_to = request.POST.getlist('inc_to[]')
        num_hours = request.POST.getlist('num_hours[]')
        trainingtype = request.POST.getlist('type[]')
        conducted = request.POST.getlist('conducted[]')

        data = [{'training': t, 'inc_from': ic_fr, 'inc_to': ic_to, 'num_hours': nh, 'trainingtype': tt, 'conducted': c}
                for t, ic_fr, ic_to, nh, tt, c in zip(training, inc_from, inc_to, num_hours, trainingtype, conducted)]

        check = Training.objects.filter(pi_id=pi_id)
        store = [row.id for row in check]

        if check:
            y = 1
            x = 0
            for row in data:
                if y > len(check):
                    Training.objects.create(
                        training=row['training'],
                        tr_from=row['inc_from'],
                        tr_to=row['inc_to'],
                        tr_hours=row['num_hours'],
                        training_type=row['trainingtype'],
                        conducted=row['conducted'],
                        pi_id=pi_id)
                else:
                    Training.objects.filter(id=store[x]).update(
                        training=row['training'],
                        tr_from=row['inc_from'],
                        tr_to=row['inc_to'],
                        tr_hours=row['num_hours'],
                        training_type=row['trainingtype'],
                        conducted=row['conducted'],
                        pi_id=pi_id)
                    x += 1
                    y += 1
        else:
            for row in data:
                Training.objects.create(
                    training=row['training'],
                    tr_from=row['inc_from'],
                    tr_to=row['inc_to'],
                    tr_hours=row['num_hours'],
                    training_type=row['trainingtype'],
                    conducted=row['conducted'],
                    pi_id=pi_id)

        check_upd_tracking = PdsUpdateTracking.objects.filter(
            Q(pi_id=pi_id) & Q(pis_config_id=7)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=7)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=7,
                pi_id=pi_id
            )
        return JsonResponse({'data': 'success'})

    context = {
        'trainingtype': Trainingtype.objects.all(),
        'training': Trainingtitle.objects.filter(
            (Q(tt_status=0) & Q(pi_id=pi_id)) | Q(tt_status=1)).order_by('tt_name'),
        'data': Training.objects.filter(pi_id=pi_id).order_by('-tr_to'),
        'form': TrainingtitleForm(),
        'pi_id': pi_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, "backend/employee_data/staff/pds/trainings.html", context)


@login_required
@permission_required("auth.edit_employee")
def add_trainingtitle(request, pi_id):
    if request.method == "POST":
        form = TrainingtitleForm(request.POST)

        if form.is_valid():
            check = Trainingtitle.objects.filter(Q(tt_name=form.cleaned_data['tt_name']) & Q(pi_id=pi_id))
            if check:
                return JsonResponse({'msg': 'This training title is already added.'})
            else:
                tt = form.save(commit=False)
                tt.tt_status = 1 if request.user.id == 1 else 0
                tt.pi_id = pi_id
                tt.save()
                return JsonResponse({'error': False, 'tt_id': tt.id, 'tt_name': tt.tt_name})


@login_required
@permission_required("auth.edit_employee")
def update_others(request, pi_id):
    if request.method == "POST":
        skills = request.POST.getlist('skills[]')
        nonacad = request.POST.getlist('nonacad[]')
        org = request.POST.getlist('org[]')

        check_skills = Skills.objects.filter(pi_id=pi_id)
        store_skills = [row.id for row in check_skills]
        check_nonacad = Recognition.objects.filter(pi_id=pi_id)
        store_nonacad = [row.id for row in check_nonacad]

        check_membership = Membership.objects.filter(pi_id=pi_id)
        store_membership = [row.id for row in check_membership]

        if check_skills:
            y = 1
            x = 0
            for row in skills:
                if y > len(check_skills):
                    Skills.objects.create(
                        hob_id=row,
                        pi_id=pi_id)
                else:
                    Skills.objects.filter(id=store_skills[x]).update(
                        hob_id=row,
                        pi_id=pi_id)
                    y += 1
                    x += 1
        else:
            for row in skills:
                Skills.objects.create(
                    hob_id=row,
                    pi_id=pi_id)

        if check_nonacad:
            y = 1
            x = 0
            for row in nonacad:
                if y > len(check_nonacad):
                    Recognition.objects.create(
                        na_id=row,
                        pi_id=pi_id)
                else:
                    Recognition.objects.filter(id=store_nonacad[x]).update(
                        na_id=row,
                        pi_id=pi_id)
                    y += 1
                    x += 1
        else:
            for row in nonacad:
                Recognition.objects.create(
                    na_id=row,
                    pi_id=pi_id)

        if check_membership:
            y = 1
            x = 0
            for row in org:
                if y > len(check_membership):
                    Membership.objects.create(
                        org_id=row,
                        pi_id=pi_id)
                else:
                    Membership.objects.filter(id=store_membership[x]).update(
                        org_id=row,
                        pi_id=pi_id)
                    y += 1
                    x += 1
        else:
            for row in org:
                Membership.objects.create(
                    org_id=row,
                    pi_id=pi_id)

        check_upd_tracking = PdsUpdateTracking.objects.filter(
            Q(pi_id=pi_id) & Q(pis_config_id=8)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=8)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=8,
                pi_id=pi_id
            )

        return JsonResponse({'data': 'success'})

    context = {
        'hobbies': Hobbies.objects.filter(
            (Q(pi_id=pi_id) & Q(hob_status=0)) | Q(hob_status=1)).order_by('hob_name'),
        'nonacad': Nonacad.objects.filter(
            (Q(pi_id=pi_id) & Q(na_status=0)) | Q(na_status=1)).order_by('na_name'),
        'org': Organization.objects.filter(
            (Q(pi_id=pi_id) & Q(org_status=0)) | Q(org_status=1)).order_by('org_name'),
        'skills': Skills.objects.filter(pi_id=pi_id),
        'reg': Recognition.objects.filter(pi_id=pi_id),
        'membership': Membership.objects.filter(pi_id=pi_id),
        'form_hob': HobbiesForm(),
        'form_na': NonacadForm(),
        'form_org': OrganizationForm(),
        'pi_id': pi_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, "backend/employee_data/staff/pds/others.html", context)


@login_required
@permission_required("auth.edit_employee")
def add_hobbies(request, pi_id):
    if request.method == "POST":
        form = HobbiesForm(request.POST)

        if form.is_valid():
            check = Hobbies.objects.filter(Q(hob_name=form.cleaned_data['hob_name']) & Q(pi_id=pi_id))
            if check:
                return JsonResponse({'msg': 'This skill/hobby is already added.'})
            else:
                hob = form.save(commit=False)
                hob.hob_status = 1 if request.user.id == 1 else 0
                hob.pi_id = pi_id
                hob.save()
                return JsonResponse({'error': False, 'hob_id': hob.id, 'hob_name': hob.hob_name})


@login_required
@permission_required("auth.edit_employee")
def add_recog(request, pi_id):
    if request.method == "POST":
        form = NonacadForm(request.POST)

        if form.is_valid():
            check = Nonacad.objects.filter(Q(na_name=form.cleaned_data['na_name']) & Q(pi_id=pi_id))
            if check:
                return JsonResponse({'msg': 'This non-academic/recognition is already added.'})
            else:
                na = form.save(commit=False)
                na.na_status = 1 if request.user.id == 1 else 0
                na.pi_id = pi_id
                na.save()
                return JsonResponse({'error': False, 'na_id': na.id, 'na_name': na.na_name})


@login_required
@permission_required("auth.edit_employee")
def update_additional_information(request, pi_id):
    if request.method == "POST":
        answers = [request.POST.get('q1'), request.POST.get('q2'), request.POST.get('q3'), request.POST.get('q4'),
                  request.POST.get('q4'), request.POST.get('q5'), request.POST.get('q6'), request.POST.get('q7'),
                  request.POST.get('q8'), request.POST.get('q9'), request.POST.get('q10'), request.POST.get('q11'),
                  request.POST.get('q12')]

        answer_detail = ['', request.POST.get('answer1'), request.POST.get('answer2'), request.POST.get('answer3'), request.POST.get('answer4'),
                  request.POST.get('answer5'), request.POST.get('answer6'), request.POST.get('answer7'), request.POST.get('answer8'),
                  request.POST.get('answer9'), request.POST.get('answer10'), request.POST.get('answer11'), request.POST.get('answer12')]

        check = Additional.objects.filter(pi_id=pi_id)
        if check:
            store = [row.id for row in check]
            count = 0
            for row in store:
                Additional.objects.filter(id=row).update(
                    question=answers[count],
                    answers=answer_detail[count],
                    ad_status=1,
                    pi_id=pi_id)
                count = count + 1
        else:
            count = 0
            for row in answers:
                Additional.objects.create(
                    question=answers[count],
                    answers=answer_detail[count],
                    ad_status=1,
                    pi_id=pi_id)
                count = count + 1

        check_upd_tracking = PdsUpdateTracking.objects.filter(
            Q(pi_id=pi_id) & Q(pis_config_id=9)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=9)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=9,
                pi_id=pi_id
            )

        return JsonResponse({'data': 'success'})

    additional = Additional.objects.filter(pi_id=pi_id)
    data = [dict(question=row.question, answers=row.answers) for row in additional]

    context = {
        'additional': data,
        'pi_id': pi_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, "backend/employee_data/staff/pds/additional.html", context)


@login_required
@permission_required("auth.edit_employee")
def update_reference(request, pi_id):
    if request.method == "POST":
        name = request.POST.getlist('name[]')
        address = request.POST.getlist('address[]')
        tel_no = request.POST.getlist('tel_no[]')

        data = [{'name': n, 'address': a, 'tel_no': tn}
                for n, a, tn in zip(name, address, tel_no)]

        check = Reference.objects.filter(pi_id=pi_id)
        store = [row.id for row in check]

        if check:
            y = 1
            x = 0
            for row in data:
                if y > len(check):
                    Reference.objects.create(
                        name=row['name'],
                        address=row['address'],
                        tel_no=row['tel_no'],
                        pi_id=pi_id)
                else:
                    Reference.objects.filter(id=store[x]).update(
                        name=row['name'],
                        address=row['address'],
                        tel_no=row['tel_no'],
                        pi_id=pi_id)
                    y += 1
                    x += 1
        else:
            for row in data:
                Reference.objects.create(
                    name=row['name'],
                    address=row['address'],
                    tel_no=row['tel_no'],
                    pi_id=pi_id)

        check_upd_tracking = PdsUpdateTracking.objects.filter(
            Q(pi_id=pi_id) & Q(pis_config_id=10)).first()
        if check_upd_tracking:
            PdsUpdateTracking.objects.filter(Q(pi_id=pi_id) & Q(pis_config_id=10)).update(
                date=datetime.now())
        else:
            PdsUpdateTracking.objects.create(
                pis_config_id=10,
                pi_id=pi_id
            )

        return JsonResponse({'data': 'success'})
    context = {
        'reference': Reference.objects.filter(pi_id=pi_id),
        'pi_id': pi_id,
        'shortcuts': PisConfig.objects.all()
    }
    return render(request, "backend/employee_data/staff/pds/reference.html", context)
